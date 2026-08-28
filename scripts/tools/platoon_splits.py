#!/usr/bin/env python3
"""platoon_splits.py — batter/pitcher handedness splits for a target player list.

Produces, for every player in TARGETS:
  hitters  : one row per split (All / vs RHP / vs LHP) per level
  pitchers : one row per split (All / vs RHH / vs LHH) per level
  arsenal  : one row per (pitch type x split) per level

Every metric that appears on the Huronalytics player cards and can be computed
from a pitch subset is included, plus a percentile rank against the MLB pool
evaluated on the SAME split.

Design decisions (see the README block the writer emits):

  1. League anchors stay FULL SEASON, MLB. Only the player's own pitch subset
     is split. So "SD+ vs LHP = 112" means 112 against the overall MLB average
     hitter, which makes vs-R and vs-L directly comparable to each other and to
     the season card. Splitting the baseline too would put each column in its
     own currency.

  2. Percentile pool = MLB players who qualify under the site's own rule
     (3.1 PA x team games for hitters; 1.0 IP x TG for SP / 0.5 for RP),
     evaluated on the same split. Pool size is reported per metric.

  3. ROC / AAA players are ranked RAW against the MLB distribution with no
     level translation, exactly as the site and the cards do.

Sources
  data/all_pitches_rs_cache.pkl  — retagged league pitch data (the site's source)
  data/metadata_rs.json          — shipped league tables / constants
  NLE2026 'NEW' tab              — the three players new to the org, whose data
                                   is NOT in the pipeline cache

Usage:  python3 scripts/tools/platoon_splits.py [--outdir ~/Downloads] [--no-new-tab]
"""

import argparse
import json
import math
import os
import pickle
import sys
from collections import defaultdict

REPO = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.insert(0, REPO)

from pipeline.utils import (                                    # noqa: E402
    safe_float, spray_angle, spray_direction, compute_in_zone,
    runexp_scale_from_json, runexp_factor,
    hitter_pa_per_game, pitcher_ip_per_game, SP_GS_RATIO, ip_str_to_float,
    BUNT_BB_TYPES, MLB_TEAMS, AAA_TEAMS,
)
from pipeline.compute import (                                  # noqa: E402
    compute_stats, compute_expected_stats, compute_hitter_stats,
    compute_pitcher_batted_ball, compute_xrv, build_bip_count_means,
)
import pipeline.sdplus as SD                                    # noqa: E402
import pipeline.contact as CT                                   # noqa: E402
from pipeline.xwoba3d import classify_bip as xw3d_classify      # noqa: E402

DATA_DIR = os.path.join(REPO, 'data')

# ── Target list ──────────────────────────────────────────────────────────
# (display name, sheet name, role, tab)
TARGETS = [
    ('Jackson Kent',    'Kent, Jackson',    'P', 'ROC'),
    ('Luis Perales',    'Perales, Luis',    'P', 'ROC'),
    ('Erik Tolman',     'Tolman, Erik',     'P', 'ROC'),
    ('Jack Sinclair',   'Sinclair, Jack',   'P', 'ROC'),
    ('Abimelec Ortiz',  'Ortiz, Abimelec',  'H', 'AAA'),
    ('Yohandy Morales', 'Morales, Yohandy', 'H', 'AAA'),
    ('Seaver King',     'King, Seaver',     'H', 'AAA'),
    ('Andrew Pinckney', 'Pinckney, Andrew', 'H', 'AAA'),
    ('Phillip Glasser', 'Glasser, Phillip', 'H', 'AAA'),
    ('Yovanny Cruz',    'Cruz, Yovanny',    'P', 'NEW'),
    ('Jake Bird',       'Bird, Jake',       'P', 'NEW'),
    ('Will Dion',       'Dion, Will',       'P', 'NEW'),
]

# MLB team of a NEW-tab player's 2026 major league stint (pickle lookup key).
# A NEW-tab player can have MULTIPLE MLB stints (all three were called up by
# Washington in August 2026), so each maps to the list of cache PTeam keys.
NEW_TAB_MLB_TEAM = {
    'Cruz, Yovanny': ['NYY', 'WSH'],
    'Bird, Jake':    ['NYY', 'WSH'],
    'Dion, Will':    ['CLE', 'WSH'],
}

MIN_PITCHES_FOR_ROW = 75       # display floor for a level (~20 PA)
MIN_LEVEL_PA = 40              # a level needs this many PA/TBF to get rows
MIN_SPLIT_PA = 20              # a hand split needs this many PA/TBF

# --days N scales the floors down: a 30-day window holds roughly a quarter of
# a season's PAs, so season floors would empty the vs-LHP column entirely.
# These are conventions chosen to keep the window's columns populated, not
# measured stabilization points — the rows are form, not skill estimates.
WINDOW_FLOOR_SCALE = 0.5
DATE_FROM = None               # set by --days; split_pitches honors it
MIN_ARSENAL_PITCHES = 25       # a pitch type x split cell needs this many
MIN_GROUP_PITCHES = 40         # a hitter pitch-group x split cell needs this many
POOL_MIN_PA  = 25              # split-sample floor for the hitter pool
POOL_MIN_TBF = 25              # split-sample floor for the pitcher pool

# 2026 FanGraphs Guts wOBA weights. FanGraphs sits behind a Cloudflare
# challenge from here, so when the live scrape 403s these are used instead.
# They are RECOVERED from the shipped data, and pinned rather than estimated:
# least squares of the pitcher leaderboard's wOBA on each pitcher's per-PA
# event rates gives [.6974 .7292 .8894 1.2580 1.5939 2.0452], and an exhaustive
# sweep of all 531,441 three-decimal sets within +/-.004 of it leaves EXACTLY
# ONE that reproduces all 589 shipped pitcher wOBA values to their rounding
# (max |error| < .0005). That set is the one below, so this is the 2026 Guts
# row itself, not an approximation of it. Re-derive with
# scripts/derive_guts_weights.py if a season rolls over.
# (The 2025 set — .689/.720/.882/1.254/1.590/2.048 — missed every row by ~.002.)
WOBA_WEIGHTS_FALLBACK = {'BB': 0.698, 'HBP': 0.729, '1B': 0.889,
                         '2B': 1.259, '3B': 1.593, 'HR': 2.044}

# The matching 2026 Guts league constants, read from a metadata_rs.json that
# was generated on a run where the FanGraphs scrape SUCCEEDED. A run where it
# fails writes process_data's 2025 fallbacks instead, and metadata carries no
# marker saying which happened — so the fallback values are pinned here and
# detected by equality below. Without that check this script would silently
# inherit a degraded lgWOBA/wOBAScale into every xRV, SD+ and CT+ number.
GUTS_2026 = {'lgWOBA': 0.3164, 'wOBAScale': 1.2342, 'lgRPA': 0.1188}
GUTS_2025_FALLBACK = {'lgWOBA': 0.317, 'wOBAScale': 1.25, 'lgRPA': 0.119}

FB_TYPES = {'FF', 'SI'}
CF_TO_FC_PITCHERS = {'Ashcraft, Graham', 'Doval, Camilo', 'Fluharty, Mason',
                     'Funderburk, Kody', 'Jansen, Kenley', 'Maton, Phil'}

# LA bins for the SACQ / xwOBAsp lookup (mirrors process_data.LA_BINS).
LA_BINS = [(-9999, -10), (-10, 0), (0, 5), (5, 10), (10, 15), (15, 20),
           (20, 25), (25, 30), (30, 35), (35, 40), (40, 50), (50, 9999)]


def la_bin_idx(la):
    for i, (lo, hi) in enumerate(LA_BINS):
        if lo <= la < hi:
            return i
    return None


def mean(vals):
    vals = [v for v in vals if v is not None]
    return sum(vals) / len(vals) if vals else None


# ═════════════════════════════════════════════════════════════════════════
#  LOAD + PREPROCESS
# ═════════════════════════════════════════════════════════════════════════

def load_pitches(use_new_tab=True):
    with open(os.path.join(DATA_DIR, 'all_pitches_rs_cache.pkl'), 'rb') as f:
        pitches = pickle.load(f)
    print(f"  Loaded {len(pitches)} pitches from the pipeline cache")

    new_rows = []
    if use_new_tab:
        import pipeline.fetch as pf
        gc = pf._gspread_client()
        raw = pf.read_pitches_from_sheet(
            gc, pf.DIVISION_WORKBOOK_IDS['NLE2026'], extra_tabs={'NEW'})
        raw = [r for r in raw if r.get('_sheet_tab') == 'NEW'
               and r.get('Pitcher') in NEW_TAB_MLB_TEAM]
        # The NEW tab is a superset: it re-lists the player's MLB pitches. Those
        # already live in the cache (with Stuff+/Loc+ grades attached), so keep
        # only the games the cache does not have — the minor league side.
        mlb_gamepks = {(p.get('PitchID') or '').split('_')[0]
                       for p in pitches if p.get('_sheet_tab') in MLB_TEAMS}
        for r in raw:
            gpk = (r.get('PitchID') or '').split('_')[0]
            if gpk in mlb_gamepks:
                continue
            # NEW-tab BTeam labelling is scratch-quality and PTeam is always
            # WSH; retag as an AAA-source pitch so the RunExp currency fix and
            # the xwOBA fill treat it like any other minor league pitch.
            # Own _source tag so the RunExp scale is derived for these rows
            # specifically (the cached ROC/AAA rows are already corrected;
            # these, read straight from the Sheet, are not).
            r['_source'] = 'NEW'
            new_rows.append(r)
        print(f"  Loaded {len(new_rows)} minor league pitches from the NEW tab")
    return pitches, new_rows


def preprocess(all_pitches, new_rows, metadata, woba_weights):
    """The CI-released pickle is ALREADY through process_game_type's
    preprocessing (InZone recomputed, CF reclassified, ROC/AAA tagged, MiLB
    RunExp rescaled to MLB currency, minor league xwOBA filled). A pickle from
    refresh_pickle.py (Sheets -> pickle) is NOT: it has InZone and ROC tags but
    no CF remap, no RunExp rescale, no xwOBA fill. Every correction here is
    therefore either self-detecting (the RunExp scale measures 1.000 on
    already-corrected rows) or idempotent (CF remap, xwOBA fill only where
    None), so both pickle provenances come out identical."""
    ep_pitchers = {(p['Pitcher'], p['PTeam']) for p in all_pitches
                   if p.get('Pitch Type') == 'EP'}

    # CF -> FF/FC remap over EVERYTHING (matches process_data; no-op on a
    # CI pickle where it already ran).
    n_cf = 0
    for p in all_pitches:
        if p.get('Pitch Type') == 'CF':
            p['Pitch Type'] = 'FC' if p.get('Pitcher') in CF_TO_FC_PITCHERS else 'FF'
            n_cf += 1
    if n_cf:
        print(f"  CF remapped on {n_cf} pitches (Sheets-fresh pickle)")

    for p in new_rows:
        p['InZone'] = compute_in_zone(p)
        # Pitcher side is the tracked player; every batter here is a minor
        # league opponent who must stay out of hitter aggregations.
        p['_roc_pitcher_pitch'] = True

    # Statcast denominates delta_run_exp in each league's own run-expectancy
    # matrix. Derive the factor from the data itself rather than trusting a
    # stored constant: rows already in MLB currency come back at 1.000 and the
    # rescale becomes a no-op, so this can never double-correct.
    from pipeline.utils import compute_runexp_scale
    scale = compute_runexp_scale(all_pitches)
    for src, sc in sorted(scale.items()):
        print(f"  RunExp scale {src}: global {sc['global']:.4f} "
              f"({len(sc['cell'])} cell factors)")
    n_fixed = 0
    for p in all_pitches:
        sc = scale.get(p.get('_source'))
        if not sc:
            continue
        v = safe_float(p.get('RunExp'))
        if v is None:
            continue
        f = runexp_factor(sc, p.get('Description'), p.get('Count'))
        if f and abs(f - 1.0) > 1e-6:
            p['RunExp'] = v / f
            n_fixed += 1
    print(f"  RunExp rescaled on {n_fixed} pitches")

    # xwOBA fill for the NEW-tab minor league rows, from the shipped 3D
    # EV x LA x spray x bats table (Savant publishes no per-pitch xwOBA there).
    table = {}
    for k, v in (metadata.get('xwOBA3DTable') or {}).items():
        ev, la, sp, bats = k.split('|')
        table[(int(ev[2:]), int(la[2:]), sp, bats)] = v['rv']
    bb_w, hbp_w = woba_weights.get('BB', 0.69), woba_weights.get('HBP', 0.72)
    from pipeline.utils import BB_EVENTS, HBP_EVENTS, K_EVENTS
    n_fill = 0
    for p in all_pitches:
        if p.get('_source', 'MLB') == 'MLB' or p.get('xwOBA') is not None:
            continue
        ev = p.get('Event')
        if not ev:
            continue
        key = xw3d_classify(p)
        if key is not None and key in table:
            p['xwOBA'] = round(table[key], 4); n_fill += 1
        elif ev in BB_EVENTS and ev != 'Intent Walk':
            p['xwOBA'] = bb_w; n_fill += 1
        elif ev in HBP_EVENTS:
            p['xwOBA'] = hbp_w; n_fill += 1
        elif ev in K_EVENTS:
            p['xwOBA'] = 0.0; n_fill += 1
    print(f"  xwOBA filled on {n_fill} PA-ending pitches")

    return ep_pitchers


# ═════════════════════════════════════════════════════════════════════════
#  LEAGUE TABLES (built once, full season, MLB-baselined)
# ═════════════════════════════════════════════════════════════════════════

class League:
    def __init__(self, all_pitches, ep_pitchers, metadata, woba_weights):
        self.woba_weights = woba_weights
        guts = dict(metadata['gutsConstants'])
        if all(abs(guts.get(k, 0) - v) < 1e-9 for k, v in GUTS_2025_FALLBACK.items()):
            print("  *** metadata gutsConstants are process_data's 2025 "
                  "FALLBACKS — that run's FanGraphs scrape failed. Using the "
                  "pinned 2026 constants instead. ***")
            guts = dict(GUTS_2026)
        self.lg_woba = guts['lgWOBA']
        self.woba_scale = guts['wOBAScale']
        self.lg_rpa = guts.get('lgRPA')     # wRC/wRC+ denominator
        self.siera_constant = metadata['sieraConstant']
        self.lg_xwobacon = metadata['hitterLeagueAverages']['xwOBAcon']
        self.reanchor = metadata['plusReanchor']
        self.hstd = metadata['hitterPlusStandardization']
        self.team_games = metadata['teamGamesPlayed']
        self.max_tg = max(self.team_games.values())

        mlb_for_xrv = [p for p in all_pitches
                       if p.get('_source', 'MLB') == 'MLB'
                       and (p.get('Pitcher'), p.get('PTeam')) not in ep_pitchers]
        self.xrv_offsets = SD.build_bip_count_offsets(
            mlb_for_xrv, self.lg_woba, self.woba_scale)
        self.xrv_bip_means = build_bip_count_means(
            mlb_for_xrv, self.lg_woba, self.woba_scale, self.xrv_offsets)

        # SACQ zone lookup (spray x LA x bats), hand table with pooled fallback.
        self.sacq_hand, self.sacq_pool = {}, {}
        for z in metadata.get('sacqZones', []):
            key = (z['spray'], z['laBin'])
            if z.get('bats') in ('L', 'R'):
                self.sacq_hand[(z['bats'],) + key] = z
            elif z.get('bats') is None:
                self.sacq_pool[key] = z

        # SD+ / CT+ cell tables — rebuilt from the same MLB-only inputs the
        # pipeline uses, so the cells are identical to the shipped ones.
        print("  Building SD+ cell table...")
        sd_elig = [p for p in all_pitches
                   if p.get('_source', 'MLB') == 'MLB' and SD.is_eligible(p)
                   and (p.get('Pitcher'), p.get('PTeam')) not in ep_pitchers]
        sd_off = SD.build_bip_count_offsets(sd_elig, self.lg_woba, self.woba_scale)
        sd_rv = SD.make_rv_xrv(self.lg_woba, self.woba_scale, sd_off)
        self.sd_table = SD.shrink_table(SD.build_weight_table(sd_elig, sd_rv),
                                        SD.zone_level_means(sd_elig, sd_rv))
        zc = defaultdict(int)
        for p in sd_elig:
            zc[SD.classify_zone(p)] += 1
        tot = sum(zc.values())
        self.sd_zone_w = {z: n / tot for z, n in zc.items()} if tot else None

        print("  Building CT+ cell table...")
        ct_sw = [p for p in all_pitches
                 if p.get('_source', 'MLB') == 'MLB' and CT.is_ct_eligible(p)
                 and (p.get('Pitcher'), p.get('PTeam')) not in ep_pitchers]
        ct_off = SD.build_bip_count_offsets(ct_sw, self.lg_woba, self.woba_scale)
        ct_rv = SD.make_rv_xrv(self.lg_woba, self.woba_scale, ct_off)
        self.ct_table = CT.shrink_contact_cells(
            CT.build_contact_cell_weights(ct_sw, ct_rv),
            CT.zone_level_contact_means(ct_sw, ct_rv))

        # League anchors for SD+/CT+ come from the FULL-SEASON MLB hitter pool,
        # so a split row is scored against the same yardstick as a season row.
        self.sd_lg_raw = self.sd_lg_mean = None
        self.ct_lg_raw = self.ct_lg_mean = None

    def anchor_from_full_season(self, hitter_groups_full):
        """Derive the league anchors (lg_raw, lg_mean) exactly as
        pipeline_sdplus/contact.regress_and_normalize does, from full-season
        MLB hitter groups. Split rows then reuse these anchors."""
        raw = SD.compute_hitter_sd(hitter_groups_full, self.sd_table, self.sd_zone_w)
        elig = {k: v for k, v in raw.items()
                if v['n_decisions'] >= SD.MIN_HITTER_DECISIONS}
        self.sd_lg_raw = sum(v['raw_sd'] for v in elig.values()) / len(elig)
        adj = [(v['n_decisions'] * v['raw_sd'] + SD.HITTER_PRIOR_N * self.sd_lg_raw)
               / (v['n_decisions'] + SD.HITTER_PRIOR_N) for v in elig.values()]
        self.sd_lg_mean = sum(adj) / len(adj)

        raw = CT.compute_hitter_ct(hitter_groups_full, self.ct_table)
        elig = {k: v for k, v in raw.items()
                if v['n_swings'] >= CT.MIN_HITTER_SWINGS}
        self.ct_lg_raw = sum(v['raw_ct'] for v in elig.values()) / len(elig)
        adj = [(v['n_swings'] * v['raw_ct'] + CT.HITTER_PRIOR_N * self.ct_lg_raw)
               / (v['n_swings'] + CT.HITTER_PRIOR_N) for v in elig.values()]
        self.ct_lg_mean = sum(adj) / len(adj)
        print(f"  SD+ anchor lg_raw={self.sd_lg_raw:.5f} lg_mean={self.sd_lg_mean:.5f}; "
              f"CT+ anchor lg_raw={self.ct_lg_raw:.5f} lg_mean={self.ct_lg_mean:.5f}")

    # ── per-subset scorers ───────────────────────────────────────────────
    def sacq_lookup(self, direction, lb, bats):
        z = self.sacq_hand.get((bats, direction, lb))
        if z and z.get('count', 0) >= 20:
            v = z.get('wobacon', z.get('woba'))
            if v is not None:
                return v
        z = self.sacq_pool.get((direction, lb))
        if z and z.get('count', 0) >= 20:
            v = z.get('wobacon', z.get('woba'))
            if v is not None:
                return v
        return None

    def xwobasp(self, pitches):
        vals = []
        for p in pitches:
            bb = p.get('BBType')
            if not bb or bb in BUNT_BB_TYPES:
                continue
            hx, hy = safe_float(p.get('HC_X')), safe_float(p.get('HC_Y'))
            la, bats = safe_float(p.get('LaunchAngle')), p.get('Bats')
            if la is None or hx is None or hy is None or not bats:
                continue
            d = spray_direction(spray_angle(hx, hy), bats)
            lb = la_bin_idx(la)
            if not d or lb is None:
                continue
            v = self.sacq_lookup(d, lb, bats)
            if v is not None:
                vals.append(v)
        return round(sum(vals) / len(vals), 3) if vals else None

    def sd_plus(self, pitches):
        raw = SD.compute_hitter_sd({('x', 'x'): pitches}, self.sd_table,
                                   self.sd_zone_w).get(('x', 'x'))
        if not raw:
            return None, 0
        n = raw['n_decisions']
        adj = ((n * raw['raw_sd'] + SD.HITTER_PRIOR_N * self.sd_lg_raw)
               / (n + SD.HITTER_PRIOR_N))
        return round(100.0 * adj / self.sd_lg_mean * self.reanchor['sdPlus'], 1), n

    def ct_plus(self, pitches):
        raw = CT.compute_hitter_ct({('x', 'x'): pitches}, self.ct_table).get(('x', 'x'))
        if not raw:
            return None, 0
        n = raw['n_swings']
        adj = ((n * raw['raw_ct'] + CT.HITTER_PRIOR_N * self.ct_lg_raw)
               / (n + CT.HITTER_PRIOR_N))
        return round(100.0 * adj / self.ct_lg_mean * self.reanchor['ctPlus'], 1), n

    def bb_plus(self, xwobacon, n_bip):
        if xwobacon is None or n_bip < 30:
            return None
        raw = 100.0 * xwobacon / self.lg_xwobacon
        v = (n_bip * raw + 60 * 100.0) / (n_bip + 60)
        return round(v * self.reanchor['bbPlus'], 1)

    def hitter_plus(self, bbp, sdp, ctp):
        """Same three steps the pipeline applies, in order: composite z on the
        standardized components, the all-MLB re-anchor shift, then the wRC+
        spread match. The z is invariant to the component re-anchor (numerator
        and denominator scale together), so post-re-anchor inputs are correct."""
        if bbp is None or sdp is None or ctp is None:
            return None
        s = self.hstd
        z = (s['weights']['bb'] * (bbp - s['bbPlus']['mean']) / s['bbPlus']['sd']
             + s['weights']['sd'] * (sdp - s['sdPlus']['mean']) / s['sdPlus']['sd']
             + s['weights']['ct'] * (ctp - s['ctPlus']['mean']) / s['ctPlus']['sd'])
        hp = s['scale'] * z + self.reanchor['hitterPlusShift']
        wsm = s.get('wrcScaleMatch')
        if wsm and wsm.get('factor'):
            hp *= wsm['factor']
        return round(100 + hp, 1)

    def xrv100(self, pitches, negate):
        n = len(pitches)
        if not n:
            return None
        x = compute_xrv(pitches, lg_woba=self.lg_woba, woba_scale=self.woba_scale,
                        count_offsets=self.xrv_offsets,
                        bip_count_means=self.xrv_bip_means, negate=negate)
        xv = x.get('xRunValue')
        return xv / n * 100 if xv is not None else None


# ═════════════════════════════════════════════════════════════════════════
#  ROW BUILDERS
# ═════════════════════════════════════════════════════════════════════════

def grade_atom_mean(pitches, key):
    vals = [int(float(p[key])) for p in pitches
            if p.get(key) not in (None, '') and str(p[key]).strip() != '']
    return (sum(vals) / len(vals)) if vals else None


def slash_from_pitches(pitches):
    """Slash line / K% / BB% from pitch events.

    The leaderboard takes these from the official boxscore (which catches
    no-pitch IBBs), but a boxscore has no handedness split, so a split row has
    to be pitch-derived. Expect a hair of drift from the season card: Ortiz is
    364 pitch-derived PA against 365 official."""
    from pipeline.utils import (HIT_EVENTS, K_EVENTS, BB_EVENTS, HBP_EVENTS,
                                SF_EVENTS, SH_EVENTS, CI_EVENTS, NON_PA_EVENTS)
    pa = [p for p in pitches if p.get('Event') and p['Event'] not in NON_PA_EVENTS]
    n_pa = len(pa)
    if not n_pa:
        return {}
    ev = [p['Event'] for p in pa]
    h = sum(1 for e in ev if e in HIT_EVENTS)
    hr = ev.count('Home Run'); d2 = ev.count('Double'); d3 = ev.count('Triple')
    d1 = max(0, h - hr - d2 - d3)
    bb = sum(1 for e in ev if e in BB_EVENTS)
    hbp = sum(1 for e in ev if e in HBP_EVENTS)
    sf = sum(1 for e in ev if e in SF_EVENTS)
    sh = sum(1 for e in ev if e in SH_EVENTS)
    ci = sum(1 for e in ev if e in CI_EVENTS)
    k = sum(1 for e in ev if e in K_EVENTS)
    ab = n_pa - bb - hbp - sf - sh - ci
    tb = d1 + 2 * d2 + 3 * d3 + 4 * hr
    obp_den = ab + bb + hbp + sf
    out = {'hr': hr, 'doubles': d2, 'triples': d3,
           'avg': round(h / ab, 3) if ab else None,
           'obp': round((h + bb + hbp) / obp_den, 3) if obp_den else None,
           'slg': round(tb / ab, 3) if ab else None,
           'kPct': k / n_pa, 'bbPct': bb / n_pa,
           'bbToK': (bb / k) if k else None}
    if out['obp'] is not None and out['slg'] is not None:
        out['ops'] = round(out['obp'] + out['slg'], 3)
        out['iso'] = round(out['slg'] - out['avg'], 3)
    return out


def hitter_row(pitches, skill_pitches, lg):
    """skill_pitches = pitches minus position-player pitching. The pipeline
    scores xwOBAcon / xwOBAsp / SD+ / CT+ on that subset (a catcher's 48 mph
    eephus is not evidence about a hitter), while the slash line and wOBA keep
    every PA."""
    row = {}
    row.update(compute_hitter_stats(pitches))
    row.update(compute_expected_stats(pitches, woba_weights=lg.woba_weights))
    row.update(slash_from_pitches(pitches))
    row['nPitches'] = len(pitches)
    if len(skill_pitches) != len(pitches):
        row['xwOBAcon'] = compute_expected_stats(
            skill_pitches, woba_weights=lg.woba_weights).get('xwOBAcon')
    row['xwOBAsp'] = lg.xwobasp(skill_pitches)
    row['xRv100'] = lg.xrv100(pitches, negate=True)
    row['sdPlus'], row['sdPlusN'] = lg.sd_plus(skill_pitches)
    row['ctPlus'], row['ctPlusN'] = lg.ct_plus(skill_pitches)
    row['bbPlus'] = lg.bb_plus(row.get('xwOBAcon'), row.get('nBip') or 0)
    row['hitterPlus'] = lg.hitter_plus(row.get('bbPlus'), row.get('sdPlus'),
                                       row.get('ctPlus'))
    return row


def pitcher_row(pitches, lg, gs_ratio):
    row = {}
    row.update(compute_stats(pitches))
    row.update(compute_pitcher_batted_ball(pitches))
    row.update(compute_expected_stats(pitches, woba_weights=lg.woba_weights))
    n = len(pitches)
    row['nPitches'] = n
    row['tbf'] = row.get('pa') or 0
    row['xRv100'] = lg.xrv100(pitches, negate=False)
    rv = row.get('runValue')
    row['rv100'] = (rv / n * 100) if (rv is not None and n) else None
    row['stuffPlus'] = grade_atom_mean(pitches, 'Stuff+')
    row['locPlus'] = grade_atom_mean(pitches, 'Loc+')

    by_type = defaultdict(list)
    for p in pitches:
        if p.get('Pitch Type') in FB_TYPES:
            v = safe_float(p.get('Velocity'))
            if v is not None:
                by_type[p['Pitch Type']].append(v)
    if by_type:
        primary = max(by_type, key=lambda t: len(by_type[t]))
        row['fbVelo'] = round(mean(by_type[primary]), 1)
    else:
        row['fbVelo'] = None

    row['siera'] = split_siera(row, lg, gs_ratio)
    return row


def split_siera(row, lg, gs_ratio):
    """SIERA on the split's own K/BB/GB/FB rates, with the shipped constant.
    Uses pitch-derived events rather than the boxscore, so it will differ by a
    hair from the season SIERA on the card."""
    tbf = row.get('tbf') or 0
    n_bip = row.get('nBip') or 0
    if tbf < 1 or n_bip < 1:
        return None
    so_pa = row.get('kPct')
    bb_pa = row.get('bbPct')
    if so_pa is None or bb_pa is None:
        return None
    gb = round((row.get('gbPct') or 0) * n_bip)
    fb = round(((row.get('fbPct') or 0) + (row.get('puPct') or 0)) * n_bip)
    net_gb_pa = (gb - fb) / tbf
    sign = -1.0 if gb >= fb else 1.0
    raw = (- 15.518 * so_pa + 9.146 * so_pa ** 2
           + 8.648 * bb_pa + 27.252 * bb_pa ** 2
           - 2.298 * net_gb_pa + sign * 4.920 * net_gb_pa ** 2
           - 4.036 * so_pa * bb_pa + 5.155 * so_pa * net_gb_pa
           + 4.546 * bb_pa * net_gb_pa + 0.367 * gs_ratio)
    return round(raw + lg.siera_constant, 2)


def arsenal_rows(pitches, lg, total_pitches):
    out = {}
    by_type = defaultdict(list)
    for p in pitches:
        pt = p.get('Pitch Type')
        if pt:
            by_type[pt].append(p)
    for pt, ps in by_type.items():
        n = len(ps)
        st = compute_stats(ps)
        bb = compute_pitcher_batted_ball(ps)
        ex = compute_expected_stats(ps, woba_weights=lg.woba_weights)
        rv = st.get('runValue')
        out[pt] = {
            'count': n,
            'usagePct': n / total_pitches if total_pitches else None,
            'velocity': mean([safe_float(p.get('Velocity')) for p in ps]),
            'maxVelo': max([v for v in (safe_float(p.get('Velocity')) for p in ps)
                            if v is not None], default=None),
            'spinRate': mean([safe_float(p.get('Spin Rate')) for p in ps]),
            'indVertBrk': mean([safe_float(p.get('xIndVrtBrk')) for p in ps]),
            'horzBrk': mean([safe_float(p.get('xHorzBrk')) for p in ps]),
            'extension': mean([safe_float(p.get('Extension')) for p in ps]),
            'armAngle': mean([safe_float(p.get('ArmAngle')) for p in ps]),
            'stuffPlus': grade_atom_mean(ps, 'Stuff+'),
            'locPlus': grade_atom_mean(ps, 'Loc+'),
            'izPct': st.get('izPct'),
            'swStrPct': st.get('swStrPct'),
            'izWhiffPct': st.get('izWhiffPct'),
            'chasePct': st.get('chasePct'),
            'cswPct': st.get('cswPct'),
            'gbPct': st.get('gbPct'),
            'hardHitPct': bb.get('hardHitPct'),
            'barrelPctAgainst': bb.get('barrelPctAgainst'),
            'xwOBAcon': ex.get('xwOBAcon'),
            'nBip': st.get('nBip'),
            'rv100': (rv / n * 100) if rv is not None else None,
            'xRv100': lg.xrv100(ps, negate=False),
        }
    return out


# ═════════════════════════════════════════════════════════════════════════
#  PERCENTILES
# ═════════════════════════════════════════════════════════════════════════

def percentile(value, pool, higher_is_better):
    """Site convention: (below + 0.5*equal) / n, inverted for lower-is-better."""
    if value is None or not pool:
        return None
    import bisect
    s = sorted(pool)
    below = bisect.bisect_left(s, value)
    equal = bisect.bisect_right(s, value) - below
    p = (below + 0.5 * equal) / len(s) * 100
    if not higher_is_better:
        p = 100 - p
    return max(0, min(100, round(p)))


# ── Metric display config: (key, label, decimals or 'pct', higher_is_better)
HITTER_METRICS = [
    ('pa',              'PA',           0,     None),
    ('nBip',            'BIP',          0,     None),
    ('avg',             'AVG',          'avg', True),
    ('obp',             'OBP',          'avg', True),
    ('slg',             'SLG',          'avg', True),
    ('ops',             'OPS',          'avg', True),
    ('iso',             'ISO',          'avg', True),
    ('wOBA',            'wOBA',         'avg', True),
    ('xwOBA',           'xwOBA',        'avg', True),
    ('hitterPlus',      'Hitter+',      1,     True),
    ('sdPlus',          'SD+',          1,     True),
    ('ctPlus',          'CT+',          1,     True),
    ('bbPlus',          'BB+',          1,     True),
    ('xwOBAcon',        'xwOBAcon',     'avg', True),
    ('xwOBAsp',         'xwOBAsp',      'avg', True),
    ('babip',           'BABIP',        'avg', True),
    ('avgEVAll',        'Avg EV',       1,     True),
    ('ev50',            'EV50',         1,     True),
    ('maxEV',           'Max EV',       1,     True),
    ('hardHitPct',      'Hard-Hit%',    'pct', True),
    ('barrelPct',       'Barrel%',      'pct', True),
    ('airPullPct',      'Air Pull%',    'pct', True),
    ('pullPct',         'Pull%',        'pct', True),
    ('gbPct',           'GB%',          'pct', False),
    ('ldPct',           'LD%',          'pct', True),
    ('fbPct',           'FB%',          'pct', True),
    ('puPct',           'PU%',          'pct', False),
    ('bbPct',           'BB%',          'pct', True),
    ('kPct',            'K%',           'pct', False),
    ('chasePct',        'Chase%',       'pct', False),
    ('izSwingPct',      'Z-Swing%',     'pct', True),
    ('swingPct',        'Swing%',       'pct', False),
    ('whiffPct',        'Whiff%',       'pct', False),
    ('contactPct',      'Contact%',     'pct', True),
    ('izContactPct',    'Z-Contact%',   'pct', True),
    ('twoStrikeWhiffPct', '2K-Whiff%',  'pct', False),
    ('firstPitchSwingPct', 'FP-Swing%', 'pct', False),
    ('xRv100',          'xRV/100',      2,     True),
]

PITCHER_METRICS = [
    ('nPitches',        'Pitches',      0,     None),
    ('tbf',             'TBF',          0,     None),
    ('nBip',            'BIP',          0,     None),
    ('kPct',            'K%',           'pct', True),
    ('bbPct',           'BB%',          'pct', False),
    ('kbbPct',          'K-BB%',        'pct', True),
    ('siera',           'SIERA (split)', 2,    False),
    ('wOBA',            'wOBA',         'avg', False),
    ('xwOBA',           'xwOBA',        'avg', False),
    ('xRv100',          'xRV/100',      2,     True),
    ('rv100',           'RV/100',       2,     True),
    ('swStrPct',        'Whiff%',       'pct', True),
    ('izWhiffPct',      'Z-Whiff%',     'pct', True),
    ('chasePct',        'Chase%',       'pct', True),
    ('cswPct',          'CSW%',         'pct', True),
    ('izPct',           'Zone%',        'pct', True),
    ('strikePct',       'Strike%',      'pct', True),
    ('fpsPct',          'FP-Strike%',   'pct', True),
    ('twoStrikeWhiffPct', '2K-Whiff%',  'pct', True),
    ('xwOBAcon',        'xwOBAcon',     'avg', False),
    ('babip',           'BABIP',        'avg', False),
    ('avgEVAgainst',    'Avg EV',       1,     False),
    ('hardHitPct',      'Hard-Hit%',    'pct', False),
    ('barrelPctAgainst', 'Barrel%',     'pct', False),
    ('gbPct',           'GB%',          'pct', True),
    ('ldPct',           'LD%',          'pct', False),
    ('fbPct',           'FB%',          'pct', False),
    ('puPct',           'PU%',          'pct', True),
    ('hrFbPct',         'HR/FB%',       'pct', False),
    ('fbVelo',          'FB Velo',      1,     True),
    ('stuffPlus',       'Stuff+',       1,     True),
    ('locPlus',         'Loc+',         1,     True),
]

# Hitter pitch-group x hand: the "what is beating him" layer. Groups rather
# than individual pitch types because an 84-PA vs-LHP sample cannot carry a
# per-type breakdown. Matches HitterCards.PITCH_GROUPS (KN in Offspeed).
PITCH_GROUPS = {
    'Hard':     {'FF', 'SI'},
    'Breaking': {'FC', 'SL', 'ST', 'CU', 'SV'},
    'Offspeed': {'CH', 'FS', 'KN'},
}

HGROUP_METRICS = [
    ('nPitches',        'Pitches',      0,     None),
    ('seenPct',         'Seen%',        'pct', None),
    ('swingPct',        'Swing%',       'pct', None),
    ('chasePct',        'Chase%',       'pct', False),
    ('izSwingPct',      'Z-Swing%',     'pct', True),
    ('whiffPct',        'Whiff%',       'pct', False),
    ('contactPct',      'Contact%',     'pct', True),
    ('izContactPct',    'Z-Contact%',   'pct', True),
    ('nBip',            'BIP',          0,     None),
    ('xwOBAcon',        'xwOBAcon',     'avg', True),
    ('avgEVAll',        'Avg EV',       1,     True),
    ('hardHitPct',      'Hard-Hit%',    'pct', True),
    ('barrelPct',       'Barrel%',      'pct', True),
    ('airPullPct',      'Air Pull%',    'pct', True),
    ('gbPct',           'GB%',          'pct', False),
    ('xRv100',          'xRV/100',      2,     True),
]


def hitter_group_rows(pitches, lg, total_pitches):
    out = {}
    for gname, types in PITCH_GROUPS.items():
        ps = [p for p in pitches if p.get('Pitch Type') in types]
        if not ps:
            continue
        st = compute_hitter_stats(ps)
        ex = compute_expected_stats(ps, woba_weights=lg.woba_weights)
        st['nPitches'] = len(ps)
        st['seenPct'] = len(ps) / total_pitches if total_pitches else None
        st['xwOBAcon'] = ex.get('xwOBAcon')
        st['xRv100'] = lg.xrv100(ps, negate=True)
        out[gname] = st
    return out


ARSENAL_METRICS = [
    ('count',           'Count',        0,     None),
    ('usagePct',        'Usage%',       'pct', None),
    ('velocity',        'Velo',         1,     True),
    ('maxVelo',         'Max Velo',     1,     True),
    ('spinRate',        'Spin',         0,     None),
    ('indVertBrk',      'IVB',          1,     None),
    ('horzBrk',         'HB',           1,     None),
    ('extension',       'Ext',          2,     True),
    ('armAngle',        'Arm Angle',    1,     None),
    ('stuffPlus',       'Stuff+',       1,     True),
    ('locPlus',         'Loc+',         1,     True),
    ('izPct',           'Zone%',        'pct', True),
    ('swStrPct',        'Whiff%',       'pct', True),
    ('izWhiffPct',      'Z-Whiff%',     'pct', True),
    ('chasePct',        'Chase%',       'pct', True),
    ('cswPct',          'CSW%',         'pct', True),
    ('nBip',            'BIP',          0,     None),
    ('gbPct',           'GB%',          'pct', True),
    ('hardHitPct',      'Hard-Hit%',    'pct', False),
    ('barrelPctAgainst', 'Barrel%',     'pct', False),
    ('xwOBAcon',        'xwOBAcon',     'avg', False),
    ('rv100',           'RV/100',       2,     True),
    ('xRv100',          'xRV/100',      2,     True),
]


def fmt(value, spec):
    if value is None:
        return ''
    if spec == 'pct':
        return f"{value * 100:.1f}"
    if spec == 'avg':
        s = f"{value:.3f}"
        return s[1:] if s.startswith('0.') else (
            '-' + s[2:] if s.startswith('-0.') else s)
    return f"{value:.{spec}f}"


# ═════════════════════════════════════════════════════════════════════════
#  MAIN
# ═════════════════════════════════════════════════════════════════════════

SPLITS_H = [('All', None), ('vs RHP', 'R'), ('vs LHP', 'L')]
SPLITS_P = [('All', None), ('vs RHH', 'R'), ('vs LHH', 'L')]


def split_pitches(pitches, hand, field):
    """Hand split, plus the rolling-window filter when --days is set.

    The window is applied HERE so it hits the target players and the MLB
    percentile pool identically: a 30-day cell is ranked against every
    qualified MLB player's own last 30 days. League anchors deliberately stay
    full season (SD+/CT+ cell tables, BB+ denominator, SACQ zones, xRV count
    offsets) — a 30-day cell table would be noise, and keeping them fixed is
    what makes the window comparable to the season rows above it."""
    if DATE_FROM is not None:
        pitches = [p for p in pitches if (p.get('Game Date') or '') >= DATE_FROM]
    if hand is None:
        return pitches
    return [p for p in pitches if p.get(field) == hand]


def main():
    global DATE_FROM, MIN_LEVEL_PA, MIN_SPLIT_PA, POOL_MIN_PA, POOL_MIN_TBF
    global MIN_PITCHES_FOR_ROW, MIN_ARSENAL_PITCHES, MIN_GROUP_PITCHES
    ap = argparse.ArgumentParser()
    ap.add_argument('--outdir', default=os.path.expanduser('~/Downloads'))
    ap.add_argument('--days', type=int, default=None, metavar='N',
                    help='rolling window: keep only the last N days of games '
                         '(counted back from the latest game in the data). '
                         'Applies to the MLB percentile pool too.')
    ap.add_argument('--no-new-tab', action='store_true')
    ap.add_argument('--validate', action='store_true',
                    help='diff the engine against the shipped leaderboards')
    args = ap.parse_args()

    with open(os.path.join(DATA_DIR, 'metadata_rs.json')) as f:
        metadata = json.load(f)

    # Prefer the weights the pipeline shipped in metadata (2026-08-04): that is
    # the set the leaderboard was actually built with, so reading it keeps this
    # script in step with the site across a season rollover without anyone
    # remembering to edit a constant. Fall back to a live scrape, then to the
    # pinned set. Verified 2026-08-04: all three agree exactly.
    woba_weights = metadata.get('wobaWeights')
    if woba_weights and set(WOBA_WEIGHTS_FALLBACK) <= set(woba_weights):
        print(f"  Using the wOBA weights shipped in metadata: {woba_weights}")
    else:
        try:
            from pipeline.fetch import fetch_guts_constants
            woba_weights, _fip, _extra = fetch_guts_constants(2026)
            print("  Fetched live FanGraphs Guts wOBA weights")
        except Exception as e:                                   # noqa: BLE001
            print(f"  Guts fetch failed ({e}); using the pinned 2026 weights")
            woba_weights = WOBA_WEIGHTS_FALLBACK

    print("\n=== Loading pitch data ===")
    pitches, new_rows = load_pitches(use_new_tab=not args.no_new_tab)
    all_pitches = pitches + new_rows

    if args.days:
        import datetime as _dt
        last = max((p.get('Game Date') or '') for p in all_pitches)
        DATE_FROM = (_dt.date.fromisoformat(last)
                     - _dt.timedelta(days=args.days - 1)).isoformat()
        for _n in ('MIN_LEVEL_PA', 'MIN_SPLIT_PA', 'POOL_MIN_PA', 'POOL_MIN_TBF',
                   'MIN_PITCHES_FOR_ROW', 'MIN_ARSENAL_PITCHES', 'MIN_GROUP_PITCHES'):
            globals()[_n] = max(5, int(globals()[_n] * WINDOW_FLOOR_SCALE))
        MIN_LEVEL_PA, MIN_SPLIT_PA = globals()['MIN_LEVEL_PA'], globals()['MIN_SPLIT_PA']
        POOL_MIN_PA, POOL_MIN_TBF = globals()['POOL_MIN_PA'], globals()['POOL_MIN_TBF']
        MIN_PITCHES_FOR_ROW = globals()['MIN_PITCHES_FOR_ROW']
        MIN_ARSENAL_PITCHES = globals()['MIN_ARSENAL_PITCHES']
        MIN_GROUP_PITCHES = globals()['MIN_GROUP_PITCHES']
        print(f"\n=== Rolling window: {args.days} days, {DATE_FROM} to {last} ===")
        print(f"  floors scaled x{WINDOW_FLOOR_SCALE}: level {MIN_LEVEL_PA} PA, "
              f"split {MIN_SPLIT_PA} PA, pool {POOL_MIN_PA} PA / {POOL_MIN_TBF} TBF")

    print("\n=== Preprocessing ===")
    ep_pitchers = preprocess(all_pitches, new_rows, metadata, woba_weights)

    print("\n=== Building league tables ===")
    lg = League(all_pitches, ep_pitchers, metadata, woba_weights)

    # ── Group pitches ────────────────────────────────────────────────────
    # Hitters: keyed (name, team); pitchers: keyed (name, team).
    hitter_groups = defaultdict(list)
    pitcher_groups = defaultdict(list)
    for p in all_pitches:
        if not p.get('_roc_pitcher_pitch'):
            b, bt = p.get('Batter'), p.get('BTeam')
            if b and bt:
                hitter_groups[(b, bt)].append(p)
        if not p.get('_roc_hitter_pitch'):
            pt_ = (p.get('Pitcher'), p.get('PTeam'))
            if pt_[0] and pt_[1] and pt_ not in ep_pitchers:
                pitcher_groups[pt_].append(p)

    # SD+/CT+ groups exclude position-player pitching (skill-metric rule).
    sd_groups = {k: [p for p in v
                     if (p.get('Pitcher'), p.get('PTeam')) not in ep_pitchers]
                 for k, v in hitter_groups.items()}
    lg.anchor_from_full_season(sd_groups)

    # ── Percentile pools: MLB players qualified under the site's own rule ──
    print("\n=== Building MLB percentile pools ===")
    with open(os.path.join(DATA_DIR, 'hitter_leaderboard_rs.json')) as f:
        h_lb = json.load(f)
    with open(os.path.join(DATA_DIR, 'pitcher_leaderboard_rs.json')) as f:
        p_lb = json.load(f)

    qual_hitters = []
    for r in h_lb:
        t = r.get('team')
        if t not in MLB_TEAMS:
            continue
        tg = lg.team_games.get(t, lg.max_tg)
        if (r.get('pa') or 0) >= hitter_pa_per_game(False) * tg:
            qual_hitters.append((r['hitter'], t))
    qual_pitchers = []
    for r in p_lb:
        t = r.get('team')
        if t not in MLB_TEAMS:
            continue
        tg = lg.team_games.get(t, lg.max_tg)
        ip_f = ip_str_to_float(r.get('ip')) if r.get('ip') is not None else 0
        g, gs = r.get('g') or 0, r.get('gs') or 0
        is_sp = g > 0 and (gs / g) > SP_GS_RATIO
        if ip_f >= tg * pitcher_ip_per_game(is_sp, False):
            qual_pitchers.append((r['pitcher'], t, gs / g if g else 0.0))
    print(f"  {len(qual_hitters)} qualified MLB hitters, "
          f"{len(qual_pitchers)} qualified MLB pitchers")

    hitter_pool = {label: defaultdict(list) for label, _ in SPLITS_H}
    hgroup_pool = {label: defaultdict(list) for label, _ in SPLITS_H}
    for key in qual_hitters:
        ps_all = hitter_groups.get(key) or []
        sd_all = sd_groups.get(key) or []
        for label, hand in SPLITS_H:
            ps = split_pitches(ps_all, hand, 'Throws')
            if not ps:
                continue
            row = hitter_row(ps, split_pitches(sd_all, hand, 'Throws'), lg)
            if (row.get('pa') or 0) < POOL_MIN_PA:
                continue
            for k, _lbl, _f, hib in HITTER_METRICS:
                if hib is not None and row.get(k) is not None:
                    hitter_pool[label][k].append(row[k])
            for gname, g in hitter_group_rows(ps, lg, len(ps)).items():
                if g['nPitches'] < 50:
                    continue
                for k, _lbl, _f, hib in HGROUP_METRICS:
                    if hib is not None and g.get(k) is not None:
                        hgroup_pool[label][(gname, k)].append(g[k])
    print("  Hitter pools built")

    pitcher_pool = {label: defaultdict(list) for label, _ in SPLITS_P}
    arsenal_pool = {label: defaultdict(list) for label, _ in SPLITS_P}
    for name, team, gs_ratio in qual_pitchers:
        ps_all = pitcher_groups.get((name, team)) or []
        for label, hand in SPLITS_P:
            ps = split_pitches(ps_all, hand, 'Bats')
            if not ps:
                continue
            row = pitcher_row(ps, lg, gs_ratio)
            if (row.get('tbf') or 0) < POOL_MIN_TBF:
                continue
            for k, _lbl, _f, hib in PITCHER_METRICS:
                if hib is not None and row.get(k) is not None:
                    pitcher_pool[label][k].append(row[k])
            for pt, ar in arsenal_rows(ps, lg, len(ps)).items():
                if ar['count'] < 25:
                    continue
                for k, _lbl, _f, hib in ARSENAL_METRICS:
                    if hib is not None and ar.get(k) is not None:
                        arsenal_pool[label][(pt, k)].append(ar[k])
    print("  Pitcher + arsenal pools built")

    # ── Target rows ──────────────────────────────────────────────────────
    print("\n=== Computing target players ===")
    h_out, p_out, a_out, g_out = [], [], [], []
    for display, sheet_name, role, tab in TARGETS:
        levels = []
        if role == 'H':
            levels.append(('AAA', hitter_groups.get((sheet_name, 'ROC')) or [],
                           sd_groups.get((sheet_name, 'ROC')) or []))
            for t in MLB_TEAMS:
                g = hitter_groups.get((sheet_name, t))
                if g:
                    levels.append(('MLB (' + t + ')', g,
                                   sd_groups.get((sheet_name, t)) or []))
        else:
            if tab == 'ROC':
                levels.append(('AAA', pitcher_groups.get((sheet_name, 'ROC')) or [], None))
            else:
                milb = [p for p in new_rows if p.get('Pitcher') == sheet_name]
                if milb:
                    levels.append(('AAA', milb, None))
                for mt in (NEW_TAB_MLB_TEAM.get(sheet_name) or []):
                    g = pitcher_groups.get((sheet_name, mt))
                    if g:
                        levels.append(('MLB (' + mt + ')', g, None))

        for level, ps_all, sd_all in levels:
            if len(ps_all) < MIN_PITCHES_FOR_ROW:
                continue
            from pipeline.utils import NON_PA_EVENTS as _NPE
            n_level_pa = sum(1 for p in ps_all
                             if p.get('Event') and p['Event'] not in _NPE)
            if n_level_pa < MIN_LEVEL_PA:
                print(f"  {display:16s} {level:12s} skipped ({n_level_pa} PA)")
                continue
            splits = SPLITS_H if role == 'H' else SPLITS_P
            field = 'Throws' if role == 'H' else 'Bats'
            gs_ratio = 0.0
            if role == 'P':
                lb_row = next((r for r in p_lb if r.get('pitcher') == sheet_name), None)
                if lb_row and (lb_row.get('g') or 0):
                    gs_ratio = min((lb_row.get('gs') or 0) / lb_row['g'], 1.0)
            for label, hand in splits:
                ps = split_pitches(ps_all, hand, field)
                if not ps:
                    continue
                base = {'Player': display, 'Level': level, 'Split': label}
                if role == 'H':
                    row = hitter_row(ps, split_pitches(sd_all, hand, 'Throws'), lg)
                    out, pool = h_out, hitter_pool[label]
                    metrics = HITTER_METRICS
                else:
                    row = pitcher_row(ps, lg, gs_ratio)
                    out, pool = p_out, pitcher_pool[label]
                    metrics = PITCHER_METRICS
                if hand is not None and (row.get('pa') or 0) < MIN_SPLIT_PA:
                    continue
                if role == 'H':
                    for gname, g in hitter_group_rows(ps, lg, len(ps)).items():
                        if g['nPitches'] < MIN_GROUP_PITCHES:
                            continue
                        grow = {'Player': display, 'Level': level,
                                'Split': label, 'Pitch Group': gname}
                        for k, lbl, spec, hib in HGROUP_METRICS:
                            grow[lbl] = fmt(g.get(k), spec)
                            if hib is not None:
                                grow[lbl + ' %ile'] = percentile(
                                    g.get(k), hgroup_pool[label].get((gname, k), []), hib)
                        g_out.append(grow)
                if role == 'P':
                    for pt, ar in arsenal_rows(ps, lg, len(ps)).items():
                        if ar['count'] < MIN_ARSENAL_PITCHES:
                            continue
                        arow = {'Player': display, 'Level': level,
                                'Split': label, 'Pitch': pt}
                        for k, lbl, spec, hib in ARSENAL_METRICS:
                            arow[lbl] = fmt(ar.get(k), spec)
                            if hib is not None:
                                arow[lbl + ' %ile'] = percentile(
                                    ar.get(k), arsenal_pool[label].get((pt, k), []), hib)
                        a_out.append(arow)
                rec = dict(base)
                for k, lbl, spec, hib in metrics:
                    rec[lbl] = fmt(row.get(k), spec)
                    if hib is not None:
                        rec[lbl + ' %ile'] = percentile(row.get(k), pool.get(k, []), hib)
                if role == 'H':
                    rec['SD+ n'] = row.get('sdPlusN')
                    rec['CT+ n'] = row.get('ctPlusN')
                out.append(rec)
                print(f"  {display:16s} {level:12s} {label:7s} "
                      f"{'PA' if role == 'H' else 'TBF'}="
                      f"{row.get('pa') if role == 'H' else row.get('tbf')}")

    if args.validate:
        validate(lg, pitcher_groups, hitter_groups, sd_groups, p_lb, h_lb,
                 qual_pitchers)

    write_outputs(args.outdir, h_out, p_out, a_out, g_out, metadata,
                  window_days=args.days)


def validate(lg, pitcher_groups, hitter_groups, sd_groups, p_lb, h_lb,
             qual_pitchers):
    """Prove the split engine reproduces the pipeline.

    (a) The shipped pitcher leaderboard already carries 23 stats split by
        batter hand (kPct_vsL, chasePct_vsR, ...). Recompute them here and
        diff. These should match to rounding.
    (b) Season-total ("All") rows for the target-style players are compared
        against the shipped season row. Slash-line stats are boxscore-sourced
        upstream, so a small gap there is expected and reported separately.
    """
    print("\n=== VALIDATION (a): pitcher hand splits vs shipped _vsL/_vsR ===")
    check = ['kPct', 'bbPct', 'kbbPct', 'chasePct', 'swStrPct', 'izWhiffPct',
             'cswPct', 'izPct', 'strikePct', 'fpsPct', 'gbPct', 'babip',
             'hardHitPct', 'ldPct', 'fbPct', 'puPct', 'wOBA', 'xwOBA',
             'xwOBAcon', 'twoStrikeWhiffPct']
    worst = {k: (0.0, None) for k in check}
    n_cmp = 0
    lb_by_key = {(r['pitcher'], r['team']): r for r in p_lb}
    for name, team, gs_ratio in qual_pitchers[:60]:
        ref = lb_by_key.get((name, team))
        ps_all = pitcher_groups.get((name, team)) or []
        if not ref or not ps_all:
            continue
        for hand, suffix in (('R', '_vsR'), ('L', '_vsL')):
            mine = pitcher_row([p for p in ps_all if p.get('Bats') == hand],
                               lg, gs_ratio)
            n_cmp += 1
            for k in check:
                a, b = mine.get(k), ref.get(k + suffix)
                if a is None or b is None:
                    continue
                d = abs(a - b)
                if d > worst[k][0]:
                    worst[k] = (d, f"{name} {suffix}")
    print(f"  compared {n_cmp} pitcher-hand rows")
    bad = False
    for k, (d, who) in sorted(worst.items(), key=lambda x: -x[1][0]):
        flag = '' if d < 5e-4 else '   <-- MISMATCH'
        if flag:
            bad = True
        print(f"    {k:20s} max |diff| = {d:.6f}  ({who}){flag}")
    print("  RESULT: " + ("MISMATCHES FOUND" if bad else "all within rounding"))

    print("\n=== VALIDATION (b): season totals vs the shipped leaderboard ===")
    for name, team in [('Perales, Luis', 'ROC'), ('Kent, Jackson', 'ROC'),
                       ('Bird, Jake', 'NYY')]:
        ref = lb_by_key.get((name, team))
        ps = pitcher_groups.get((name, team)) or []
        if not ref or not ps:
            continue
        g = ref.get('g') or 1
        mine = pitcher_row(ps, lg, min((ref.get('gs') or 0) / g, 1.0))
        print(f"  {name} ({team})")
        for k in ['kPct', 'bbPct', 'chasePct', 'swStrPct', 'izWhiffPct',
                  'xwOBA', 'xwOBAcon', 'babip', 'gbPct', 'hardHitPct',
                  'barrelPctAgainst', 'xRv100', 'stuffPlus', 'locPlus',
                  'fbVelo']:
            a, b = mine.get(k), ref.get(k)
            if a is None or b is None:
                print(f"    {k:18s} mine={a} shipped={b}")
                continue
            print(f"    {k:18s} mine={a:9.4f} shipped={b:9.4f} diff={a - b:+.4f}")

    h_by_key = {(r['hitter'], r['team']): r for r in h_lb}
    for name, team in [('Ortiz, Abimelec', 'ROC'), ('Pinckney, Andrew', 'ROC')]:
        ref = h_by_key.get((name, team))
        ps = hitter_groups.get((name, team)) or []
        if not ref or not ps:
            continue
        mine = hitter_row(ps, sd_groups.get((name, team)) or [], lg)
        print(f"  {name} ({team})")
        for k in ['pa', 'avg', 'obp', 'slg', 'wOBA', 'xwOBA', 'xwOBAcon',
                  'xwOBAsp', 'babip', 'kPct', 'bbPct', 'hardHitPct',
                  'barrelPct', 'airPullPct', 'chasePct', 'izContactPct',
                  'maxEV', 'bbPlus', 'sdPlus', 'ctPlus', 'hitterPlus']:
            a, b = mine.get(k), ref.get(k)
            if a is None or b is None:
                print(f"    {k:18s} mine={a} shipped={b}")
                continue
            print(f"    {k:18s} mine={a:9.4f} shipped={b:9.4f} diff={a - b:+.4f}")


# ═════════════════════════════════════════════════════════════════════════
#  OUTPUT
# ═════════════════════════════════════════════════════════════════════════

def write_outputs(outdir, h_out, p_out, a_out, g_out, metadata, window_days=None):
    import csv
    os.makedirs(outdir, exist_ok=True)
    stamp = metadata.get('generatedAt', '')[:10] or 'latest'
    if window_days:
        stamp = f'last{window_days}d_{stamp}'
    files = []

    def dump(name, rows):
        if not rows:
            return None
        path = os.path.join(outdir, name)
        cols = list(rows[0].keys())
        with open(path, 'w', newline='') as f:
            w = csv.DictWriter(f, fieldnames=cols)
            w.writeheader()
            w.writerows(rows)
        files.append(path)
        return path

    dump(f'wsh_platoon_hitters_{stamp}.csv', h_out)
    dump(f'wsh_platoon_pitchers_{stamp}.csv', p_out)
    dump(f'wsh_platoon_arsenal_{stamp}.csv', a_out)
    dump(f'wsh_platoon_hitter_pitchgroups_{stamp}.csv', g_out)

    xlsx = os.path.join(outdir, f'wsh_platoon_splits_{stamp}.xlsx')
    write_xlsx(xlsx, [('Hitters', h_out), ('Hitter Pitch Groups', g_out),
                      ('Pitchers', p_out), ('Arsenal', a_out)])
    files.append(xlsx)

    print("\nWrote:")
    for f in files:
        print("  " + f)


README_LINES = [
    'Nationals AAA platoon splits, 2026 season',
    '',
    'What each sheet holds:',
    '    Hitters              one row per player x level x split (All / vs RHP / vs LHP)',
    '    Hitter Pitch Groups  the same splits broken down by Hard / Breaking / Offspeed',
    '    Pitchers             one row per player x level x split (All / vs RHH / vs LHH)',
    '    Arsenal              pitch type x split, for every pitch a pitcher threw 25+ of',
    '    ...  %ile            the same sheet with the raw percentile numbers, for sorting',
    '',
    'Cell colors:',
    '    Percentile against MLB players who qualify under the site rule, measured on the',
    '    SAME split. So a colored "vs LHP" cell is ranked against every qualified MLB',
    '    hitter\'s own vs-LHP number, not against their season line.',
    '',
    'Reading the AAA rows:',
    '    AAA numbers are ranked RAW against the MLB distribution with no level',
    '    translation, exactly as the player cards and the site do. A 70th-percentile',
    '    AAA cell is not a 70th-percentile major leaguer.',
    '',
    'Baselines:',
    '    Only the player\'s own pitches are split. Every league anchor (SD+/CT+ cell',
    '    tables, BB+ denominator, Hitter+ standardization, SACQ zones, xRV count',
    '    offsets) stays full-season and MLB, so vs-R and vs-L are on one scale and',
    '    comparable to the season card.',
    '',
    'Known deviations from the card:',
    '    AVG / OBP / SLG / wOBA / K% / BB% on the card come from the official boxscore,',
    '    which catches no-pitch intentional walks. A split has no boxscore, so these are',
    '    pitch-derived here. Ortiz is 364 pitch-derived PA against 365 official.',
    '    SIERA is recomputed on the split\'s own K/BB/GB/FB rates with the shipped',
    '    constant, so it will differ slightly from the season SIERA.',
    '',
    'Sample floors:',
    '    A level needs 40 PA/TBF, a hand split needs 20. SD+ n and CT+ n columns show',
    '    the decision/swing counts behind those two grades; both regress to league at',
    '    n0 = 200 (SD+) and 65 (CT+), so a small-n split is pulled hard toward 100.',
    '',
    'Legend:',
]


def pctl_fill(p):
    """Card palette: deep terracotta at the top, slate blue at the bottom,
    warm paper in the middle. Returns an aRGB hex string."""
    if p is None:
        return None
    t = (p - 50) / 50.0
    if t >= 0:
        base = (196, 74, 56)     # terracotta
    else:
        base = (94, 124, 153)    # slate blue
        t = -t
    bg = (245, 241, 232)         # cream
    a = 0.10 + 0.75 * min(1.0, t)
    rgb = tuple(round(bg[i] + (base[i] - bg[i]) * a) for i in range(3))
    return 'FF%02X%02X%02X' % rgb


def as_number(s):
    """Turn a formatted string back into (value, Excel number format) so the
    workbook holds real numbers — text cells sort lexicographically, which
    would put -2.64 next to -0.44 and .162 next to .1. Display is unchanged:
    the format code reproduces the string exactly, leading-zero rule included."""
    if s is None or s == '':
        return None, None
    if not isinstance(s, str):
        return s, None
    t = s.lstrip('-')
    if not t.replace('.', '', 1).isdigit():
        return s, None
    try:
        v = float(s) if not t.startswith('.') else float(s.replace('.', '0.', 1))
    except ValueError:
        return s, None
    dec = len(t.split('.')[1]) if '.' in t else 0
    if t.startswith('.'):
        fmt_code = '.' + '0' * dec          # .416, no leading zero
    else:
        fmt_code = '0' + ('.' + '0' * dec if dec else '')
    return v, fmt_code


def write_xlsx(path, sheets):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet('Read me')
    for line in README_LINES:
        ws.append([line])
    ws.column_dimensions['A'].width = 118
    for r in ws.iter_rows():
        r[0].alignment = Alignment(wrap_text=False)
        if r[0].value and r[0].value.endswith(':'):
            r[0].font = Font(bold=True)
    for i, (lab, p) in enumerate([('99th', 99), ('90th', 90), ('75th', 75),
                                  ('50th', 50), ('25th', 25), ('10th', 10),
                                  ('1st', 1)], start=len(README_LINES) + 2):
        ws.cell(row=i, column=1, value=f'    {lab} percentile')
        c = ws.cell(row=i, column=2, value=p)
        c.fill = PatternFill('solid', fgColor=pctl_fill(p))
        c.alignment = Alignment(horizontal='center')

    for title, rows in sheets:
        ws = wb.create_sheet(title)
        if not rows:
            continue
        cols = [c for c in rows[0].keys() if not c.endswith('%ile')]
        ws.append(cols)
        for c in ws[1]:
            c.font = Font(bold=True, size=9)
            c.alignment = Alignment(horizontal='center', wrap_text=True)
        for ri, r in enumerate(rows, start=2):
            for ci, col in enumerate(cols, start=1):
                v, code = as_number(r.get(col))
                cell = ws.cell(row=ri, column=ci, value=v)
                if code:
                    cell.number_format = code
                p = r.get(col + ' %ile')
                if p in (None, ''):
                    continue
                p = int(p)
                cell.fill = PatternFill('solid', fgColor=pctl_fill(p))
                cell.font = Font(size=9, bold=p >= 80 or p <= 20)
                cell.alignment = Alignment(horizontal='center')
        ws.freeze_panes = ('E2' if title in ('Arsenal', 'Hitter Pitch Groups')
                           else 'D2')
        for ci, col in enumerate(cols, start=1):
            width = max(len(str(col)) + 2, 8)
            ws.column_dimensions[get_column_letter(ci)].width = min(width, 16)

    # A companion sheet with the raw percentile numbers, for sorting.
    for title, rows in sheets:
        if not rows:
            continue
        ws = wb.create_sheet(title + ' %ile')
        cols = list(rows[0].keys())
        ws.append(cols)
        for c in ws[1]:
            c.font = Font(bold=True, size=9)
        for ri, r in enumerate(rows, start=2):
            for ci, col in enumerate(cols, start=1):
                v, code = as_number(r.get(col))
                cell = ws.cell(row=ri, column=ci, value=v)
                if code:
                    cell.number_format = code
    wb.save(path)


if __name__ == '__main__':
    main()
