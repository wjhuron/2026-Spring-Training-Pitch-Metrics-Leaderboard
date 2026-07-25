"""
Export ROC pitch-mix comps to xlsx.
Sheets:
  1. Top Comps (soft hand)      -- top 3, soft same-hand preference
  2. Same-Hand Only             -- top 3 restricted to same throwing hand
  3. PitchByPitch (soft)        -- per-pitch best-match detail for the soft-hand top 3
  4. PitchByPitch (same-hand)   -- per-pitch best-match detail for the same-hand top 3
HB and relX are shown in a common right-handed frame (LHP mirrored), the same
frame used for matching, so cross-hand comps line up. velo/ivb/relz/ext are
frame-independent.
"""
import json, math, os
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

D = json.load(open('data/pitch_leaderboard_rs.json'))

MLB30 = {'ARI','ATH','ATL','BAL','BOS','CHC','CIN','CLE','COL','CWS','DET','HOU',
         'KCR','LAA','LAD','MIA','MIL','MIN','NYM','NYY','PHI','PIT','SDP','SEA',
         'SFG','STL','TBR','TEX','TOR','WSH'}
AGG = {'2TM','3TM'}
FEATS = ['velocity','indVertBrk','horzBrk','relPosZ','relPosX','extension']
W = {'velocity':1.0,'indVertBrk':1.0,'horzBrk':1.0,'relPosZ':0.6,'relPosX':0.6,'extension':0.6}
HAND_PEN = 1.08
KSIM = 0.55

TARGETS = ['Champlain, Chandler','Cranz, Robert','Kent, Jackson','Kent, Zak',
           'Lara, Andry','Penrod, Zach','Perales, Luis','Sinclair, Jack',
           'Tolman, Erik','Yean, Eddy','Young, Luke']

def valid(r):
    return all(r.get(f) is not None for f in FEATS)

def normframe(r):
    hb = r['horzBrk']; rx = r['relPosX']
    if r['throws'] == 'L':
        hb = -hb; rx = -rx
    return {'velocity':r['velocity'],'indVertBrk':r['indVertBrk'],'horzBrk':hb,
            'relPosZ':r['relPosZ'],'relPosX':rx,'extension':r['extension']}

by_id = defaultdict(list)
for r in D:
    by_id[r['mlbId']].append(r)

target_ids = set()
for t in TARGETS:
    for r in D:
        if r['pitcher']==t and r['team']=='ROC':
            target_ids.add(r['mlbId'])

candidates = {}
for mid, rows in by_id.items():
    if mid in target_ids: continue
    teams = set(r['team'] for r in rows)
    if teams & AGG:      use = [r for r in rows if r['team'] in AGG]
    elif teams & MLB30:  use = [r for r in rows if r['team'] in MLB30]
    else:                continue
    use = [r for r in use if valid(r) and r['count'] >= 15]
    if not use or sum(r['count'] for r in use) < 200: continue
    pitches = [{'pt':r['pitchType'],'usage':r['usagePct'],'vec':normframe(r)} for r in use]
    us = sum(p['usage'] for p in pitches)
    for p in pitches: p['w'] = p['usage']/us if us else 0
    candidates[mid] = {'name':use[0]['pitcher'],'throws':use[0]['throws'],'pitches':pitches}

pool = [p['vec'] for c in candidates.values() for p in c['pitches']]
def msd(f):
    xs=[v[f] for v in pool]; m=sum(xs)/len(xs)
    return m,(sum((x-m)**2 for x in xs)/len(xs))**0.5
STATS = {f:msd(f) for f in FEATS}
def z(vec): return {f:(vec[f]-STATS[f][0])/STATS[f][1] for f in FEATS}

targets = {}
for t in TARGETS:
    rows=[r for r in D if r['pitcher']==t and r['team']=='ROC' and valid(r)]
    pitches=[{'pt':r['pitchType'],'usage':r['usagePct'],'vec':normframe(r)} for r in rows]
    us=sum(p['usage'] for p in pitches)
    for p in pitches: p['w']=p['usage']/us if us else 0
    targets[t]={'throws':rows[0]['throws'],'pitches':pitches}

def pdist(za,zb):
    return math.sqrt(sum(W[f]*(za[f]-zb[f])**2 for f in FEATS))

def arsenal_dist(tgt,cand):
    tz=[(p['w'],z(p['vec'])) for p in tgt['pitches']]
    cz=[(p['w'],z(p['vec'])) for p in cand['pitches']]
    fwd=sum(w*min(pdist(zv,czv) for _,czv in cz) for w,zv in tz)
    rev=sum(w*min(pdist(czv,zv) for _,zv in tz) for w,czv in cz)
    return 0.65*fwd+0.35*rev

def rank(t, same_hand_only):
    tgt=targets[t]; scored=[]
    for mid,c in candidates.items():
        if same_hand_only and c['throws']!=tgt['throws']: continue
        d=arsenal_dist(tgt,c)
        if not same_hand_only and c['throws']!=tgt['throws']: d*=HAND_PEN
        scored.append((100*math.exp(-KSIM*d), c))
    scored.sort(key=lambda x:-x[0])
    return scored[:3]

def arsenal_str(tgt):
    return ", ".join(f"{p['pt']}{int(round(p['usage']*100))}"
                     for p in sorted(tgt['pitches'],key=lambda x:-x['usage']))

def disp_name(nm):  # "Last, First" -> "First Last"
    if ", " in nm:
        l,f=nm.split(", ",1); return f"{f} {l}"
    return nm

# ---------- workbook ----------
wb=openpyxl.Workbook()
HDR=Font(bold=True,color='FFFFFF'); HFILL=PatternFill('solid',fgColor='2F5496')
SUB=Font(bold=True,color='FFFFFF',size=9); SFILL=PatternFill('solid',fgColor='8EAADB')
TITLE=Font(bold=True,size=12); NOTE=Font(italic=True,size=9,color='666666')
thin=Side(style='thin',color='D9D9D9'); BORDER=Border(*[thin]*4)
CEN=Alignment(horizontal='center'); LEFT=Alignment(horizontal='left')

def style_header(ws,row,ncols):
    for c in range(1,ncols+1):
        cell=ws.cell(row=row,column=c); cell.font=HDR; cell.fill=HFILL
        cell.alignment=CEN; cell.border=BORDER

def comps_sheet(ws, same_hand_only, title):
    ws.append([title]); ws['A1'].font=TITLE
    ws.append(['Similarity index (0-100): higher = closer whole-arsenal shape match. '
               '73+ tight, 66-72 solid, <66 looser.']); ws['A2'].font=NOTE
    ws.append([])
    hdr=['ROC Pitcher','Hand','Arsenal (usage%)','Rank','MLB Comp','Comp Hand','Similarity']
    ws.append(hdr); style_header(ws,4,len(hdr))
    for t in TARGETS:
        tgt=targets[t]; top=rank(t,same_hand_only)
        for i,(sim,c) in enumerate(top,1):
            ws.append([disp_name(t) if i==1 else '',
                       tgt['throws'] if i==1 else '',
                       arsenal_str(tgt) if i==1 else '',
                       i, disp_name(c['name']), c['throws'], round(sim,1)])
        # thin separator between pitchers
    widths=[22,6,26,6,22,10,11]
    for i,w in enumerate(widths,1): ws.column_dimensions[chr(64+i)].width=w
    for row in ws.iter_rows(min_row=4):
        for cell in row:
            cell.border=BORDER
            if cell.column>=4: cell.alignment=CEN

def pbp_sheet(ws, same_hand_only, title):
    ws.append([title]); ws['A1'].font=TITLE
    ws.append(['HB and relX are in a right-handed frame (LHP mirrored) -- the frame used for matching. '
               'velo/ivb/relz/ext are frame-independent. Each row: ROC value / comp value.'])
    ws['A2'].font=NOTE
    ws.append([])
    hdr=['ROC Pitcher','Rank','MLB Comp','Comp Hand','Sim','ROC Pitch','Usage%','Comp Pitch',
         'velo R','velo C','ivb R','ivb C','hb R','hb C','relz R','relz C','relx R','relx C','ext R','ext C']
    ws.append(hdr); style_header(ws,4,len(hdr))
    for t in TARGETS:
        tgt=targets[t]; top=rank(t,same_hand_only)
        for i,(sim,c) in enumerate(top,1):
            first=True
            for p in sorted(tgt['pitches'],key=lambda x:-x['usage']):
                zp=z(p['vec'])
                best=min(c['pitches'],key=lambda cc:pdist(zp,z(cc['vec'])))
                pv=p['vec']; bv=best['vec']
                ws.append([disp_name(t) if first else '', i if first else '',
                           disp_name(c['name']) if first else '',
                           c['throws'] if first else '', round(sim,1) if first else '',
                           p['pt'], int(round(p['usage']*100)), best['pt'],
                           round(pv['velocity'],1),round(bv['velocity'],1),
                           round(pv['indVertBrk'],1),round(bv['indVertBrk'],1),
                           round(pv['horzBrk'],1),round(bv['horzBrk'],1),
                           round(pv['relPosZ'],1),round(bv['relPosZ'],1),
                           round(pv['relPosX'],1),round(bv['relPosX'],1),
                           round(pv['extension'],1),round(bv['extension'],1)])
                first=False
    widths=[22,5,22,6,6,10,7,10]+[7]*12
    for i,w in enumerate(widths,1): ws.column_dimensions[chr(64+i) if i<=26 else 'A'+chr(64+i-26)].width=w
    for row in ws.iter_rows(min_row=4):
        for cell in row:
            cell.border=BORDER
            if cell.column>=2: cell.alignment=CEN

ws1=wb.active; ws1.title='Top Comps (soft hand)'
comps_sheet(ws1, False, 'ROC Pitch-Mix Comps -- Soft Handedness Preference')
comps_sheet(wb.create_sheet('Same-Hand Only'), True, 'ROC Pitch-Mix Comps -- Same Throwing Hand Only')
pbp_sheet(wb.create_sheet('PitchByPitch (soft)'), False, 'Pitch-by-Pitch Breakdown -- Soft Hand (top 3)')
pbp_sheet(wb.create_sheet('PitchByPitch (same-hand)'), True, 'Pitch-by-Pitch Breakdown -- Same Hand (top 3)')

out=os.path.expanduser('~/Downloads/ROC_pitch_comps_2026-07-22.xlsx')
wb.save(out)
print('saved',out)
