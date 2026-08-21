#!/usr/bin/env python3
"""
Backfill supplemental Statcast data into the Google Sheet.

Scans each team tab for rows that have a PitchID but are missing supplemental
columns (ArmAngle, BatSpeed, SwingLength, AttackAngle,
AttackDirection, SwingPathTilt). Downloads the Statcast Search CSV for the
relevant team/date ranges from Baseball Savant and fills in the empty cells.

Configuration: edit the variables below before running.
"""

# Runnable as a file from any directory (IDE run buttons included):
# put the repo root on sys.path before the intra-repo package imports.
import os as _os, sys as _sys
_ROOT = _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__)))
if _ROOT not in _sys.path:
    _sys.path.insert(0, _ROOT)


import argparse
import gspread
import requests
import pandas as pd
from io import StringIO
import os
import time
from datetime import datetime, timedelta

# ── USER CONFIGURATION ──────────────────────────────────────────────────────
# Set date range (inclusive). Leave both as None to backfill all dates.
start_date = None
end_date   = None

# Set specific teams, or None for all teams.  e.g. ["BOS", "NYY"]
filter_teams = None

# Produce an Excel report of all changes? "yes" or "no"
produce_report = "no"

# Also re-sync recent games' PlateZ to the current feed? "yes" or "no".
# Statcast reprocesses plate_z weeks after games (esp. the early-season vertical
# recalibration); this catches up the last RESYNC_PLATEZ_DAYS of games. Off by
# default — flip to "yes" (or pass --resync-platez yes) when you want it.
resync_platez = "no"
resync_platez_days = 200

# Also run the MLB Stats API feed pass over ROC/AAA? "yes" or "no". This fills
# Outs and Runners from the feed. It is separate from (and complementary to)
# the Savant minor-league supplement pass, which the main loop now runs for
# ROC/AAA to fill MILB_SUPPLEMENT_COLS — see the MiLB block below. Fill-only.
backfill_milb = "yes"
# ─────────────────────────────────────────────────────────────────────────────

# Six 2026 per-division workbooks — single-homed in pipeline.fetch so the
# audit layer and the pipeline can never point at different books. main()
# opens each book and walks its team tabs; NLE2026 also carries ROC/AAA/FCL.
from pipeline.fetch import DIVISION_WORKBOOK_IDS as SPREADSHEET_IDS

# Spreadsheet column name -> Statcast CSV column name
SUPPLEMENT_MAP = {
    'ArmAngle': 'arm_angle',
    'BatSpeed': 'bat_speed',
    'SwingLength': 'swing_length',
    'AttackAngle': 'attack_angle',
    'AttackDirection': 'attack_direction',
    'SwingPathTilt': 'swing_path_tilt',
    'RunExp': 'delta_pitcher_run_exp',
    'xBA': 'estimated_ba_using_speedangle',
    'xSLG': 'estimated_slg_using_speedangle',
    'xwOBA': 'estimated_woba_using_speedangle',
    'Outs': 'outs_when_up',
    'Event': 'events',
    'Barrel': 'launch_speed_angle',
}

# Swing-tracking cluster: if BatSpeed is missing or sub-50, the entire
# cluster is treated as invalid and dropped together (matches Pitcher2026).
SWING_CLUSTER_COLS = {'BatSpeed', 'SwingLength', 'AttackAngle',
                      'AttackDirection', 'SwingPathTilt'}

# Columns that store raw integer values from Statcast (no rounding needed)
INT_COLS = {'Outs', 'Barrel'}  # Raw integer values (Barrel = launch_speed_angle 1-6)

# Columns that store free-form strings (no numeric coercion, custom translator).
STRING_COLS = {'Event'}

# Columns where official Statcast data should always overwrite estimates
# (even if the cell already has a value from the initial download). Barrel is
# included because Pitcher2026 seeds it with the code_barrel estimate (6 or
# blank), which the official launch_speed_angle (1-6) should replace.
ALWAYS_OVERWRITE_COLS = {'ArmAngle', 'Barrel'}

# Columns that only ever OVERWRITE existing values; they are never used to
# fill a blank cell. Intended for scoring-change corrections (e.g., official
# scorer flips a play from hit to error), where the initial download already
# populated the cell via the MLB Stats API feed.
OVERWRITE_ONLY_COLS = {'Event'}

# Statcast `events` code -> MLB Stats API event string (the format Wally's
# sheet already stores, produced by Pitcher2026.py via play.result.event).
# Only scoring-change-relevant codes are mapped. Statcast's generic
# `field_out` is intentionally OMITTED: MLB Stats API keeps Groundout /
# Flyout / Lineout / Pop Out as distinct events, and we have no way to
# disambiguate from Statcast alone. A missing mapping means "skip; do not
# overwrite the existing sheet value."
STATCAST_TO_MLB_EVENT = {
    'single': 'Single',
    'double': 'Double',
    'triple': 'Triple',
    'home_run': 'Home Run',
    'strikeout': 'Strikeout',
    'strikeout_double_play': 'Strikeout Double Play',
    'walk': 'Walk',
    'intent_walk': 'Intent Walk',
    'hit_by_pitch': 'Hit By Pitch',
    'sac_fly': 'Sac Fly',
    'sac_fly_double_play': 'Sac Fly Double Play',
    'sac_bunt': 'Sac Bunt',
    'sac_bunt_double_play': 'Sac Bunt Double Play',
    'catcher_interf': 'Catcher Interference',
    'field_error': 'Field Error',
    'fielders_choice': 'Fielders Choice',
    'fielders_choice_out': 'Fielders Choice Out',
    'grounded_into_double_play': 'Grounded Into DP',
    'double_play': 'Double Play',
    'triple_play': 'Triple Play',
    'force_out': 'Forceout',
}

# Number format pinned on every numeric supplement column we write.
#
# These columns used to be written with value_input_option='RAW', which stores
# "53.6" as the STRING "53.6". The values read back fine (safe_float parses
# them, so the website was never wrong), but Sheets sorts them lexicographically
# — "9.9" above "74.8" — which makes the column useless to sort or filter in the
# spreadsheet itself. Writing USER_ENTERED stores real numbers and fixes that.
#
# The pinned format is what makes the switch invisible to everything else: this
# script decides "already filled" and "identical to existing" by string compare
# against get_all_values(), which returns the FORMATTED value. Pinning ArmAngle
# to 0.0 keeps 41 reading back as "41.0", and RunExp to 0.000 keeps 0 reading
# back as "0.000" — exactly the strings RAW stored. Without the pin, Automatic
# format would render those as "41" and "0" and every later run would see a
# mismatch and phantom-overwrite the cell.
#
# Single-homed in scrapers/sheet_precision.py as of 2026-08-17, because
# backfill_full.py writes the same columns. Two copies of a depth meant the two
# scripts overwrote each other's cells on every run: one wrote 53.6, the other
# 53.600, and each read the other's value as a change forever.
#
# Event is a free-form string (STRING_COLS), so it has no format entry.
from scrapers.sheet_precision import NUMBER_FORMATS as SUPPLEMENT_NUMBER_FORMATS


def pin_supplement_formats(ws, header, cols):
    """Pin SUPPLEMENT_NUMBER_FORMATS on the full data range of `cols`.

    Applied to the whole column rather than just the rows written, so it is
    idempotent and also corrects historical rows whose format was left
    Automatic. One batch_update per tab.
    """
    reqs = []
    for name in cols:
        fmt = SUPPLEMENT_NUMBER_FORMATS.get(name)
        if fmt is None or name not in header:
            continue
        idx0 = header.index(name)
        reqs.append({'repeatCell': {
            'range': {'sheetId': ws.id, 'startRowIndex': 1,
                      'startColumnIndex': idx0, 'endColumnIndex': idx0 + 1},
            'cell': {'userEnteredFormat': {'numberFormat': fmt}},
            'fields': 'userEnteredFormat.numberFormat'}})
    if not reqs:
        return
    _retry_sheets_call(lambda: ws.spreadsheet.batch_update({'requests': reqs}),
                       'number-format pin')


# Per-column rounding, single-homed in scrapers/sheet_precision.py alongside the
# number formats. Default is 1 decimal for anything not listed.
#
# AttackAngle, AttackDirection and SwingPathTilt moved from 1 decimal to 3 on
# 2026-08-17. Savant serves those three as raw doubles (13 to 17 decimals) while
# bat_speed and swing_length come back at a clean 1, so rounding all five alike
# was throwing away real digits on three of them. See sheet_precision.py for the
# measurement and for why 3 is a convention rather than a measured optimum.
from scrapers.sheet_precision import PRECISION as ROUND_DECIMALS

# Team abbreviation mapping: spreadsheet tab name -> Statcast Search abbreviation
STATCAST_TEAM_MAP = {
    'ATH': 'OAK',
    'KCR': 'KC',
    'SDP': 'SD',
    'SFG': 'SF',
    'TBR': 'TB',
}

MLB_TEAMS = {
    'ARI', 'ATH', 'ATL', 'BAL', 'BOS', 'CHC', 'CIN', 'CLE', 'COL', 'CWS',
    'DET', 'HOU', 'KCR', 'LAA', 'LAD', 'MIA', 'MIL', 'MIN', 'NYM', 'NYY',
    'PHI', 'PIT', 'SDP', 'SEA', 'SFG', 'STL', 'TBR', 'TEX', 'TOR', 'WSH',
}
# ── MiLB (Savant minor-league Statcast Search) ───────────────────────────────
# Savant serves the affiliate levels from a SEPARATE endpoint
# (statcast-search-minors), which Pitcher2026 already uses in player_id mode.
# Verified 2026-07-25 against 7 ROC games: arm_angle, release_pos_y and
# delta_pitcher_run_exp come back on 99.6% of pitches, and the batted-ball
# fields (launch_speed_angle, bb_type, hc_x/hc_y, expected stats) on ~100% of
# balls in play. Arm angle is the same measurement as MLB's, not an estimate:
# across 138 pitchers who threw at both levels in 2026, MLB vs MiLB season
# means correlate r=0.991 with a 1.4 deg mean absolute difference.
#
# Bat tracking was 100% null for MiLB when this was verified (2026-07-25).
# As of 2026-08-21 Savant serves bat_speed on ROC EVENT pitches (balls in play
# and strikeouts); those were applied to the AAA tab by hand from a Savant
# search export. The minors CSV path here still does not write the cluster.
#
# Two quirks drive the code below:
#   1. `team` is honoured only as the NUMERIC MiLB club id, not the three-letter
#      code: team=ROC returns 0 rows, team=534 returns Rochester's. Verified
#      2026-07-26. Because it filters on the queried side, one request covers
#      only half of a ROC game — player_type=pitcher gives ROC's own pitchers,
#      player_type=batter gives the pitches thrown TO ROC batters, i.e. the
#      opponent's. Both are needed, and their union is exactly the set the old
#      client-side home_team/away_team filter produced: on 2026-05-29..31,
#      477 + 482 = 959 keys, identical to the 959 filtered out of an unfiltered
#      18,486-row pull.
#   2. Responses are hard-capped at 25,000 rows with no error, so a range must
#      be requested in windows or it is silently truncated.
#
# Server-side filtering is worth roughly 25x: the unfiltered query returns every
# affiliate level (ROC is ~5% of it), so the old path moved ~470 MB across 41
# windows to keep ~29k rows. The full season now costs 18.5 MB.
MILB_TEAMS = {'ROC', 'AAA'}

# Tab -> club code as it appears in the minors CSV's home_team/away_team.
# Both tabs cover the same Rochester games (ROC tab = ROC pitchers, AAA tab =
# the opposing pitchers in those games), so both filter on the same club.
MILB_SAVANT_TEAM = {'ROC': 'ROC', 'AAA': 'ROC'}

# Tab -> numeric MiLB club id for the server-side `team` filter (quirk 1 above).
# Rochester Red Wings = 534. Both tabs are the same Rochester games.
MILB_SAVANT_TEAM_ID = {'ROC': 534, 'AAA': 534}

# Both sides of every ROC game. See quirk 1: neither alone is complete.
MILB_PLAYER_TYPES = ('pitcher', 'batter')

# Which supplement columns the MiLB path may write. Everything here was
# verified present in the minors search on real ROC games (2026-07-25):
#   ArmAngle 99.6% of pitches | RunExp 99.6% of pitches
#   Barrel (official launch_speed_angle 1-6) 100% of balls in play
#   xwOBA 98.3% of in-play plus every K/BB/HBP | xBA/xSLG 96% of in-play
# Deliberately EXCLUDED:
#   - bat tracking (BatSpeed/SwingLength/Attack*/SwingPathTilt): not served
#     on this CSV; BatSpeed on the AAA tab came from a Savant search export.
#   - Outs/Runners/Description: already 100% from the MLB Stats API feed.
#   - ExitVelo/LaunchAngle/BBType/HC_X/HC_Y: already complete for balls in
#     play. Savant additionally carries EV/LA on FOULS, which the feed does
#     not; filling from here would silently redefine ExitVelo from "in play"
#     to "in play + fouls" and shift every EV-based metric.
MILB_SUPPLEMENT_COLS = {'ArmAngle', 'RunExp', 'xBA', 'xSLG', 'xwOBA', 'Barrel'}

# Days per minors request. Was 3 when every request dragged back all four
# affiliate levels (~5.5k rows/day vs the 25k cap). With the server-side club
# filter a request carries only ROC's own pitches, ~120/day/side, so 30 days
# runs ~3.6k rows and keeps a ~7x margin under the cap even in the densest
# stretch. The cap check below is still the thing that enforces this.
MILB_CHUNK_DAYS = 30
MILB_ROW_CAP = 25000

ALL_TRACKED_TEAMS = set(MLB_TEAMS) | MILB_TEAMS


def date_in_range(date_str):
    """Check if a date string falls within the configured range (inclusive)."""
    if start_date is None and end_date is None:
        return True
    if start_date and date_str < start_date:
        return False
    if end_date and date_str > end_date:
        return False
    return True


# requests' `timeout` is per socket operation, not a budget for the request: the
# clock restarts every time any bytes arrive, so a server dribbling one chunk
# per read interval keeps a GET alive forever. These responses are megabytes
# delivered over hundreds of chunks, which is exactly the shape that can stall
# without ever tripping it. SAVANT_TOTAL_TIMEOUT is the real ceiling, enforced
# below against the wall clock; the other two only bound individual operations.
# The largest response observed is ~11.5 MB in ~25s, so 180s is ~7x headroom.
SAVANT_CONNECT_TIMEOUT = 15
SAVANT_READ_TIMEOUT = 60
SAVANT_TOTAL_TIMEOUT = 180


def _get_savant_body(url, params, session):
    """GET a Savant CSV under a hard wall-clock deadline.

    Streams the body so elapsed time can be checked between chunks, and raises
    requests.exceptions.Timeout — the same type a per-socket timeout raises, so
    the caller's retry arm already handles it — once SAVANT_TOTAL_TIMEOUT is
    exceeded. Returns (status_code, text); text is '' for non-200 responses.
    """
    started = time.monotonic()
    with session.get(url, params=params, stream=True,
                     timeout=(SAVANT_CONNECT_TIMEOUT, SAVANT_READ_TIMEOUT)) as response:
        if response.status_code != 200:
            return response.status_code, ''
        body = bytearray()
        for chunk in response.iter_content(chunk_size=1 << 16):
            body.extend(chunk)
            elapsed = time.monotonic() - started
            if elapsed > SAVANT_TOTAL_TIMEOUT:
                raise requests.exceptions.Timeout(
                    f"exceeded {SAVANT_TOTAL_TIMEOUT}s total after "
                    f"{len(body)} bytes")
        # Matches response.text: requests decodes with the header charset and
        # errors='replace', falling back to a sniff when the header omits one.
        encoding = response.encoding or response.apparent_encoding or 'utf-8'
        return response.status_code, str(bytes(body), encoding, errors='replace')


def _fetch_savant_csv(url, params, session, label):
    """GET a Savant CSV endpoint and parse it, retrying transient failures.

    Retries 5xx, timeouts and connection resets, and also malformed bodies:
    Savant intermittently returns an HTML error page with a 200, which pandas
    surfaces as a ParserError rather than an HTTP failure. Returns a DataFrame,
    or None when the query legitimately has no rows.
    """
    for attempt in range(4):
        try:
            status_code, csv_text = _get_savant_body(url, params, session)
            if status_code != 200:
                if status_code >= 500 and attempt < 3:
                    wait = 5 * (2 ** attempt)  # 5, 10, 20 s
                    print(f"    Savant returned {status_code} for {label}, "
                          f"retrying in {wait}s...")
                    time.sleep(wait)
                    continue
                print(f"    Statcast returned status {status_code}")
                return None

            if not csv_text or csv_text.strip() == '' or 'No Results' in csv_text[:100]:
                print(f"    No Statcast data available yet ({label})")
                return None

            df = pd.read_csv(StringIO(csv_text), low_memory=False)
            if df.empty:
                print(f"    Empty DataFrame ({label})")
                return None
            return df

        except (requests.exceptions.Timeout,
                requests.exceptions.ConnectionError,
                requests.exceptions.ChunkedEncodingError,
                pd.errors.ParserError) as e:
            if attempt < 3:
                wait = 5 * (2 ** attempt)
                print(f"    Savant request failed for {label} "
                      f"({type(e).__name__}), retrying in {wait}s...")
                time.sleep(wait)
                continue
            print(f"    Giving up on {label} ({type(e).__name__})")
            return None
    return None


def _date_windows(date_min, date_max, span_days):
    """Split an inclusive date range into consecutive windows of span_days."""
    d = datetime.strptime(date_min, '%Y-%m-%d').date()
    end = datetime.strptime(date_max, '%Y-%m-%d').date()
    out = []
    while d <= end:
        stop = min(d + timedelta(days=span_days - 1), end)
        out.append((d.isoformat(), stop.isoformat()))
        d = stop + timedelta(days=1)
    return out


# (club, player_type, window_start, window_end) -> club-filtered DataFrame (or
# None if the request failed). Lives for the process so a multi-tab run
# downloads once: ROC and AAA are the same Rochester games from opposite sides.
_MINORS_WINDOW_CACHE = {}


def _download_statcast_minors(team_tab, date_min, date_max, session):
    """Fetch the minor-league Statcast Search for one affiliate's games.

    Requests are filtered server-side by numeric club id, once per player_type
    so both sides of each ROC game come back (see the MiLB block above). The
    response is capped at MILB_ROW_CAP rows with no error, so the range is still
    windowed and a capped window is refused rather than passed off as complete.
    """
    club = MILB_SAVANT_TEAM.get(team_tab, team_tab)
    club_id = MILB_SAVANT_TEAM_ID.get(team_tab)
    if club_id is None:
        # Config error, not a runtime condition: without an id every request
        # would go out as team=None and come back empty, which reads like a
        # season with no games rather than a missing mapping.
        raise KeyError(f"{team_tab} is in MILB_TEAMS but has no "
                       f"MILB_SAVANT_TEAM_ID entry; add its numeric club id")
    windows = _date_windows(date_min, date_max, MILB_CHUNK_DAYS)
    reqs = [(w, pt) for w in windows for pt in MILB_PLAYER_TYPES]
    print(f"    Downloading minor-league Statcast for {team_tab} (club {club}) "
          f"{date_min} to {date_max} in {len(windows)} windows "
          f"x {len(MILB_PLAYER_TYPES)} sides...")

    url = "https://baseballsavant.mlb.com/statcast-search-minors/csv"
    frames = []
    failed = []
    kept_rows = 0
    for r_i, ((w_start, w_end), player_type) in enumerate(reqs, 1):
        cache_key = (club, player_type, w_start, w_end)
        if cache_key in _MINORS_WINDOW_CACHE:
            cached = _MINORS_WINDOW_CACHE[cache_key]
            if cached is not None and len(cached):
                frames.append(cached)
                kept_rows += len(cached)
            continue
        params = {
            'all': 'true',
            'type': 'details',
            'game_date_gt': w_start,
            'game_date_lt': w_end,
            'team': str(club_id),
            'player_type': player_type,
            'min_pitches': '0',
            'min_results': '0',
            'sort_col': 'pitches',
            'sort_order': 'desc',
            'minors': 'true',
        }
        label = f"{team_tab} {w_start}..{w_end} {player_type}"
        df = _fetch_savant_csv(url, params, session, label)
        if df is None:
            # None is ambiguous here: a legitimately empty window (off days) and
            # a window whose retries were exhausted both land on it. Only the
            # latter leaves a hole, and _fetch_savant_csv has already said which
            # it was — record it so the end-of-download summary can flag it.
            _MINORS_WINDOW_CACHE[cache_key] = None
            failed.append(label)
            continue
        if len(df) >= MILB_ROW_CAP:
            # Never let a capped window pass as complete data.
            print(f"    WARNING: {label} returned {len(df)} rows "
                  f"(cap {MILB_ROW_CAP}) — results were truncated. Lower "
                  f"MILB_CHUNK_DAYS and re-run; this window is being skipped.")
            failed.append(f"{label} (truncated)")
            continue
        if 'home_team' not in df.columns or 'away_team' not in df.columns:
            print(f"    WARNING: {label} has no home/away columns; skipping")
            failed.append(f"{label} (no team cols)")
            continue
        # Redundant while the server-side filter holds, and that is the point:
        # `team` is undocumented, and an id Savant stops honouring goes back to
        # returning every affiliate level rather than erroring. This keeps a
        # silently-ignored filter from widening the result set.
        keep = df[(df['home_team'] == club) | (df['away_team'] == club)]
        if len(keep) != len(df):
            print(f"    WARNING: {label} returned {len(df) - len(keep)} non-{club} "
                  f"rows — the server-side team={club_id} filter was ignored. "
                  f"Falling back to the client-side filter for this request.")
        _MINORS_WINDOW_CACHE[cache_key] = keep
        if len(keep):
            frames.append(keep)
            kept_rows += len(keep)
        # A long run of silent requests reads as a hang, especially right after
        # a 502-retry line. Say where we are on every one.
        print(f"      [{r_i}/{len(reqs)}] {w_start}..{w_end} {player_type}: "
              f"{len(keep)} {club} (running {kept_rows})", flush=True)
        time.sleep(1.5)  # be polite to Savant

    if failed:
        print(f"    NOTE: {len(failed)} of {len(reqs)} requests returned no "
              f"data (off days, or exhausted retries — see the messages above): "
              f"{', '.join(failed)}")
    if not frames:
        print(f"    No {club} rows found in the minor-league search")
        return None
    # The two sides are disjoint (a pitch is thrown either by ROC or to ROC),
    # but dedupe on the PitchID key rather than assume it.
    out = pd.concat(frames, ignore_index=True)
    return out.drop_duplicates(['game_pk', 'at_bat_number', 'pitch_number'],
                               ignore_index=True)


def download_statcast(team_tab, date_min, date_max, session):
    """Download Statcast Search CSV for a team and date range.
    Returns a dict keyed by (game_pk, at_bat_number, pitch_number) -> row dict.

    MLB tabs use the standard Statcast Search; ROC/AAA use the minor-league
    search, which needs windowed requests and a client-side club filter.
    """
    is_milb = team_tab.upper() in MILB_TEAMS
    allowed_cols = MILB_SUPPLEMENT_COLS if is_milb else None

    if is_milb:
        df = _download_statcast_minors(team_tab, date_min, date_max, session)
    else:
        statcast_team = STATCAST_TEAM_MAP.get(team_tab, team_tab)
        print(f"    Downloading Statcast for {team_tab} ({statcast_team}) "
              f"{date_min} to {date_max}...")
        df = _fetch_savant_csv(
            "https://baseballsavant.mlb.com/statcast_search/csv",
            {
                'all': 'true',
                'type': 'details',
                'game_date_gt': date_min,
                'game_date_lt': date_max,
                'team': statcast_team,
                'player_type': 'pitcher',
                'min_pitches': '0',
                'min_results': '0',
                'sort_col': 'pitches',
                'sort_order': 'desc',
            },
            session, team_tab)

    if df is None:
        return None

    try:
        # Verify merge keys exist
        for col in ['game_pk', 'at_bat_number', 'pitch_number']:
            if col not in df.columns:
                print(f"    Missing merge key: {col}")
                return None

        # Build lookup dict keyed by PitchID components
        # Use vectorised pandas ops where possible, then convert to dict
        statcast_cols = list(SUPPLEMENT_MAP.values())
        available = [c for c in statcast_cols if c in df.columns]

        # ---- Align Savant pitch_number with the MLB feed numbering ----
        # Savant counts automatic balls/strikes (pitch-timer violations,
        # intentional walks) as pitch_number entries; the MLB Stats API feed —
        # and therefore Wally's PitchID — counts only real pitches. Left as-is,
        # every pitch AFTER an auto ball keys to the wrong Savant row, so its
        # supplement fields (xwOBA, RunExp, etc.) get filled from the neighbour.
        # Fix: within each plate appearance, drop the automatic rows and
        # renumber the real pitches 1..N so the key matches the sheet.
        if 'description' in df.columns:
            is_auto = df['description'].astype(str).str.lower().str.contains('automatic', na=False)
        else:
            is_auto = pd.Series(False, index=df.index)
        _order = df.sort_values(['game_pk', 'at_bat_number', 'pitch_number']).index
        _sdf = df.loc[_order]
        feed_num = (~is_auto.loc[_order]).groupby(
            [_sdf['game_pk'], _sdf['at_bat_number']]).cumsum().reindex(df.index)

        # Build string keys once (vectorised). k2 uses the feed-aligned number.
        keys_df = pd.DataFrame({
            'k0': df['game_pk'].astype(int).astype(str),
            'k1': df['at_bat_number'].astype(int).astype(str),
            'k2': feed_num.astype(int).astype(str),
        })

        # Swing-cluster invalidity mask: BatSpeed missing or <50 means the
        # entire swing-tracking frame is unreliable; null all members together.
        if 'bat_speed' in df.columns:
            bs_numeric = pd.to_numeric(df['bat_speed'], errors='coerce')
            swing_invalid = bs_numeric.isna() | (bs_numeric < 50)
        else:
            swing_invalid = pd.Series(False, index=df.index)

        # Pre-format each supplement column into a string Series
        formatted = {}
        for sheet_col, csv_col in SUPPLEMENT_MAP.items():
            if csv_col not in df.columns:
                continue
            # MiLB: only the columns the minors search actually serves and that
            # we've opted into (MILB_SUPPLEMENT_COLS).
            if allowed_cols is not None and sheet_col not in allowed_cols:
                continue
            series = df[csv_col]
            if sheet_col in STRING_COLS:
                # String column: custom translator. For Event, translate Statcast
                # lowercase_underscore codes to MLB Stats API title-case strings
                # that Wally's sheet uses. Unmapped codes (including `field_out`)
                # are dropped so the downstream overwrite step leaves the cell
                # alone.
                if sheet_col == 'Event':
                    mapped = series.map(STATCAST_TO_MLB_EVENT)
                    formatted[sheet_col] = mapped.dropna()
                else:
                    formatted[sheet_col] = series.dropna().astype(str)
            elif sheet_col in INT_COLS:
                # Integer columns: raw int as string, NaN -> None
                s = series.dropna().astype(float).astype(int).astype(str)
                formatted[sheet_col] = s
            else:
                decimals = ROUND_DECIMALS.get(sheet_col, 1)
                numeric = pd.to_numeric(series, errors='coerce')
                if sheet_col in SWING_CLUSTER_COLS:
                    numeric = numeric.where(~swing_invalid)
                rounded = numeric.round(decimals)
                # Format to fixed decimal string; NaN rows excluded via dropna
                fmt_func = (lambda d: lambda v: f"{v:.{d}f}")(decimals)
                s = rounded.dropna().map(fmt_func)
                formatted[sheet_col] = s

        # Runners column (vectorised)
        has_runners = all(c in df.columns for c in ['on_1b', 'on_2b', 'on_3b'])
        if has_runners:
            r1 = df['on_1b'].notna()
            r2 = df['on_2b'].notna()
            r3 = df['on_3b'].notna()
            # Build runners string per row
            runners = pd.Series('0', index=df.index)
            # Assign combinations (most common first for speed)
            mask_any = r1 | r2 | r3
            if mask_any.any():
                parts = []
                for mask, label in [(r1, '1'), (r2, '2'), (r3, '3')]:
                    parts.append(mask.map({True: label, False: ''}))
                runners = (parts[0] + '+' + parts[1] + '+' + parts[2]).str.strip('+').str.replace(r'\++', '+', regex=True)
                runners = runners.replace('', '0')

        # Assemble lookup dict
        lookup = {}
        for i in df.index:
            if is_auto.at[i]:
                continue   # automatic ball/strike: not a real pitch, no key
            key = (keys_df.at[i, 'k0'], keys_df.at[i, 'k1'], keys_df.at[i, 'k2'])
            data = {}
            for sheet_col in formatted:
                if i in formatted[sheet_col].index:
                    data[sheet_col] = formatted[sheet_col].at[i]
                elif sheet_col in ALWAYS_OVERWRITE_COLS:
                    data[sheet_col] = ''
            if has_runners and (allowed_cols is None or 'Runners' in allowed_cols):
                data['Runners'] = runners.at[i]
            if data:
                lookup[key] = data

        print(f"    Got {len(lookup)} pitches with supplement data "
              f"(columns: {available})")
        return lookup

    except requests.exceptions.Timeout:
        print(f"    Timeout downloading Statcast data")
        return None
    except Exception as e:
        print(f"    Error: {e}")
        return None


_TRANSIENT_SHEETS_CODES = ('429', '500', '502', '503', '504')


def _retry_sheets_call(fn, label, max_retries=5):
    """Retry a Sheets API call on rate-limit (429), transient 5xx errors, and
    network-level drops (connection reset, timeout)."""
    for attempt in range(max_retries):
        try:
            return fn()
        except gspread.exceptions.APIError as e:
            msg = str(e)
            code = next((c for c in _TRANSIENT_SHEETS_CODES if c in msg), None)
            if code and attempt < max_retries - 1:
                wait = min(60, 5 * (2 ** attempt))  # 5, 10, 20, 40, 60 s
                kind = 'Rate limited' if code == '429' else f'Transient {code}'
                print(f"    {kind} during {label}, waiting {wait}s before retry "
                      f"({attempt + 1}/{max_retries - 1})...")
                time.sleep(wait)
            else:
                raise
        except (requests.exceptions.ConnectionError,
                requests.exceptions.Timeout,
                requests.exceptions.ChunkedEncodingError) as e:
            # Transient network blip (e.g. ConnectionResetError [Errno 54] from a
            # dropped TLS connection during a write). gspread surfaces these as
            # requests errors, not APIError, so they need their own retry path.
            # The Sheets writes are idempotent (same cells and values), so
            # resending after a reset is safe even if Google applied the lost
            # request.
            if attempt < max_retries - 1:
                wait = min(60, 5 * (2 ** attempt))  # 5, 10, 20, 40, 60 s
                print(f"    Connection error during {label} "
                      f"({type(e).__name__}), waiting {wait}s before retry "
                      f"({attempt + 1}/{max_retries - 1})...")
                time.sleep(wait)
            else:
                raise


def read_sheet_with_retry(ws, max_retries=5):
    return _retry_sheets_call(ws.get_all_values, 'sheet read', max_retries)


def update_cells_with_retry(ws, cells, max_retries=5, **kwargs):
    return _retry_sheets_call(
        lambda: ws.update_cells(cells, **kwargs), 'cell write', max_retries,
    )


def write_report(report_data, output_dir=os.path.join(os.path.expanduser('~'), 'Downloads', '')):
    """Write an Excel report with one tab per team showing all changed rows.
    Changed cells are bold; column header row shows which columns had changes."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    bold = Font(bold=True)
    teams_with_data = sorted(t for t, rows in report_data.items() if rows)

    if not teams_with_data:
        print("  No changes to report.")
        return None

    for team in teams_with_data:
        ws = wb.create_sheet(title=team)
        entries = report_data[team]
        header = entries[0]['header']

        # Write header row (bold)
        for c, col_name in enumerate(header, start=1):
            cell = ws.cell(row=1, column=c, value=col_name)
            cell.font = bold

        # Write data rows
        for r, entry in enumerate(entries, start=2):
            row_vals = entry['row_values']
            changes = entry['changes']  # {col_idx: 'new'|'overwrite'}
            for c, val in enumerate(row_vals):
                cell = ws.cell(row=r, column=c + 1, value=val)
                if c in changes:
                    cell.font = bold

    stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    path = os.path.join(output_dir, f'backfill_report_{stamp}.xlsx')
    wb.save(path)
    return path


def main():
    print(f"Date range: {start_date or '(all)'} to {end_date or '(all)'}")
    if filter_teams:
        print(f"Teams: {', '.join(sorted(filter_teams))}")

    # Default gspread service account (~/.config/gspread/service_account.json =
    # huronalytics), the writer on all six division books and the same account
    # the append path (sheets_append.py) uses. The old repo-local
    # service_account.json was the st-leaderboard reader, which has neither
    # access to nor write permission on the new books.
    gc = gspread.service_account()

    session = requests.Session()
    session.headers.update({
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)',
        'Accept': 'text/csv',
    })

    total_filled = 0
    total_overwritten = 0
    report_data = {}  # team -> list of {header, row_values, changes}

    for sheet_label, sheet_id in SPREADSHEET_IDS.items():
        # Metadata calls 503 sporadically too (open_by_key killed a CI
        # leaderboard run on 2026-07-13) — same retry as reads/writes.
        sh = _retry_sheets_call(lambda: gc.open_by_key(sheet_id), 'workbook open')
        print(f"\n{'='*60}")
        print(f"[{sheet_label}] {sh.title}")
        print(f"{'='*60}")

        for i, ws in enumerate(_retry_sheets_call(sh.worksheets, 'tab list')):
            tab_name = ws.title.upper()

            # Skip WBC and non-tracked tabs
            if tab_name not in ALL_TRACKED_TEAMS:
                continue
            if filter_teams and tab_name not in filter_teams:
                continue

            print(f"\n[{ws.title}]")
            if i > 0:
                time.sleep(1.5)

            rows = read_sheet_with_retry(ws)
            if not rows or len(rows) < 2:
                print(f"  Empty sheet")
                continue

            header = rows[0]
            col_idx = {name: j for j, name in enumerate(header) if name}

            # Verify required columns exist
            pitch_id_col = col_idx.get('PitchID')
            if pitch_id_col is None:
                print(f"  No PitchID column — skipping")
                continue

            # Find supplement column indices (SUPPLEMENT_MAP + Runners).
            # ROC/AAA are limited to MILB_SUPPLEMENT_COLS — the minors search
            # serves no bat tracking at all, and the columns it does serve
            # beyond that set are opt-in (see MILB_SUPPLEMENT_COLS).
            is_milb_tab = tab_name in MILB_TEAMS
            allowed = MILB_SUPPLEMENT_COLS if is_milb_tab else None
            supp_col_idx = {}
            for sheet_col in SUPPLEMENT_MAP:
                if allowed is not None and sheet_col not in allowed:
                    continue
                if sheet_col in col_idx:
                    supp_col_idx[sheet_col] = col_idx[sheet_col]
            if 'Runners' in col_idx and (allowed is None or 'Runners' in allowed):
                supp_col_idx['Runners'] = col_idx['Runners']
            if not supp_col_idx:
                print(f"  No supplement columns found — skipping")
                continue

            # Find rows that need filling:
            # PitchID exists AND (at least one supplement column is empty
            # OR has an always-overwrite column that might contain an estimate)
            needs_fill = []  # (row_index_1based, pitch_id, cols_to_update)
            # cols_to_update entries: (sheet_col, existing_value) — empty string means new fill
            game_dates = set()
            date_col = col_idx.get('Game Date')

            for r_idx, row in enumerate(rows[1:], start=2):
                pid = row[pitch_id_col] if pitch_id_col < len(row) else ''
                if not pid or '_' not in pid:
                    continue

                # Apply date range filter for supplement backfill
                if date_col is not None:
                    gd = row[date_col] if date_col < len(row) else ''
                    if not date_in_range(gd):
                        continue

                # Check which supplement columns need updating:
                # - Empty columns need filling (but NOT for OVERWRITE_ONLY_COLS)
                # - ALWAYS_OVERWRITE_COLS need updating even if they have a value
                #   (the existing value may be an estimate that official data should replace)
                # - OVERWRITE_ONLY_COLS update existing values only (for scoring
                #   corrections like hit↔error); never used to fill a blank cell.
                cols_to_update = []
                for sheet_col, c_idx in supp_col_idx.items():
                    val = row[c_idx] if c_idx < len(row) else ''
                    is_empty = (val == '' or val is None)
                    if is_empty and sheet_col in OVERWRITE_ONLY_COLS:
                        continue
                    if is_empty:
                        cols_to_update.append((sheet_col, ''))
                    elif sheet_col in ALWAYS_OVERWRITE_COLS or sheet_col in OVERWRITE_ONLY_COLS:
                        cols_to_update.append((sheet_col, val))

                if cols_to_update:
                    needs_fill.append((r_idx, pid, cols_to_update))
                    if date_col is not None:
                        gd = row[date_col] if date_col < len(row) else ''
                        if gd:
                            game_dates.add(gd)

            if not needs_fill:
                print(f"  All rows filled — nothing to do")
                continue

            print(f"  {len(needs_fill)} rows need supplement data "
                  f"(dates: {sorted(game_dates)})")

            # Download Statcast data for this team's date range
            if not game_dates:
                print(f"  No game dates found — skipping")
                continue

            date_min = min(game_dates)
            date_max = max(game_dates)

            time.sleep(3)  # Be polite to Baseball Savant
            lookup = download_statcast(ws.title, date_min, date_max, session)

            if lookup is None:
                print(f"  No Statcast data available — skipping")
                continue

            # Match and prepare cell updates
            cells_to_update = []
            new_fill_cells = 0
            overwrite_cells = 0
            team_report_rows = []

            for r_idx, pid, cols_to_update in needs_fill:
                # Split PitchID: game_pk_atbat(zero-padded)_pitch(zero-padded)
                # Strip padding to match Statcast lookup keys (unpadded ints)
                parts = pid.split('_')
                if len(parts) != 3:
                    continue
                key = (parts[0], str(int(parts[1])), str(int(parts[2])))

                statcast_row = lookup.get(key, {})
                row_changes = {}  # col_idx -> 'new'|'overwrite'

                for sheet_col, existing_val in cols_to_update:
                    if sheet_col in statcast_row:
                        val = statcast_row[sheet_col]
                        # Don't write empty values for overwrite cols —
                        # that would erase data when official data isn't ready yet
                        if not val and (sheet_col in ALWAYS_OVERWRITE_COLS
                                        or sheet_col in OVERWRITE_ONLY_COLS):
                            continue
                        # Skip if overwrite value is identical to existing
                        if existing_val and str(val) == str(existing_val):
                            continue
                        cell = gspread.Cell(
                            row=r_idx,
                            col=supp_col_idx[sheet_col] + 1,  # 1-indexed
                            value=val,
                        )
                        cells_to_update.append(cell)
                        c_idx = col_idx[sheet_col]
                        if existing_val:
                            overwrite_cells += 1
                            row_changes[c_idx] = 'overwrite'
                        else:
                            new_fill_cells += 1
                            row_changes[c_idx] = 'new'

                # Collect report data for this row if anything changed
                if row_changes and produce_report == 'yes':
                    # Build row with new values applied
                    row_vals = list(rows[r_idx - 1])
                    for c_idx, change_type in row_changes.items():
                        # Find the cell we're about to write for this column
                        for cell in cells_to_update:
                            if cell.row == r_idx and cell.col == c_idx + 1:
                                row_vals[c_idx] = cell.value
                                break
                    team_report_rows.append({
                        'header': header,
                        'row_values': row_vals,
                        'changes': row_changes,
                    })

            if cells_to_update:
                print(f"  Writing {len(cells_to_update)} cells "
                      f"({new_fill_cells} new, {overwrite_cells} overwritten)...")
                # USER_ENTERED so numeric supplement columns land as real
                # numbers and sort correctly in the sheet; the format pin below
                # keeps the read-back strings byte-identical to what RAW wrote.
                update_cells_with_retry(ws, cells_to_update,
                                        value_input_option='USER_ENTERED')
                try:
                    pin_supplement_formats(ws, header, supp_col_idx.keys())
                except Exception as _e:
                    print(f"    WARNING: number-format pin failed "
                          f"({type(_e).__name__}: {_e}); values are still "
                          f"correct, run scripts/ops/fix_text_typed_supplements.py")
                total_filled += new_fill_cells
                total_overwritten += overwrite_cells
                time.sleep(2)  # Rate limit buffer after write
            else:
                print(f"  No new data to fill, no overwrites changed.")

            if team_report_rows:
                report_data[tab_name] = team_report_rows

    parts = []
    if total_filled:
        parts.append(f"{total_filled} new cells filled")
    if total_overwritten:
        parts.append(f"{total_overwritten} cells overwritten")
    if parts:
        print(f"\nDone. {', '.join(parts)}.")
    else:
        print(f"\nDone. No new data added, no data overwritten.")

    if produce_report == 'yes':
        report_path = write_report(report_data)
        if report_path:
            print(f"Report saved to: {report_path}")

    if resync_platez == 'yes':
        print(f"\n{'='*60}\n[PlateZ re-sync] catching up recent games to the current feed\n{'='*60}")
        import importlib.util
        _rp = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                           'scripts', 'audits', 'resync_recent_platez.py')
        _spec = importlib.util.spec_from_file_location('resync_recent_platez', _rp)
        _mod = importlib.util.module_from_spec(_spec); _spec.loader.exec_module(_mod)
        _mod.resync(gc, days=resync_platez_days, apply=True)  # reuses the write-capable client

    if backfill_milb == 'yes':
        print(f"\n{'='*60}\n[MiLB backfill] Outs + Runners for ROC/AAA (feed)\n{'='*60}")
        import importlib.util
        _mp = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                           'scripts', 'builders', 'backfill_milb_feed.py')
        _spec = importlib.util.spec_from_file_location('backfill_milb_feed', _mp)
        _mod = importlib.util.module_from_spec(_spec); _spec.loader.exec_module(_mod)
        _mod.run(gc, apply=True)   # fill-only; reuses the write-capable client


if __name__ == '__main__':
    # ── CLI overrides (optional — edit start_date/end_date at top of file as before) ──
    parser = argparse.ArgumentParser(description='Backfill supplemental Statcast data into Google Sheet')
    parser.add_argument('--start', default=None, help='Start date YYYY-MM-DD, or "none" for all dates')
    parser.add_argument('--end', default=None, help='End date YYYY-MM-DD, or "none" for all dates')
    parser.add_argument('--teams', default=None, help='Comma-separated team abbreviations (e.g., BOS,NYY)')
    parser.add_argument('--report', default=None, help='"yes" to produce an Excel report of changes')
    parser.add_argument('--resync-platez', default=None,
                        help='"yes" to also re-sync recent games\' PlateZ to the current feed')
    parser.add_argument('--backfill-milb', default=None,
                        help='"yes"/"no" to backfill ROC/AAA Outs+Runners from the feed (default yes)')
    args = parser.parse_args()

    # Only override module-level globals if CLI args were explicitly passed
    if args.start is not None:
        start_date = None if args.start.lower() == 'none' else args.start
    if args.end is not None:
        end_date = None if args.end.lower() == 'none' else args.end
    if args.teams is not None:
        filter_teams = [t.strip().upper() for t in args.teams.split(',') if t.strip()]
    if args.report is not None:
        produce_report = args.report.lower()
    if args.resync_platez is not None:
        resync_platez = args.resync_platez.lower()
    if args.backfill_milb is not None:
        backfill_milb = args.backfill_milb.lower()

    main()
