#!/usr/bin/env python3
"""
Full-column backfill: reconcile every Google Sheet column against its source.

`backfill_supplement.py` fills the 12 Savant-served columns and nothing else.
This module reconciles all 43 in-scope columns, 31 of which come from the MLB
Stats API game feed and have never been backfillable. It also adds the two
things the supplement path cannot do: it finds pitches that are missing from the
sheet entirely, and it rewrites stored values at full source precision.

Out of scope by instruction (Wally, 2026-08-17): Game Date, PTeam, BTeam,
PitchID, Pitch Type, Stuff+, Loc+, Pitching+.

DRY RUN IS THE DEFAULT. Nothing is written to any sheet without --apply.

    python3 -m scrapers.backfill_full                    # dry run, whole season
    python3 -m scrapers.backfill_full --teams ARI,WSH    # dry run, two tabs
    python3 -m scrapers.backfill_full --apply            # write

Design notes that are not obvious:

  * The feed is parsed by Pitcher2026.download_game_data, not by a second copy
    of the parser here. It is called with raw_precision=True and a disk-cached
    payload. One parser means the backfill can never disagree with the scraper
    that wrote the row in the first place.

  * "Precision" changes and "drift" changes are separated, because they need
    different review. A cell is a precision change only when the stored value is
    exactly what you get by rounding today's source value to the stored value's
    own decimal depth. Anything else is real movement in the source.

  * Blank cells are not all the same thing. Wally deliberately deleted a small
    number of Velocity, Spin Rate, RTilt and movement values that were obvious
    misreads (a 55 mph slip, a 200 rpm curveball). Refilling those with the same
    number would undo the decision, so a blank-fill candidate is suppressed when
    the source still offers the value that was rejected. See DecisionLedger.
"""

# Runnable as a file from any directory (IDE run buttons included):
# put the repo root on sys.path before the intra-repo package imports.
import os as _os, sys as _sys
_ROOT = _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__)))
if _ROOT not in _sys.path:
    _sys.path.insert(0, _ROOT)

import argparse
import collections
import re
import gzip
import json
import math
import os
import time
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP

import gspread
import pandas as pd
import requests

from pipeline.fetch import DIVISION_WORKBOOK_IDS as SPREADSHEET_IDS
from scrapers.backfill_supplement import (
    MLB_TEAMS, MILB_TEAMS, ALL_TRACKED_TEAMS,
    _retry_sheets_call, read_sheet_with_retry, update_cells_with_retry,
    download_statcast,
)
from scrapers.sheet_precision import (
    PRECISION, NUMBER_FORMATS, STRING_COLS, TIME_COLS, TILT_MINUTES,
    fmt, as_float, stored_decimals, tilt_minutes, tilt_gap,
)

DATA = os.path.join(_ROOT, 'data')
FEED_CACHE = os.path.join(DATA, '_feed_cache')
LEDGER_PATH = os.path.join(DATA, 'backfill_decisions.json')
REPORT_DIR = os.path.join(os.path.expanduser('~'), 'Downloads')


# ── Column scope ─────────────────────────────────────────────────────────────
# Wally's exclusion list. Game Date/PTeam/BTeam/PitchID/Pitch Type are identity
# and must never move under a row. The three grade columns are written by
# scripts/ci/sheets_write_grades.py from the models, not by any feed.
EXCLUDED_COLS = {
    'Game Date', 'PTeam', 'BTeam', 'PitchID', 'Pitch Type',
    'Stuff+', 'Loc+', 'Pitching+',
}

# Columns whose authority is the MLB Stats API game feed. Six of these are
# computed from feed values rather than read from it (RTilt, OTilt, VAA, HAA,
# xIndVrtBrk, xHorzBrk) — Pitcher2026 owns that math.
FEED_COLS = [
    'Pitcher', 'Throws', 'Velocity', 'Spin Rate', 'RTilt', 'OTilt',
    'IndVertBrk', 'HorzBrk', 'xIndVrtBrk', 'xHorzBrk',
    'RelPosZ', 'RelPosX', 'Extension', 'PlateZ', 'PlateX', 'SzTop', 'SzBot',
    'VAA', 'HAA', 'Batter', 'Bats', 'Count', 'Outs', 'Description', 'Event',
    'ExitVelo', 'LaunchAngle', 'Distance', 'BBType', 'HC_X', 'HC_Y',
]

# Columns whose authority is Baseball Savant. Unchanged from the supplement path.
SAVANT_COLS = [
    'ArmAngle', 'xBA', 'xSLG', 'xwOBA', 'RunExp', 'Barrel', 'Runners',
    'BatSpeed', 'SwingLength', 'AttackAngle', 'AttackDirection', 'SwingPathTilt',
]

IN_SCOPE = FEED_COLS + SAVANT_COLS
SAVANT_COL_SET = set(SAVANT_COLS)
assert not (set(IN_SCOPE) & EXCLUDED_COLS), 'scope lists disagree'
assert not (set(FEED_COLS) & SAVANT_COL_SET), 'a column has two authorities'

# The precision policy and the formatter live in scrapers/sheet_precision.py so
# this module and backfill_supplement.py cannot disagree about a column's depth.


# ── Position players (EP) ────────────────────────────────────────────────────
# EP is the eephus, and in this sheet it is a clean marker for a position player
# on the mound. Measured 2026-08-17: 1,746 EP pitches from 52 pitchers, and for
# every one of the 52 the EP count equals his TOTAL pitch count, so not one of
# them is a real pitcher with an eephus in his arsenal.
#
# Wally's call 2026-08-17: do not touch the pitch-characteristic columns on these
# rows. Release, movement, spin and approach are not meaningful for a shortstop
# lobbing one in, and a measured value there is noise dressed as data.
#
# Everything else on an EP row IS kept and reviewed: Count, Description, Event,
# and the whole batted-ball block. An EP pitch still has a real count, a real
# result and a real batted ball, and one of ARI's Count corrections sits on one.
#
# PlateZ, PlateX, SzTop and SzBot are deliberately NOT in this set. Location is
# measured the same way whoever throws the pitch, and the zone belongs to the
# hitter, not the pitcher. VAA and HAA ARE in it, because they are derived from
# velocity and movement and inherit their meaninglessness here.
EP_PITCH_TYPE = 'EP'

# THE pitch-metric block. One set, two uses, and they are the same set because
# they are the same idea: these fourteen columns describe how the pitch was
# thrown, and nothing else in the schema does.
#
#   1. They are skipped entirely on EP rows. Release, movement, spin and approach
#      are not meaningful for a shortstop lobbing one in.
#   2. They are the only columns a recommendation may judge against the pitcher's
#      own median for a pitch type. Everything else is per-event or belongs to the
#      batter, so his median says nothing about it — see recommend().
#
# Deliberately absent: PlateZ and PlateX, because a pitcher throws all over the
# zone and his median location is not a yardstick for one pitch; SzTop and SzBot,
# because the zone belongs to the hitter; and the whole batted-ball and
# bat-tracking block, which describes what the hitter did.
PITCH_METRIC_COLS = {
    'Velocity', 'Spin Rate', 'RTilt', 'OTilt',
    'IndVertBrk', 'HorzBrk', 'xIndVrtBrk', 'xHorzBrk',
    'RelPosZ', 'RelPosX', 'Extension', 'ArmAngle',
    'VAA', 'HAA',
}


# ── Feed cache ───────────────────────────────────────────────────────────────
FEED_URL = 'https://statsapi.mlb.com/api/v1.1/game/{}/feed/live'


def feed_path(game_pk):
    return os.path.join(FEED_CACHE, f'{game_pk}.json.gz')


def fetch_game_json(game_pk, session, refresh=False):
    """Return one game's feed payload, from disk when cached.

    Fails closed on every path. A non-200, a short body or unparseable JSON
    raises rather than returning a partial game, because a partial game reads
    downstream as "these pitches do not exist" and would be reported as rows to
    delete or refill. Writes to a temp path and moves, so an interrupted run
    never leaves a truncated cache entry behind.
    """
    path = feed_path(game_pk)
    if os.path.exists(path) and not refresh:
        try:
            with gzip.open(path, 'rt', encoding='utf-8') as f:
                return json.load(f)
        except (OSError, EOFError, json.JSONDecodeError, gzip.BadGzipFile) as e:
            print(f"    cached feed for {game_pk} is unreadable "
                  f"({type(e).__name__}); re-fetching")

    last = None
    for attempt in range(4):
        try:
            r = session.get(FEED_URL.format(game_pk), timeout=(15, 90))
            if r.status_code != 200:
                last = f'HTTP {r.status_code}'
                if r.status_code >= 500 and attempt < 3:
                    time.sleep(5 * (2 ** attempt))
                    continue
                raise RuntimeError(f'feed {game_pk}: {last}')
            data = r.json()
            if not data or 'liveData' not in data:
                raise RuntimeError(f'feed {game_pk}: payload has no liveData')
            os.makedirs(FEED_CACHE, exist_ok=True)
            tmp = path + '.tmp'
            with gzip.open(tmp, 'wt', encoding='utf-8') as f:
                json.dump(data, f)
            os.replace(tmp, path)
            return data
        except (requests.exceptions.Timeout,
                requests.exceptions.ConnectionError,
                requests.exceptions.ChunkedEncodingError,
                json.JSONDecodeError) as e:
            last = f'{type(e).__name__}: {e}'
            if attempt < 3:
                time.sleep(5 * (2 ** attempt))
                continue
            raise RuntimeError(f'feed {game_pk}: {last}') from e
    raise RuntimeError(f'feed {game_pk}: {last}')


_SCRAPER = None


def _scraper():
    """One Pitcher2026 instance for the process. Importing it is not free."""
    global _SCRAPER
    if _SCRAPER is None:
        from scrapers.pitcher2026 import BaseballSavantFocusedDownloader
        _SCRAPER = BaseballSavantFocusedDownloader()
    return _SCRAPER


_FEED_ROWS_CACHE = {}


def feed_rows(game_pk, session, refresh=False):
    """PitchID -> {column: formatted string} for one game, at full precision.

    Parsed by Pitcher2026.download_game_data so this module cannot drift from
    the scraper. Memoised because a game is walked once per team tab.
    """
    if game_pk in _FEED_ROWS_CACHE:
        return _FEED_ROWS_CACHE[game_pk]
    payload = fetch_game_json(game_pk, session, refresh=refresh)
    df = _scraper().download_game_data(game_pk, game_json=payload,
                                       raw_precision=True)
    out = {}
    if df is not None and len(df):
        df = _scraper().normalize_aaa_labels(df)
        cols = [c for c in FEED_COLS if c in df.columns]
        for rec in df.to_dict('records'):
            pid = rec.get('PitchID')
            if not pid:
                continue
            out[pid] = {c: fmt(c, rec.get(c)) for c in cols}
            out[pid]['_PTeam'] = str(rec.get('PTeam') or '')
    _FEED_ROWS_CACHE[game_pk] = out
    return out


# ── Noise scale, for deciding what needs a human ─────────────────────────────
# Wally's rule (2026-08-17): drift that is only drift gets force-overwritten;
# changes big enough to be a real correction are the ones he decides on. That
# needs a per-column yardstick, and the natural one is the column's own spread
# within a (pitcher, pitch type) group. A 10 rpm move is 0.11 of that spread; a
# 760 rpm move is 8.5 of it. Absolute thresholds cannot do that job, because
# 10 rpm and 10 mph are not remotely the same size of claim.
#
# Robust scale (MAD x 1.4826) so the misreads we are hunting do not inflate the
# yardstick that is meant to catch them. Measured over the current sheet, at the
# sample size production runs at, using groups of 30 pitches or more.
SCALE_MIN_N = 30
BASELINE_MIN_N = 10

# A pitcher's own band: the quantile of his deviations from his centre, for one
# pitch type, that a proposed value must stay inside.
#
# This replaced an "added distance" test that was too aggressive and that Wally
# caught on 2026-08-17. Bryan Abreu's fastball tilt read 12:54, the feed proposed
# 12:42, and his average is 1:00. The old rule rejected it because the change
# moved the value from 6 minutes off his average to 18. But 18 minutes is 9
# degrees of spin axis and his own band is 20 minutes, so that is an entirely
# ordinary place for one pitch to land. Distance from the average is not evidence
# of error until it leaves the range the pitcher actually occupies, and only his
# own observed spread can say where that is.
#
# 0.99 rather than the maximum is a CONVENTION, and the reasoning is that the
# maximum is fragile: one bad value already sitting in the sheet for that pitcher
# and pitch type inflates it and the test stops rejecting anything. Measured on
# HOU's 105 candidate rows the two differ on 6, and in all 6 the maximum is the
# permissive one. BAND_MIN_N of 20 is the floor for trusting a quantile at all;
# below it the pooled-SD gate takes over.
BAND_QUANTILE = 0.99
BAND_MIN_N = 20

# Drift beyond this many scale units goes to review; below it, it is written
# without asking. CONVENTION, not a measured optimum — no objective trades review
# effort against missed corrections, so none exists to bracket. What the data
# does say is that the choice barely matters: of ARI's 17,762 scaled drift cells,
# 99.85% sit below 1.0 and the sweep gives 91 review rows at 0.5, 26 at 1.0, 10
# at 2.0 and 7 at 3.0. Every point in that range leaves a list short enough to
# actually read, so 1.0 is picked for being the round number in the flat part.
DRIFT_REVIEW_SD = 1.0

# A blank cell whose candidate value sits further than this from the pitcher's
# own median for that pitch type is not offered as a fill. Also a CONVENTION.
# Wally deleted values like a 55.9 mph fastball against a 95 mph median, which is
# 33 scale units out; the gate exists so that class never comes back. Candidates
# beyond it are NOT written and NOT silently dropped — they land in the
# _IMPLAUSIBLE FILLS tab, rolled up by game, because on ARI the tail turned out
# to be a systematic per-game defect rather than 140 unrelated cells.
FILL_REVIEW_SD = 4.0

_BASELINES = None


def baselines():
    """(scale, median) per column, from the current sheet cache.

    scale[col]              -> robust SD within a (pitcher, pitch type) group
    median[col][(p, ptype)] -> that group's median, the plausibility baseline
    """
    global _BASELINES
    if _BASELINES is not None:
        return _BASELINES
    import pickle
    import statistics
    path = os.path.join(DATA, 'all_pitches_rs_cache.pkl')
    if not os.path.exists(path):
        print(f"  WARNING: {os.path.relpath(path, _ROOT)} is missing, so no "
              f"noise scale can be measured. EVERY drift and every blank fill "
              f"will be sent to review rather than written silently. Run "
              f"python3 -m pipeline.process_data to rebuild it.")
        _BASELINES = ({}, {}, {})
        return _BASELINES

    print("  measuring the per-column noise scale from the sheet cache...")
    rows = pickle.load(open(path, 'rb'))
    cols = [c for c in IN_SCOPE if c in PRECISION]
    grp = collections.defaultdict(lambda: collections.defaultdict(list))
    for x in rows:
        key = (x.get('Pitcher'), x.get('Pitch Type'))
        for c in cols:
            v = as_float(x.get(c))
            if v is not None:
                grp[c][key].append(v)

    scale, median, band = {}, {}, {}

    def summarise(c, groups, circular):
        """Fill scale[c], median[c] and band[c] for one column.

        band[c][(pitcher, type)] is that pitcher's OWN 99th-percentile deviation
        from his centre for that pitch type. It is the yardstick a recommendation
        actually needs: "is this value outside what he has been observed to throw",
        which no single pooled SD figure can answer.
        """
        sds, med, bnd = [], {}, {}
        for key, vals in groups.items():
            if circular:
                centre = collections.Counter(vals).most_common(1)[0][0]
                devs = []
                for v in vals:
                    d = abs(v - centre) % TILT_MINUTES
                    devs.append(min(d, TILT_MINUTES - d))
            else:
                centre = statistics.median(vals)
                devs = [abs(v - centre) for v in vals]
            if len(vals) >= BASELINE_MIN_N:
                med[key] = centre
            if len(vals) >= BAND_MIN_N:
                devs.sort()
                bnd[key] = devs[min(len(devs) - 1,
                                    int(BAND_QUANTILE * len(devs)))]
            if len(vals) >= SCALE_MIN_N:
                sd = 1.4826 * statistics.median(devs)
                if sd > 0:
                    sds.append(sd)
        median[c], band[c] = med, bnd
        if sds:
            scale[c] = statistics.median(sds)

    # RTilt and OTilt need their own pass. They are clock readings, so their
    # spread has to be measured circularly around the group's modal reading, or
    # every group straddling 12:00 reports a spread of about six hours. Without a
    # scale here, every tilt change lands in review however small it is, and on
    # ARI that meant 39 rows of which most were a one or two minute wobble.
    for c in TIME_COLS:
        if c not in IN_SCOPE:
            continue
        tg = collections.defaultdict(list)
        for x in rows:
            m = tilt_minutes(x.get(c))
            if m is not None:
                tg[(x.get('Pitcher'), x.get('Pitch Type'))].append(m)
        summarise(c, tg, circular=True)

    for c in cols:
        summarise(c, grp[c], circular=False)

    print(f"    scale measured for {len(scale)} columns "
          f"({len(cols)} numeric + {len(TIME_COLS & set(IN_SCOPE))} tilt)")
    _BASELINES = (scale, median, band)
    return _BASELINES


_GROUP_N = None


def group_sizes():
    """(pitcher, pitch type) -> how many pitches the sheet has for that pair.

    Shown beside every average in the review so a 4-pitch baseline is never
    mistaken for a 400-pitch one.
    """
    global _GROUP_N
    if _GROUP_N is not None:
        return _GROUP_N
    import pickle
    path = os.path.join(DATA, 'all_pitches_rs_cache.pkl')
    out = collections.Counter()
    if os.path.exists(path):
        for x in pickle.load(open(path, 'rb')):
            out[(x.get('Pitcher'), x.get('Pitch Type'))] += 1
    _GROUP_N = out
    return out


def scale_units(col, delta, old=None, new=None):
    """|delta| in the column's own noise units, or None when it has no scale.

    For RTilt and OTilt the delta has to come from tilt_gap rather than a
    subtraction, because the clock wraps at noon.
    """
    scale, _, _ = baselines()
    s = scale.get(col)
    if s is None:
        return None
    if col in TIME_COLS:
        g = tilt_gap(old, new)
        return None if g is None else g / s
    if delta is None:
        return None
    return abs(delta) / s


def centre_distance(col, pitcher, pitch_type, value):
    """How far `value` sits from that pitcher's centre for that pitch type.

    In the column's own units, circular for a tilt. None when there is no centre
    to measure against.
    """
    _, median, _ = baselines()
    m = (median.get(col) or {}).get((pitcher, pitch_type))
    if m is None:
        return None
    if col in TIME_COLS:
        mv = tilt_minutes(value)
        if mv is None:
            return None
        d = abs(mv - m) % TILT_MINUTES
        return min(d, TILT_MINUTES - d)
    v = as_float(value)
    return None if v is None else abs(v - m)


def own_band(col, pitcher, pitch_type):
    """That pitcher's own 99th-percentile deviation for that pitch type."""
    _, _, band = baselines()
    return (band.get(col) or {}).get((pitcher, pitch_type))


def fill_units(col, pitcher, pitch_type, value):
    """How far a candidate fill sits from that pitcher's own median, in units."""
    scale, median, _ = baselines()
    s = scale.get(col)
    m = (median.get(col) or {}).get((pitcher, pitch_type))
    if s is None or m is None:
        return None
    if col in TIME_COLS:
        mv = tilt_minutes(value)
        if mv is None:
            return None
        d = abs(mv - m) % TILT_MINUTES
        return min(d, TILT_MINUTES - d) / s
    v = as_float(value)
    return None if v is None else abs(v - m) / s


# ── Prior-decision ledger ────────────────────────────────────────────────────
# Wally deleted a handful of values on purpose because they were obvious
# misreads. No ledger was ever written, but two 2026-07-06 snapshots survive and
# reconstruct what the source was offering at the time he looked:
#
#   data/_tracking_pitches.pkl   396,214 pitches, sheet value AND Savant value
#                                side by side
#   data/_statcast2026_full.pkl  435,374 Savant rows, incl. spin_axis (RTilt)
#
# Measured coverage of today's blanks on rows that do have tracking:
#   Velocity   22 blank, 17 with a record (3 held a value on Jul 6)
#   Spin Rate  2,782 blank, 1,239 with a record (56 held a value on Jul 6)
#   RTilt      1,139 blank, 1,118 with a record
#
# The snapshot cannot prove a blank was a deletion rather than a feed gap, and
# it does not need to. The test is not "was this deleted", it is "is the number
# on offer today the number that was already rejected". Exact match only —
# Wally's call 2026-08-17: no tolerance band, any difference gets reviewed.
LEDGER_COLS = {'Velocity', 'Spin Rate', 'RTilt',
               'IndVertBrk', 'HorzBrk', 'xIndVrtBrk', 'xHorzBrk'}

# (game_pk, column) pairs where the SOURCE is wrong for the whole game, so no
# cell in it may be filled however plausible the individual value looks.
#
# Mexico City series, 2026-04-25 and 2026-04-26, Estadio Alfredo Harp Helu.
# Wally blanked Extension for both games on purpose. Verified 2026-08-17: the
# feed still serves extension about half a foot below each pitcher's own median
# for that pitch type (median offered-minus-usual -0.490 ft in 825093 and
# -0.543 ft in 825094, against -0.113, 0.000 and -0.017 ft in the three other
# whole-game Extension gaps, which are ordinary tracking gaps and DO get
# filled). The per-cell plausibility gate is not enough on its own here: the
# median offset is 3.7 noise units, under FILL_REVIEW_SD, so roughly two thirds
# of the 615 cells would have passed it one at a time. A defect that applies to
# a whole game has to be blocked at the game level.
KNOWN_BAD_SOURCE = {
    ('825093', 'Extension'),
    ('825094', 'Extension'),
}

_SNAPSHOT_DATE = '2026-07-06'


class DecisionLedger:
    """Values Wally has already seen and left out. Blank-fill only."""

    def __init__(self, entries=None):
        # (PitchID, column) -> {'rejected': [str, ...], 'asof': str, 'origin': str}
        self.entries = entries or {}
        self.new_rejections = {}

    # -- persistence ----------------------------------------------------------
    @classmethod
    def load(cls):
        if os.path.exists(LEDGER_PATH):
            with open(LEDGER_PATH, encoding='utf-8') as f:
                blob = json.load(f)
            entries = {tuple(k.split('|', 1)): v
                       for k, v in blob.get('decisions', {}).items()}
            print(f"  decision ledger: {len(entries)} entries from "
                  f"{os.path.relpath(LEDGER_PATH, _ROOT)}")
            return cls(entries)
        print("  decision ledger: absent, bootstrapping from the 2026-07-06 "
              "snapshots")
        return cls(cls._bootstrap())

    def save(self):
        blob = {
            'version': 1,
            'note': ('Values deliberately left out of the sheet. A blank-fill '
                     'candidate matching one of these exactly is suppressed. '
                     'Bootstrapped from data/_tracking_pitches.pkl and '
                     'data/_statcast2026_full.pkl (both 2026-07-06); every run '
                     'since appends what the review rejected.'),
            'updated': datetime.now().strftime('%Y-%m-%d'),
            'decisions': {f'{pid}|{col}': v
                          for (pid, col), v in sorted(self.entries.items())},
        }
        tmp = LEDGER_PATH + '.tmp'
        with open(tmp, 'w', encoding='utf-8') as f:
            json.dump(blob, f, indent=1, sort_keys=True)
        os.replace(tmp, LEDGER_PATH)

    # -- bootstrap ------------------------------------------------------------
    @staticmethod
    def _bootstrap():
        """Reconstruct prior decisions from the July 6 snapshots.

        Records a rejection for every LEDGER_COLS cell that is blank in the
        current sheet cache while the July 6 snapshot shows a value for it, on
        either the sheet side or the Savant side. Both numbers are stored: a
        deliberate deletion leaves the sheet-side value, and a blank that
        predates the snapshot leaves only the Savant-side proxy.
        """
        import pickle
        entries = {}
        cur_path = os.path.join(DATA, 'all_pitches_rs_cache.pkl')
        trk_path = os.path.join(DATA, '_tracking_pitches.pkl')
        sav_path = os.path.join(DATA, '_statcast2026_full.pkl')
        for p in (cur_path, trk_path, sav_path):
            if not os.path.exists(p):
                print(f"    WARNING: {os.path.relpath(p, _ROOT)} is missing, so "
                      f"the ledger bootstrap is incomplete. Every prior "
                      f"decision it covered will be offered again for review.")
                return entries

        cur = pickle.load(open(cur_path, 'rb'))
        trk = {r['pid']: r for r in pickle.load(open(trk_path, 'rb'))}
        sav = pickle.load(open(sav_path, 'rb'))
        sav['_pid'] = (sav['game_pk'].astype(int).astype(str) + '_' +
                       sav['at_bat_number'].astype(int).astype(str).str.zfill(3)
                       + '_' +
                       sav['pitch_number'].astype(int).astype(str).str.zfill(2))
        spin_axis = sav.set_index('_pid')['spin_axis'].to_dict()

        # Savant field that stands in for each sheet column in the snapshot.
        SAV_KEY = {'Velocity': 'release_speed', 'Spin Rate': 'release_spin_rate',
                   'IndVertBrk': 'pfx_z', 'HorzBrk': 'pfx_x',
                   'xIndVrtBrk': 'pfx_z', 'xHorzBrk': 'pfx_x'}

        def filled(v):
            return v not in (None, '')

        for row in cur:
            pid = row.get('PitchID')
            if not pid or pid not in trk and pid not in spin_axis:
                continue
            # Only rows that were tracked at all. A pitch with no tracking is a
            # feed gap, not a decision.
            if not (filled(row.get('PlateZ')) and filled(row.get('IndVertBrk'))):
                continue
            snap = trk.get(pid) or {}
            snap_sav = snap.get('sav') or {}
            for col in LEDGER_COLS:
                if filled(row.get(col)):
                    continue
                rejected = []
                if col == 'RTilt':
                    ax = spin_axis.get(pid)
                    if ax is not None and not pd.isna(ax):
                        t = _scraper().spin_axis_to_tilt(float(ax))
                        if t:
                            rejected.append(t)
                else:
                    sheet_then = snap.get(col)
                    if sheet_then is not None and not pd.isna(sheet_then):
                        rejected.append(fmt(col, sheet_then))
                    sk = SAV_KEY.get(col)
                    if sk and snap_sav.get(sk) is not None:
                        v = snap_sav[sk]
                        if not pd.isna(v):
                            rejected.append(fmt(col, v))
                rejected = sorted({r for r in rejected if r})
                if rejected:
                    entries[(pid, col)] = {
                        'rejected': rejected,
                        'asof': _SNAPSHOT_DATE,
                        'origin': 'bootstrap_2026_07_06_snapshot',
                    }
        print(f"    bootstrapped {len(entries)} prior decisions")
        by_col = collections.Counter(c for _, c in entries)
        for c, n in by_col.most_common():
            print(f"      {c:12} {n}")
        return entries

    # -- use ------------------------------------------------------------------
    def suppresses(self, pid, col, candidate):
        """True when `candidate` is a value already rejected for this cell."""
        e = self.entries.get((pid, col))
        return bool(e) and candidate in e['rejected']

    def record(self, pid, col, value):
        """Note that `value` was offered for a blank cell and declined."""
        key = (pid, col)
        e = self.entries.setdefault(
            key, {'rejected': [], 'asof': datetime.now().strftime('%Y-%m-%d'),
                  'origin': 'review'})
        if value not in e['rejected']:
            e['rejected'] = sorted(e['rejected'] + [value])
            self.new_rejections[key] = value


# ── What a sweep has already surfaced ────────────────────────────────────────
# Wally's rule 2026-08-17: once a sweep picks up a data point, it must not be
# picked up again by a later sweep unless the source value changes again.
#
# So the ledger is a record of "this cell was offered this exact value", not a
# record of a manual decision. At the end of every sweep each reviewable and
# each held-back change is written into it. The next sweep compares its proposal
# against that record: the same value is skipped, a different value is surfaced.
# Nothing has to be marked up by hand.
#
# The bootstrap entries from the 2026-07-06 snapshots live in the same store and
# behave identically. They are simply the values that were already declined
# before the first sweep ran.
#
# --no-record turns the write-back off, for an exploratory run that should not
# silence anything.
RECORDED_KINDS = {'new', 'drift', 'suppressed', 'blocked_source'}

# The override column. A pre-filled character rather than a form control:
# openpyxl cannot write a real Excel checkbox, and Numbers discards Excel data
# validation on import, so there was nothing to click there.
BOX_EMPTY = '\u2610'      # ballot box
BOX_TICKED = '\u2611'     # ballot box with check


def is_override(value):
    """Did this row get marked to do the opposite of the recommendation?

    The test is "the cell is no longer an untouched box", NOT "the cell contains
    a tick". Wally reviewed the first workbook in Numbers, where the dropdown had
    been stripped and the box could not be clicked, and his workaround was to
    DELETE the character. So an empty cell has to count as an override, and so
    does a tick, an x, or any other mark. One rule, same behaviour in Excel,
    Sheets and Numbers, and it matches what he already did.

    read_decisions() guards the failure mode this creates: a workbook that comes
    back with no boxes left ANYWHERE is a format conversion dropping the glyph,
    not a decision to overturn every recommendation, and it refuses rather than
    inverting a few hundred rows.
    """
    return str(value or '').strip() != BOX_EMPTY


def record_sweep(changes, ledger):
    """Note every value this sweep surfaced or held back, so it does not recur.

    Returns how many (PitchID, column, value) triples were newly recorded.
    """
    before = sum(len(e['rejected']) for e in ledger.entries.values())
    for ch in changes:
        if ch.kind in RECORDED_KINDS and ch.new:
            ledger.record(ch.pitch_id, ch.col, ch.new)
    after = sum(len(e['rejected']) for e in ledger.entries.values())
    return after - before


def _wanted(ch, decisions, kinds):
    """Should this change be written?

    A decision read out of a reviewed workbook wins outright, including over the
    kind filter. That is what lets an override on the already-raised tab recover
    a previously declined value: those rows are not in `kinds`, so without this
    they could never come back however the box was marked.

    With no decision on file, an automatic class (extra decimals, sub-precision
    drift) is written and everything else follows its own recommendation.
    """
    if decisions is not None and (ch.pitch_id, ch.col) in decisions:
        return decisions[(ch.pitch_id, ch.col)]
    if ch.kind not in kinds:
        return False
    return not ch.rec or ch.rec == ADOPT


# ── Reading a reviewed workbook back in ──────────────────────────────────────
def read_decisions(path):
    """Turn a reviewed dry-run workbook into a decision map.

    Returns {(PitchID, column): True to write, False to leave alone}.

    The recommendation column carries the verdict; the Reject box inverts it.
    An untouched workbook therefore means "do exactly what was recommended",
    which is the safe reading — a file that never got opened cannot silently
    flip anything.
    """
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    out, flipped, seen = {}, 0, 0
    boxes_seen = boxes_intact = 0
    for name in wb.sheetnames:
        ws = wb[name]
        header = None
        for row in ws.iter_rows(values_only=True):
            if header is None:
                header = list(row)
                if not {'PitchID', 'Recommend', 'Reject'} <= set(header):
                    break            # SUMMARY, HOW TO READ, or a rollup tab
                continue
            rec_row = dict(zip(header, row))
            pid = rec_row.get('PitchID')
            if not pid:
                continue
            # On a per-column tab the column IS the tab name. On the zone tabs it
            # is in a Column cell, because those carry SzTop and SzBot together.
            col = rec_row.get('Column') or name
            verdict = str(rec_row.get('Recommend') or '').strip().lower()
            write = verdict == ADOPT
            boxes_seen += 1
            if is_override(rec_row.get('Reject')):
                write = not write
                flipped += 1
            else:
                boxes_intact += 1
            out[(str(pid), str(col))] = write
            seen += 1
    # A workbook with no boxes left ANYWHERE is a format conversion dropping the
    # glyph, not a decision to overturn every recommendation. Refuse rather than
    # invert a few hundred rows on a guess.
    if boxes_seen and boxes_intact == 0:
        raise SystemExit(
            f"REFUSING to read {os.path.basename(path)}: not one of its "
            f"{boxes_seen} override cells still holds '{BOX_EMPTY}'. That looks "
            f"like the character not surviving a format conversion rather than a "
            f"decision to overturn every recommendation. Re-export as .xlsx from "
            f"the app you reviewed it in, check the Reject column still shows "
            f"'{BOX_EMPTY}' on the rows you left alone, and run this again.")
    print(f"  read {seen} decisions from {os.path.basename(path)}: "
          f"{boxes_intact} left alone, {flipped} overridden. "
          f"{sum(out.values())} cells to write, {seen - sum(out.values())} to "
          f"leave alone.")
    return out


# ── Recommendation ───────────────────────────────────────────────────────────
# One verdict per reviewable row: adopt the new value, or leave the cell alone.
# Deliberately blind to the ledger — Wally's instruction 2026-08-17 is that the
# recommendation must stand on the data in front of it, not on whether a value
# has been seen before.
#
# The whole judgement rests on one comparison: where does the proposed value sit
# relative to THIS pitcher's own median for THIS pitch type. That is what makes
# Kai-Wei Teng's sweeper reading IVB 18.6 against his own 1.0 obviously wrong,
# and Bryan King's fastball moving 2223 -> 2126 against his own 2106 obviously
# right, even though the second is the larger raw change.
ADOPT, REJECT = 'adopt', 'do not adopt'

# Judging a value against the pitcher's own median only makes sense for the
# pitch-metric block, which is single-homed above as PITCH_METRIC_COLS.
#
# Caught 2026-08-17 on the first run that showed recommendations: xSLG
# 3.403 -> 4.000 was rejected as 3.8 units from Mike Burrows' median, when 4.000
# IS the xSLG of a home run and his median across all batted balls is about 0.24.
# The same error hit xBA, xwOBA, RunExp and the whole bat-tracking block, where
# SwingLength belongs to the hitter's swing and has nothing to do with who threw
# the pitch. A per-hitter baseline WOULD be meaningful for bat tracking, and that
# is a worthwhile refinement, but a wrong baseline is worse than none so it is not
# guessed at here.

# Plain words for the report. `suppressed` in particular read like a verdict when
# it only ever meant "this value was raised by an earlier sweep".
PLAIN_KIND = {
    'new': 'cell is blank',
    'drift': 'value changed at source',
    'blocked_source': 'whole game bad at source',
    'zone_fix': 'wrong hitter\'s zone',
    'zone_fix_nodonor': 'zone outlier, no donor',
    'suppressed': 'already raised by an earlier sweep',
    'precision': 'extra decimals only',
    'drift_sub': 'change below stored precision',
    'drift_small': 'change below one noise unit',
}


def recommend(kind, col, old, new, units, pitcher, pitch_type, game_pk):
    """Return (verdict, one-line reason)."""
    if (game_pk, col) in KNOWN_BAD_SOURCE:
        return REJECT, 'source is wrong for this whole game'

    if kind in ('zone_fix', 'zone_fix_nodonor'):
        return ADOPT, 'zone belongs to another hitter'

    if col not in PITCH_METRIC_COLS:
        if col in STRING_COLS:
            return ADOPT, 'feed is authoritative for this column'
        return ADOPT, ('per-event value, so the pitcher\'s own average is not a '
                       'yardstick for it')

    # ONE test for fills and for drift alike: does the proposed value stay inside
    # the range this pitcher actually occupies for this pitch type?
    #
    # The direction of a change is NOT evidence. A value moving from very typical
    # to merely typical is not a defect, and treating it as one is what rejected
    # Bryan Abreu's fastball tilt 12:54 -> 12:42 when his own band is 20 minutes
    # wide. What matters is only whether the value leaves the range.
    dist = centre_distance(col, pitcher, pitch_type, new)

    # No centre to measure against: a string column, or a pitcher and pitch type
    # pair too thin for a median. Description, BBType, Count and Event land here,
    # and the feed is the authority on all four — they are the scoring and tagging
    # corrections that motivated the whole exercise.
    if dist is None:
        if col in STRING_COLS:
            return ADOPT, 'feed is authoritative for this column'
        return ADOPT, 'no baseline for this pitcher and pitch type'

    unit = 'min' if col in TIME_COLS else ''
    band = own_band(col, pitcher, pitch_type)
    if band is not None:
        if dist > band:
            return REJECT, (f'{dist:.1f}{unit} from his own centre, outside the '
                            f'{band:.1f}{unit} band he throws for this pitch type')
        return ADOPT, (f'{dist:.1f}{unit} from his own centre, inside his '
                       f'{band:.1f}{unit} band')

    # Too few pitches for a band. Fall back to the pooled gate, which is coarser
    # but is all there is when a pitcher has thrown a pitch type a handful of
    # times.
    units = fill_units(col, pitcher, pitch_type, new)
    if units is None:
        return ADOPT, 'no baseline for this pitcher and pitch type'
    if units > FILL_REVIEW_SD:
        return REJECT, (f'{units:.1f} noise units out, and too few pitches of '
                        f'this type to measure his own band')
    return ADOPT, (f'{units:.1f} noise units out, and too few pitches of this '
                   f'type to measure his own band')


# ── Diff ─────────────────────────────────────────────────────────────────────
# One row per proposed cell change.
Change = collections.namedtuple(
    'Change', 'tab row col pitcher batter game_date pitch_id pitch_type '
              'old new kind delta source units rec rec_why')
# `units` is |delta| expressed in the column's own noise scale, or None for a
# string, a time value, or a column with no measurable spread.

# Written without asking. Everything else needs Wally to look at it.
AUTO_KINDS = {'precision', 'drift_sub', 'drift_small', 'zone_fix', 'zone_fix_nodonor'}
# Counted on SUMMARY only; never listed row by row.
SUMMARY_ONLY_KINDS = {'precision', 'drift_sub', 'drift_small'}

# kind:
#   new         sheet cell is blank, the source has a value
#   drift       both present and they disagree beyond stored precision
#   precision   both present, same number, the source carries more digits
#   suppressed  blank cell, but the source still offers a rejected value
#   missing     the whole pitch is absent from the sheet


def classify(col, old, new):
    """Return 'new' | 'drift' | 'drift_sub' | 'precision' | None for one cell.

    The three change classes exist because they need different review, and
    lumping them together buries the ~15k cells that carry real information
    under ~200k that do not.

      precision  The stored value is EXACTLY today's source value rounded to the
                 stored value's own decimal depth. Nothing moved; digits are
                 being added. LaunchAngle 24 -> 24.3, SzTop 3.19 -> 3.1900.
      drift_sub  The source moved, but by less than half a unit of the stored
                 depth, so the sheet was never displaying the difference. VAA
                 stored at 2 decimals moving 0.003. Reported as a count only.
      drift      The source moved by more than the sheet was showing. This is
                 the class worth a human decision: a re-tagged Description, a
                 corrected Spin Rate, a revised xwOBA.

    `precision` is a strict subset of `drift_sub`, so it is tested first.
    """
    if new == '':
        return None                      # never blank a populated cell
    if old == '':
        return 'new'
    if old == new:
        return None
    if col in STRING_COLS or col in TIME_COLS:
        return 'drift'
    o, n = as_float(old), as_float(new)
    if o is None or n is None:
        return 'drift'
    d = stored_decimals(old)
    if d is None:
        return 'drift'

    # Both comparisons run in the DECIMAL domain, not the float domain. Done in
    # floats, HAA 1.90 -> 1.905 computes |n-o| as 0.0050000000000001155 and
    # tests false against half a unit of 10^-2, so the code contradicted its own
    # rule and filed the cell as drift. That single artifact produced 2,418 of
    # ARI's 20,224 drift rows: HAA 662, PlateZ 654, VAA 328, RelPosX 87,
    # RelPosZ 25, SzTop 24, Extension 22, all sitting exactly on the boundary.
    # Decimal(old) is exact because `old` is the sheet's own string, and
    # Decimal(new) is exact because `new` came out of fmt() at a fixed depth.
    # There is no epsilon and no tolerance to tune.
    try:
        do, dn = Decimal(old), Decimal(new)
    except InvalidOperation:
        return 'drift'
    if do.quantize(Decimal(1).scaleb(-d), rounding=ROUND_HALF_UP) == \
       dn.quantize(Decimal(1).scaleb(-d), rounding=ROUND_HALF_UP):
        return 'precision'
    if abs(dn - do) <= Decimal(1).scaleb(-d) / 2:
        return 'drift_sub'
    return 'drift'


# Deviation from a hitter's own modal strike zone, beyond which the cell is not
# a rounding artifact. MEASURED, not chosen: across all 583,619 zone cells the
# distribution is sharply bimodal in inches — 573,333 within 0.06, then 3,784 to
# 0.12, then 6,411 to 0.25, then it collapses to 46 in the 0.25 to 0.5 band and
# 45 beyond. That 140x drop between the 0.12-0.25 band and the 0.25-0.5 band is
# the break, and it leaves 91 cells in the whole season above it.
ZONE_OUTLIER_FT = 0.25 / 12

_ZONE_INDEX = None


def _sheet_zone_index():
    """Season-wide zone observations, from the sheet cache.

    Returns (counts, game_hitters):
      counts[(batter, bteam)][(sztop2, szbot2)] -> how many pitches
      game_hitters[game_pk]                     -> set of (batter, bteam)

    Fresh copies each call, so the caller can fold this run's own observations in
    without polluting the cached index.
    """
    global _ZONE_INDEX
    if _ZONE_INDEX is None:
        import pickle
        path = os.path.join(DATA, 'all_pitches_rs_cache.pkl')
        counts = collections.defaultdict(collections.Counter)
        rosters = collections.defaultdict(set)
        if not os.path.exists(path):
            print(f"    WARNING: {os.path.relpath(path, _ROOT)} is missing, so "
                  f"the modal strike zone can only be built from the tabs in "
                  f"this run. Run every tab together, or no zone will be "
                  f"repaired.")
        else:
            for x in pickle.load(open(path, 'rb')):
                b, pid = x.get('Batter'), x.get('PitchID')
                t, bo = as_float(x.get('SzTop')), as_float(x.get('SzBot'))
                if not b or not pid or t is None or bo is None:
                    continue
                key = (b, x.get('BTeam'))
                counts[key][(round(t, 2), round(bo, 2))] += 1
                rosters[pid.split('_')[0]].add(key)
        _ZONE_INDEX = (counts, rosters)
    c, r = _ZONE_INDEX
    out_c = collections.defaultdict(collections.Counter)
    for k, v in c.items():
        out_c[k] = v.copy()
    out_r = collections.defaultdict(set)
    for k, v in r.items():
        out_r[k] = set(v)
    return out_c, out_r


def zone_outlier_changes(zone_obs, fix_nodonor=True):
    """Repair strike-zone cells that carry a DIFFERENT hitter's zone.

    Wally's read (2026-08-17): the operator failed to change who was batting.
    Tested directly against all 91 season cells beyond ZONE_OUTLIER_FT, and it
    holds for 40 of the 51 at-bats: the odd zone equals the modal zone of
    another hitter who batted in that same game. Rocchio's at-bat 824408_027
    carries Trevor Larnach's zone, Ward's and Beavers' in 824168 carry Colton
    Cowser's, and Altuve's 824170_065 carries Isaac Paredes'.

    A cell is only rewritten when that match exists. The other 11 at-bats have a
    different signature and are reported instead of repaired: SzTop matches the
    hitter's own modal value EXACTLY and only SzBot moves, by roughly 0.4 inch.
    A mis-attributed batter would move both together, so these are something
    else — most likely a genuine re-measurement of the bottom of the zone, which
    a change in crouch would produce without moving shoulder height.

    Keyed on (Batter, BTeam), never on Batter alone, because two players share a
    name: Max Muncy reads 3.128/1.579 for LAD and 3.228/1.629 for ATH, and both
    are correct.

    The 9 unmatched at-bats are repaired as well, tagged `zone_fix_nodonor` and
    listed in their own tab. Wally's call 2026-08-17. Measured signature: SzTop
    is off by 0.00 inch in every one of the 9, and only SzBot moves, by 0.28 to
    0.47 inch. A mis-attributed batter moves both together, so the mechanism is
    different — but the alternate SzBot holds steady for the whole at-bat, which
    reads as an artifact rather than a genuine re-read of the knee. It is 17
    cells in 583,619, all under half an inch. Pass fix_nodonor=False to leave
    them alone.

    Returns (changes, unexplained).
    """
    # The modal zone and the game rosters come from the WHOLE sheet, not from
    # this run's tabs. A hitter appears in the tab of every team that pitched to
    # him, so a --teams run sees a slice of him. Deriving the mode from that
    # slice produces false matches: on a HOU-only run Brent Rooker's HOU-slice
    # mode came out as 3.290/1.688 and "explained" Zack Gelof's outlier, which
    # the full-season data says is unexplained. The cache is today's sheet, which
    # is the right basis for "what is this hitter's usual zone", and it agrees
    # with the feed everywhere except on the outliers being hunted.
    counts, game_hitters = _sheet_zone_index()
    for o in zone_obs:
        t, b = as_float(o['sztop']), as_float(o['szbot'])
        if t is None or b is None:
            continue
        key = (o['batter'], o['bteam'])
        counts[key][(round(t, 2), round(b, 2))] += 1
        game_hitters[o['pid'].split('_')[0]].add(key)

    # Ties broken by count then by value, so a re-run cannot flip the answer.
    modal = {k: max(sorted(c), key=lambda v: c[v]) for k, c in counts.items()}

    # A modal zone is only trustworthy when it rests on enough observations and
    # is a clear plurality. This is a fail-closed guard for PARTIAL runs, not a
    # tuned threshold: a hitter appears in the tab of every team that pitched to
    # him, so `--teams HOU` might see only two of his rows, and if both carry the
    # bad zone the mode IS the bad zone. In a full-season run the affected
    # hitters carry 979 to 2,220 rows with the mode at over 99%, so this drops
    # nothing real. It only stops a narrow run from writing a wrong repair.
    ZONE_MIN_OBS, ZONE_MIN_SHARE = 20, 0.80
    weak = set()
    for k, c in counts.items():
        n = sum(c.values())
        if n < ZONE_MIN_OBS or c[modal[k]] / n < ZONE_MIN_SHARE:
            weak.add(k)
    if weak:
        print(f"    zone repair: {len(weak)} of {len(modal)} hitters have too "
              f"few or too mixed observations for a trustworthy modal zone and "
              f"are skipped. Run every tab together to clear this.")

    out, unexplained = [], []
    for o in zone_obs:
        key = (o['batter'], o['bteam'])
        if key not in modal or key in weak:
            continue
        mt, mb = modal[key]
        t, b = as_float(o['sztop']), as_float(o['szbot'])
        if t is None or b is None:
            continue
        if max(abs(t - mt), abs(b - mb)) <= ZONE_OUTLIER_FT:
            continue

        gpk = o['pid'].split('_')[0]
        match = [hk for hk in game_hitters[gpk]
                 if hk != key
                 and abs(modal[hk][0] - round(t, 2)) <= ZONE_OUTLIER_FT
                 and abs(modal[hk][1] - round(b, 2)) <= ZONE_OUTLIER_FT]
        if not match:
            unexplained.append({
                'game_date': o['game_date'], 'game_pk': gpk, 'pid': o['pid'],
                'at_bat': o['pid'].split('_')[1], 'tab': o['tab'],
                'batter': o['batter'], 'bteam': o['bteam'],
                'has': f'{t:.3f}/{b:.3f}', 'modal': f'{mt:.3f}/{mb:.3f}',
                'sztop_off_in': round((t - mt) * 12, 2),
                'szbot_off_in': round((b - mb) * 12, 2),
            })
            if not fix_nodonor:
                continue

        kind = 'zone_fix' if match else 'zone_fix_nodonor'
        for col, want in (('SzTop', mt), ('SzBot', mb)):
            sheet_val = o['sheet_' + col]
            new = fmt(col, want)
            # Only a real move counts as a repair. When one of the pair already
            # matches (Gelof's SzTop was right and only SzBot was wrong) the
            # difference is precision, and the ordinary precision pass owns it.
            if classify(col, sheet_val, new) in (None, 'precision'):
                continue
            o_f = as_float(sheet_val)
            out.append(Change(
                tab=o['tab'], row=o['row'], col=col, pitcher=o['pitcher'],
                batter=o['batter'], game_date=o['game_date'],
                pitch_id=o['pid'], pitch_type=o.get('pitch_type', ''),
                old=sheet_val, new=new, kind=kind,
                delta=(want - o_f) if o_f is not None else None, units=None,
                rec=ADOPT,
                rec_why=(f'carries {match[0][0]}\'s zone' if match
                         else 'outlier against his own zone, no donor found'),
                source=(f'zone of {match[0][0]}' if match
                        else 'no donor in this game')))
    return out, unexplained


def diff_tab(tab, rows, header, feed_by_pid, savant_lookup, ledger, zone_obs):
    """Compare one team tab against its sources. Returns (changes, missing).

    Appends one record per row to `zone_obs` so the optional modal-zone pass can
    run across every tab at once, which it must: a hitter appears in the tab of
    each team that pitched to him.
    """
    col_idx = {name: j for j, name in enumerate(header) if name}
    if 'PitchID' not in col_idx:
        return [], []
    pid_c = col_idx['PitchID']
    present = set()
    changes = []

    live = [c for c in IN_SCOPE if c in col_idx]
    for r_idx, row in enumerate(rows[1:], start=2):
        pid = row[pid_c] if pid_c < len(row) else ''
        if not pid or '_' not in pid:
            continue
        present.add(pid)
        expected = dict(feed_by_pid.get(pid) or {})
        # Savant keys on unpadded ints; the sheet pads.
        parts = pid.split('_')
        if len(parts) == 3:
            try:
                skey = (parts[0], str(int(parts[1])), str(int(parts[2])))
            except ValueError:
                skey = None
            if skey:
                for c, v in (savant_lookup.get(skey) or {}).items():
                    # download_statcast also returns Outs and Event, which the
                    # FEED owns here. Letting them through would undo the main
                    # reason for the feed pass: Savant's generic `field_out`
                    # code cannot distinguish Groundout from Lineout, so a
                    # scoring correction could never land. Authority is one
                    # source per column, enforced by the assert above.
                    if c not in SAVANT_COL_SET:
                        continue
                    expected[c] = v if c in STRING_COLS else fmt(c, as_float(v))
        if not expected:
            continue

        def cell(name):
            j = col_idx.get(name)
            return (row[j].strip() if j is not None and j < len(row) and row[j]
                    else '')

        pitcher, batter, gdate = cell('Pitcher'), cell('Batter'), cell('Game Date')
        pitch_type = cell('Pitch Type')   # read-only here; never written

        # Feed truth for the zone, recorded whether or not it differs, because
        # the modal pass needs every observation to find the mode.
        if batter and ('SzTop' in expected or 'SzBot' in expected):
            zone_obs.append({
                'tab': tab, 'row': r_idx, 'pid': pid, 'pitcher': pitcher,
                'batter': batter, 'bteam': cell('BTeam'), 'game_date': gdate,
                'pitch_type': pitch_type,
                'sztop': expected.get('SzTop', ''),
                'szbot': expected.get('SzBot', ''),
                'sheet_SzTop': cell('SzTop'), 'sheet_SzBot': cell('SzBot'),
            })

        ep_row = pitch_type.strip().upper() == EP_PITCH_TYPE

        for col in live:
            if ep_row and col in PITCH_METRIC_COLS:
                continue
            c_idx = col_idx[col]
            old = (row[c_idx] if c_idx < len(row) else '') or ''
            old = old.strip()
            new = expected.get(col)
            if new is None:
                continue
            kind = classify(col, old, new)
            if kind is None:
                continue
            o, n = as_float(old), as_float(new)
            delta = (n - o) if (o is not None and n is not None) else None
            units = None

            # The recommendation is computed FIRST, from the data alone, and is
            # deliberately blind to the ledger.
            rec = rec_why = ''
            if kind in ('new', 'drift'):
                probe = (scale_units(col, delta, old=old, new=new)
                         if kind == 'drift' else
                         fill_units(col, pitcher, pitch_type, new))
                rec, rec_why = recommend(kind, col, old, new, probe,
                                         pitcher, pitch_type, pid.split('_')[0])
                if (pid.split('_')[0], col) in KNOWN_BAD_SOURCE:
                    kind = 'blocked_source'

            # Already seen with this same value, so not raised again. Wally's
            # rule 2026-08-17. Applies to a drift overwrite as much as to a fill,
            # or every run re-proposes Hunter Brown's 1403 rpm curveball.
            if kind in ('new', 'drift') and ledger.suppresses(pid, col, new):
                kind = 'suppressed'
            elif kind == 'new':
                units = fill_units(col, pitcher, pitch_type, new)
            elif kind == 'drift':
                # Wally's rule: only drift large enough to be a real correction
                # gets a human. A column with no measurable scale (every string
                # and time column) has units None and always goes to review,
                # which is right — those are the re-tags he asked for, and there
                # are only a handful per tab.
                units = scale_units(col, delta, old=old, new=new)
                if units is not None and units <= DRIFT_REVIEW_SD:
                    kind = 'drift_small'

            changes.append(Change(
                tab=tab, row=r_idx, col=col, pitcher=pitcher, batter=batter,
                game_date=gdate, pitch_id=pid, pitch_type=pitch_type,
                old=old, new=new, kind=kind,
                delta=delta, units=units, rec=rec, rec_why=rec_why,
                source='feed' if col in FEED_COLS else 'savant'))

    missing = []
    for pid, exp in feed_by_pid.items():
        if pid in present:
            continue
        if exp.get('_PTeam', '').upper() != tab.upper():
            continue          # belongs to the other team's tab
        missing.append((pid, exp))
    return changes, missing


# ── Report ───────────────────────────────────────────────────────────────────
# Which player column each metric is reviewed by. Zone and batted-ball columns
# read naturally by hitter; everything else by pitcher.
BATTER_SIDE = {'SzTop', 'SzBot', 'Batter', 'Bats'}


def write_report(changes, missing, ledger, path, unexplained_zones=None):
    """One workbook. A summary tab, then one tab per column, then extras."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    SCALES, MEDIANS, _ = baselines()
    GROUP_N = group_sizes()

    wb = Workbook()
    wb.remove(wb.active)
    bold = Font(bold=True)

    def sheet(title, headers):
        ws = wb.create_sheet(title=title[:31])
        for i, h in enumerate(headers, start=1):
            ws.cell(row=1, column=i, value=h).font = bold
        ws.freeze_panes = 'A2'
        return ws

    def fit(ws, cap=90):
        """Size every column to its own longest cell, header included.

        Replaces the hand-written width lists, which were guesses and went stale
        the moment a column moved. This reads the sheet that was actually
        written, so a two-character Pitch Type column stays two characters wide
        and a reason column is only as wide as its longest reason.

        A column whose content runs past `cap` is capped there and wrapped, so
        one long cell cannot push every other column off the screen.
        """
        for col_cells in ws.iter_cols():
            widest = 0
            for c in col_cells:
                if c.value is None:
                    continue
                for line in str(c.value).split('\n'):
                    widest = max(widest, len(line))
            letter = get_column_letter(col_cells[0].column)
            if widest > cap:
                ws.column_dimensions[letter].width = cap
                for c in col_cells[1:]:
                    c.alignment = Alignment(wrap_text=True, vertical='top')
            else:
                # +3 covers the padding Excel adds inside a cell plus the extra
                # width a bold header needs, so no heading is clipped.
                ws.column_dimensions[letter].width = widest + 3

    def add_checkboxes(ws, col_letter, n_rows):
        """Put an empty ballot box in every cell of a column, with a two-item
        dropdown so it can be ticked without typing.

        openpyxl cannot create a real Excel form-control checkbox, and a real one
        would not survive a round trip through Sheets or Numbers anyway. A
        validated character does, and it reads back reliably. is_override()
        accepts
        a typed x or any tick glyph as well, so a round trip that rewrites the
        character cannot lose a decision.
        """
        if n_rows < 1:
            return
        dv = DataValidation(type='list',
                            formula1=f'"{BOX_EMPTY},{BOX_TICKED}"',
                            allow_blank=True, showDropDown=False)
        dv.error = f'Choose {BOX_EMPTY} or {BOX_TICKED}'
        dv.promptTitle = 'Override the recommendation'
        dv.prompt = ('Tick to do the OPPOSITE of the Recommend column on this '
                     'row. Leave it alone to accept the recommendation.')
        ws.add_data_validation(dv)
        dv.add(f'{col_letter}2:{col_letter}{n_rows + 1}')
        for r in range(2, n_rows + 2):
            c = ws[f'{col_letter}{r}']
            c.value = BOX_EMPTY
            c.alignment = Alignment(horizontal='center')

    ws = sheet('HOW TO READ', ['', ''])
    for k, v in [
        ('What this file is',
         'A dry run. Nothing has been written to any sheet.'),
        ('Each value appears once',
         'Every value in this file has been recorded in '
         'data/backfill_decisions.json. A later sweep will NOT raise it again '
         'unless the source value changes to something different.'),
        ('If you do nothing',
         'The change is not applied, and it will not come back. Re-run with '
         '--apply to write, or delete the cell entry from the ledger to see it '
         'again.'),
        ('To see everything again',
         'Delete data/backfill_decisions.json, or run with --no-record so a '
         'sweep does not record what it surfaced.'),
        ('Recommend',
         'My verdict on this row: "%s" or "%s". Computed from the data only — it '
         'ignores whether an earlier sweep already raised this value.'
         % (ADOPT, REJECT)),
        ('Reject',
         'CHANGE the box to do the OPPOSITE of Recommend. Deleting the box '
         'counts, and so does a tick or a typed x — whatever your app lets you '
         'do. Numbers strips the dropdown, so deleting is the way there. Leave a '
         'box untouched to accept the recommendation on that row.'),
        ('Reject, worked example',
         'Recommended "%s" plus an override means leave the cell alone. '
         'Recommended "%s" plus an override means write it.' % (ADOPT, REJECT)),
        ('Send it back',
         'python3 -m scrapers.backfill_full --apply --decisions-from <this file>'),
        ('Rows are sorted',
         'Everything I recommend against is at the TOP of each tab, so the rows '
         'most likely to need an override are the first ones you see.'),
        ('Change: cell is blank',
         'The cell is empty and the source has a value.'),
        ('Change: value changed at source',
         'Both have a value and they differ by more than %s of the column\'s '
         'own spread within a pitcher and pitch type.' % DRIFT_REVIEW_SD),
        ('How a row is judged',
         'A value is recommended against only when it falls OUTSIDE the range '
         'that pitcher actually occupies for that pitch type, measured from his '
         'own pitches. The direction of a change is not held against it, so a '
         'value moving from very typical to merely typical is still adopted.'),
        ('Off usual',
         'Proposed value minus that pitcher\'s median for that pitch type. Shown '
         'only for the pitch-metric columns, because his median says nothing '
         'about a per-event value like xSLG or a hitter\'s swing length.'),
        ('_FILLS BY PITCH',
         'A read-only rollup: one row per pitch instead of per cell, so a pitch '
         'that gained its whole tracking block reads as one thing. Decide on the '
         'per-column tabs.'),
        ('_ALREADY RAISED (not written)',
         'Values an earlier sweep already put in front of you, plus the ones you '
         'had deleted before the first sweep ran. Nothing here is written; it is '
         'a record. Override a row to bring that value back.'),
        ('Not listed here',
         'Extra-decimal rewrites, drift below stored precision, and drift below '
         'one noise unit. Counted on SUMMARY and written without asking.'),
        ('EP pitches',
         'Position players. Release, movement, spin and approach columns are '
         'skipped for them. Count, Description, Event and batted-ball data are '
         'kept.'),
    ]:
        ws.append([k, v])
    fit(ws)

    by_col = collections.defaultdict(list)
    for ch in changes:
        by_col[ch.col].append(ch)

    # ---- Summary -----------------------------------------------------------
    ws = sheet('SUMMARY', ['Column', 'Source', 'Decimals',
                           'Fills to review', 'Fills implausible',
                           'Drift to review', 'Drift auto-written',
                           'Precision rewrites', 'Zone fixes', 'Suppressed',
                           'Median |drift| reviewed', 'Max |drift| reviewed',
                           'Noise scale (1 unit)', 'Rows to read'])
    order = sorted(by_col, key=lambda c: (c not in FEED_COLS, c))
    for c in order:
        rows_ = by_col[c]
        kinds = collections.Counter(r.kind for r in rows_)
        # Magnitudes describe the REVIEWED drift only. Mixing in the
        # sub-precision class would drag every median toward zero and make a
        # real correction look like noise.
        deltas = sorted(abs(r.delta) for r in rows_
                        if r.delta is not None and r.kind == 'drift')
        med = deltas[len(deltas) // 2] if deltas else None
        auto = kinds['drift_sub'] + kinds['drift_small']
        to_read = kinds['new'] + kinds['drift']
        sc = SCALES.get(c)
        ws.append([c, 'feed' if c in FEED_COLS else 'savant',
                   PRECISION.get(c, 'string/time'),
                   kinds['new'], kinds['fill_implausible'],
                   kinds['drift'], auto,
                   kinds['precision'],
                   kinds['zone_fix'] + kinds['zone_fix_nodonor'],
                   kinds['suppressed'],
                   round(med, 6) if med is not None else '',
                   round(deltas[-1], 6) if deltas else '',
                   round(sc, 4) if sc else '',
                   to_read])
    total = collections.Counter(r.kind for r in changes)
    ws.append([])
    ws.append([])
    ws.append(['TOTAL', '', '', total['new'], total['fill_implausible'],
               total['drift'], total['drift_sub'] + total['drift_small'],
               total['precision'],
               total['zone_fix'] + total['zone_fix_nodonor'], total['suppressed'],
               '', '', '',
               total['new'] + total['drift']])
    ws.append(['Missing pitches', '', '', len(missing)])
    ws.append(['Zone at-bats unexplained', '', '', len(unexplained_zones or [])])
    ws.append([])
    ws.append(['Drift is sent to review above %s noise units; blank fills are '
               'offered below %s. See DRIFT_REVIEW_SD / FILL_REVIEW_SD.'
               % (DRIFT_REVIEW_SD, FILL_REVIEW_SD)])
    fit(ws)

    # ---- One tab per column ------------------------------------------------
    # `precision` and `drift_sub` are counted on SUMMARY and deliberately NOT
    # listed: Wally's call 2026-08-17, summary only. Together they run to
    # millions of rows and carry no judgement, because neither one moves the
    # value by as much as the sheet was already displaying.
    # Already-seen rows are left OFF the tabs entirely. Wally's rule: a value a
    # sweep has raised once does not come back. They stay in the SUMMARY counts.
    HIDDEN = SUMMARY_ONLY_KINDS | {'suppressed'}
    for c in order:
        rows_ = [r for r in by_col[c] if r.kind not in HIDDEN]
        if not rows_:
            continue
        keyer = ((lambda r: (r.tab, r.batter, r.game_date, r.pitch_id))
                 if c in BATTER_SIDE else
                 (lambda r: (r.tab, r.pitcher, r.game_date, r.pitch_id)))
        rows_.sort(key=keyer)
        # Pitch Type and the pitcher's own average for that pitch type sit next
        # to the proposed value. Wally's ask 2026-08-17: HorzBrk 11.2 means
        # nothing until you know it is a slider and that his sliders average
        # 10.8. `Off usual` is the proposed value minus that average, which is
        # the number that actually decides whether a change is believable.
        # Recommendation and the override box sit immediately after Proposed,
        # where the eye already is. Wally's layout, 2026-08-17.
        ws = sheet(c, ['Team', 'Pitcher', 'Pitch Type', 'Batter', 'Game Date',
                       'GameID', 'PitchID', 'Sheet row', 'Change',
                       'Current', 'Proposed',
                       'Recommend', 'Reject', 'Why',
                       'Pitcher avg on this type', 'Off usual', 'n for avg',
                       'Delta', 'Noise units'])
        REC_LETTER, BOX_LETTER = 'L', 'M'
        colmed = MEDIANS.get(c) or {}
        rows_.sort(key=lambda r: (r.rec != REJECT, keyer(r)))
        show_avg = c in PITCH_METRIC_COLS
        for r in rows_:
            avg = colmed.get((r.pitcher, r.pitch_type)) if show_avg else None
            n_for_avg = GROUP_N.get((r.pitcher, r.pitch_type), '')
            if c in TIME_COLS and avg is not None:
                avg_disp = f'{(int(avg) // 60) or 12}:{int(avg) % 60:02d}'
                off = tilt_gap(r.new, avg_disp)
            else:
                avg_disp = round(avg, PRECISION.get(c, 3)) if avg is not None else ''
                nv = as_float(r.new)
                off = (round(nv - avg, PRECISION.get(c, 3))
                       if (avg is not None and nv is not None) else '')
            ws.append([r.tab, r.pitcher, r.pitch_type, r.batter, r.game_date,
                       r.pitch_id.split('_')[0], r.pitch_id,
                       r.row, PLAIN_KIND.get(r.kind, r.kind), r.old, r.new,
                       r.rec, None, r.rec_why,
                       avg_disp, off if off is not None else '', n_for_avg,
                       round(r.delta, 6) if r.delta is not None else '',
                       round(r.units, 2) if r.units is not None else ''])
        add_checkboxes(ws, BOX_LETTER, len(rows_))
        fit(ws)

    # ---- Missing pitches ---------------------------------------------------
    if missing:
        ws = sheet('_MISSING PITCHES', ['Team', 'PitchID', 'Game Date',
                                        'Pitcher', 'Batter', 'Pitch Type',
                                        'Count', 'Description'])
        rows_ = sorted(missing, key=lambda m: (m[1].get('_PTeam', ''),
                                               m[1].get('Pitcher', ''), m[0]))
        for pid, exp in rows_:
            ws.append([exp.get('_PTeam', ''), pid, exp.get('Game Date', ''),
                       exp.get('Pitcher', ''), exp.get('Batter', ''),
                       exp.get('Pitch Type', ''), exp.get('Count', ''),
                       exp.get('Description', '')])
        fit(ws)

    # ---- Blank fills, one row per PITCH -------------------------------------
    # The per-column tabs are the authority, but they over-count the decision:
    # on ARI 1,310 fills were really 290 pitches, and 113 of those were a single
    # pitch that had no tracking at all and now has all ten columns. Approving
    # that is one judgement, not ten.
    fills = [r for r in changes if r.kind == 'new']
    if fills:
        ws = sheet('_FILLS BY PITCH',
                   ['Team', 'Pitcher', 'Pitch Type', 'Game Date', 'GameID',
                    'PitchID', 'Sheet row', 'Columns gained',
                    'Max noise units',
                    'Proposed values (pitcher avg) — READ ONLY, decide on the '
                    'per-column tabs'])
        byp = collections.defaultdict(list)
        for r in fills:
            byp[(r.tab, r.pitcher, r.pitch_type, str(r.game_date),
                 r.pitch_id, r.row)].append(r)
        for (tab, pit, pt, gd, pid, row), rs in sorted(byp.items()):
            us = [r.units for r in rs if r.units is not None]
            bits = []
            for r in sorted(rs, key=lambda r: r.col):
                avg = (MEDIANS.get(r.col) or {}).get((pit, pt))
                if avg is None:
                    bits.append(f'{r.col}={r.new}')
                elif r.col in TIME_COLS:
                    bits.append(f'{r.col}={r.new} '
                                f'(avg {(int(avg) // 60) or 12}:{int(avg) % 60:02d})')
                else:
                    bits.append(f'{r.col}={r.new} '
                                f'(avg {avg:.{PRECISION.get(r.col, 3)}f})')
            ws.append([tab, pit, pt, gd, pid.split('_')[0], pid, row,
                       len(rs), round(max(us), 2) if us else '',
                       ', '.join(bits)])
        fit(ws)

    # ---- Drift auto-written, rolled up by column and month ------------------
    # Wally's call 2026-08-17: summarise these rather than list them. PlateX
    # alone runs to about 80,000 cells season-wide and its whole story is the
    # month it lands in, not the individual rows.
    auto = [r for r in changes if r.kind in ('drift_small', 'drift_sub')]
    if auto:
        ws = sheet('_AUTO-WRITTEN BY MONTH',
                   ['Column', 'Month', 'Cells', 'Median |delta|', 'Max |delta|',
                    'Max noise units'])
        buck = collections.defaultdict(list)
        for r in auto:
            buck[(r.col, str(r.game_date)[:7])].append(r)
        for (c, m), rs in sorted(buck.items()):
            ds = sorted(abs(r.delta) for r in rs if r.delta is not None)
            us = [r.units for r in rs if r.units is not None]
            ws.append([c, m, len(rs),
                       round(ds[len(ds) // 2], 6) if ds else '',
                       round(ds[-1], 6) if ds else '',
                       round(max(us), 3) if us else ''])
        fit(ws)

    # ---- Zone repairs ------------------------------------------------------
    zc = [r for r in changes if r.kind == 'zone_fix']
    if zc:
        ws = sheet('_ZONE FIXED', ['Game Date', 'GameID', 'At-bat', 'Team',
                                   'Column', 'Batter', 'PitchID', 'Sheet row',
                                   'Current', 'Corrected',
                                   'Recommend', 'Reject',
                                   'Delta', 'Whose zone it was'])
        zc.sort(key=lambda r: (str(r.game_date), r.pitch_id, r.col))
        for r in zc:
            parts = r.pitch_id.split('_')
            ws.append([r.game_date, parts[0], parts[1], r.tab, r.col, r.batter,
                       r.pitch_id, r.row, r.old, r.new,
                       r.rec, None,
                       round(r.delta, 6) if r.delta is not None else '',
                       r.source])
        add_checkboxes(ws, 'L', len(zc))
        fit(ws)

    # ---- Zone outliers with no donor in their game -------------------------
    # Repaired too (Wally, 2026-08-17), but kept separate: the mechanism is not a
    # mis-attributed batter, so if one of these later proves to be a real
    # measurement the list of what was changed is right here.
    zn = [r for r in changes if r.kind == 'zone_fix_nodonor']
    if zn:
        ws = sheet('_ZONE FIXED (NO DONOR)',
                   ['Game Date', 'GameID', 'At-bat', 'Team', 'Column', 'Batter',
                    'PitchID', 'Sheet row', 'Current', 'Corrected',
                    'Recommend', 'Reject', 'Delta (in)'])
        zn.sort(key=lambda r: (str(r.game_date), r.pitch_id, r.col))
        for r in zn:
            parts = r.pitch_id.split('_')
            ws.append([r.game_date, parts[0], parts[1], r.tab, r.col, r.batter,
                       r.pitch_id, r.row, r.old, r.new, r.rec, None,
                       round(r.delta * 12, 2) if r.delta is not None else ''])
        add_checkboxes(ws, 'L', len(zn))
        fit(ws)

    if unexplained_zones:
        ws = sheet('_ZONE UNEXPLAINED',
                   ['Game Date', 'GameID', 'At-bat', 'Team', 'Batter',
                    'Has (top/bot)', 'Hitter usual', 'PitchID',
                    'SzTop off (in)', 'SzBot off (in)', 'Repaired?'])
        for u in sorted(unexplained_zones, key=lambda u: (str(u['game_date']),
                                                          u['game_pk'])):
            ws.append([u['game_date'], u['game_pk'], u['at_bat'], u['tab'],
                       u['batter'], u['has'], u['modal'], u['pid'],
                       u['sztop_off_in'], u['szbot_off_in'],
                       'yes, see _ZONE FIXED (NO DONOR)' if zn else 'no'])
        fit(ws)

    # ---- Blank fills rejected as implausible, rolled up by game -----------
    imp = [r for r in changes if r.kind in ('fill_implausible', 'blocked_source')]
    if imp:
        ws = sheet('_IMPLAUSIBLE FILLS',
                   ['Column', 'Game Date', 'GameID', 'Team', 'Why', 'Cells',
                    'Median noise units', 'Max noise units', 'Example'])
        buck = collections.defaultdict(list)
        for r in imp:
            buck[(r.col, str(r.game_date), r.pitch_id.split('_')[0], r.tab,
                  r.kind)].append(r)
        for (c, gd, gpk, tab, kind), rs in sorted(buck.items(),
                                                  key=lambda kv: -len(kv[1])):
            us = sorted(r.units for r in rs if r.units is not None)
            worst = max(rs, key=lambda r: r.units or 0)
            why = ('source bad for the whole game' if kind == 'blocked_source'
                   else f'beyond {FILL_REVIEW_SD} noise units')
            ws.append([c, gd, gpk, tab, why, len(rs),
                       round(us[len(us) // 2], 2) if us else '',
                       round(us[-1], 2) if us else '',
                       f'{worst.pitcher}: blank -> {worst.new}'])
        fit(ws)

    # ---- Values an earlier sweep already raised -----------------------------
    # These are NOT written. They are here so the record is visible, and they
    # carry the same Pitch Type, pitcher average and override box as the review
    # tabs — Wally's ask 2026-08-17, because a row cannot be judged without
    # knowing what the pitch was and what he normally does with it. Overriding
    # one here brings it back: _wanted() lets a decision beat the kind filter.
    supp = [r for r in changes if r.kind == 'suppressed']
    if supp:
        ws = sheet('_ALREADY RAISED (not written)',
                   ['Team', 'Column', 'Pitcher', 'Pitch Type', 'Game Date',
                    'GameID', 'PitchID', 'Sheet row', 'Current', 'Proposed',
                    'Recommend', 'Reject', 'Why',
                    'Pitcher avg on this type', 'Off usual', 'n for avg',
                    'First raised', 'Values already declined'])
        supp.sort(key=lambda r: (r.rec != REJECT, r.col, r.tab, r.pitcher,
                                 r.pitch_id))
        for r in supp:
            e = ledger.entries.get((r.pitch_id, r.col)) or {}
            avg = ((MEDIANS.get(r.col) or {}).get((r.pitcher, r.pitch_type))
                   if r.col in PITCH_METRIC_COLS else None)
            if avg is None:
                avg_disp, off = '', ''
            elif r.col in TIME_COLS:
                avg_disp = f'{(int(avg) // 60) or 12}:{int(avg) % 60:02d}'
                off = tilt_gap(r.new, avg_disp)
            else:
                dp = PRECISION.get(r.col, 3)
                avg_disp = round(avg, dp)
                nv = as_float(r.new)
                off = round(nv - avg, dp) if nv is not None else ''
            ws.append([r.tab, r.col, r.pitcher, r.pitch_type, r.game_date,
                       r.pitch_id.split('_')[0], r.pitch_id, r.row,
                       r.old, r.new, r.rec, None, r.rec_why,
                       avg_disp, off if off is not None else '',
                       GROUP_N.get((r.pitcher, r.pitch_type), ''),
                       e.get('asof', ''),
                       ', '.join(e.get('rejected', []))])
        add_checkboxes(ws, 'L', len(supp))
        fit(ws)

    wb.save(path)
    return path


# ── Apply ────────────────────────────────────────────────────────────────────
def pin_formats(ws, header, cols):
    """Pin NUMBER_FORMATS over each column's full data range. Idempotent."""
    reqs = []
    for name in cols:
        f = NUMBER_FORMATS.get(name)
        if f is None or name not in header:
            continue
        i = header.index(name)
        reqs.append({'repeatCell': {
            'range': {'sheetId': ws.id, 'startRowIndex': 1,
                      'startColumnIndex': i, 'endColumnIndex': i + 1},
            'cell': {'userEnteredFormat': {'numberFormat': f}},
            'fields': 'userEnteredFormat.numberFormat'}})
    if reqs:
        _retry_sheets_call(lambda: ws.spreadsheet.batch_update({'requests': reqs}),
                           'number-format pin')


def apply_changes(ws, header, changes, chunk=40000):
    """Write the accepted changes. USER_ENTERED so numbers sort in the sheet."""
    col_idx = {n: j for j, n in enumerate(header) if n}
    cells = [gspread.Cell(row=c.row, col=col_idx[c.col] + 1, value=c.new)
             for c in changes if c.col in col_idx]
    for i in range(0, len(cells), chunk):
        batch = cells[i:i + chunk]
        print(f"      writing cells {i + 1}..{i + len(batch)} of {len(cells)}",
              flush=True)
        update_cells_with_retry(ws, batch, value_input_option='USER_ENTERED')
        time.sleep(2)
    pin_formats(ws, header, {c.col for c in changes})
    return len(cells)


# ── Main ─────────────────────────────────────────────────────────────────────
def main(filter_teams=None, apply=False, refresh_feed=False, kinds=None,
         report=True, record=True, decisions_from=None):
    kinds = kinds or {'new', 'drift', 'drift_sub', 'drift_small',
                  'precision', 'zone_fix', 'zone_fix_nodonor'}
    print(f"Mode: {'APPLY' if apply else 'DRY RUN (nothing is written)'}")
    print(f"Change classes in scope: {', '.join(sorted(kinds))}")
    print(f"Columns: {len(IN_SCOPE)} ({len(FEED_COLS)} feed, "
          f"{len(SAVANT_COLS)} savant)")

    decisions = read_decisions(decisions_from) if decisions_from else None
    if apply and decisions is None:
        print("  NOTE: no --decisions-from workbook, so every reviewable change "
              "follows its own recommendation with no overrides.")

    ledger = DecisionLedger.load()
    gc = gspread.service_account()
    session = requests.Session()
    session.headers.update({
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)',
        'Accept': 'application/json, text/csv',
    })

    all_changes, all_missing, zone_obs = [], [], []
    feed_failures = []
    # tab -> (worksheet, header) so the optional modal-zone pass can write after
    # the cross-tab mode is known.
    tab_handles = {}

    for label, sheet_id in SPREADSHEET_IDS.items():
        sh = _retry_sheets_call(lambda: gc.open_by_key(sheet_id), 'workbook open')
        print(f"\n{'=' * 62}\n[{label}] {sh.title}\n{'=' * 62}")

        for i, ws in enumerate(_retry_sheets_call(sh.worksheets, 'tab list')):
            tab = ws.title.upper()
            if tab not in ALL_TRACKED_TEAMS:
                continue
            if filter_teams and tab not in filter_teams:
                continue
            print(f"\n[{ws.title}]")
            if i > 0:
                time.sleep(1.5)

            rows = read_sheet_with_retry(ws)
            if not rows or len(rows) < 2:
                print("  empty tab")
                continue
            header = rows[0]
            if 'PitchID' not in header:
                print("  no PitchID column, skipping")
                continue

            pid_c = header.index('PitchID')
            game_pks, dates = set(), set()
            d_c = header.index('Game Date') if 'Game Date' in header else None
            for row in rows[1:]:
                pid = row[pid_c] if pid_c < len(row) else ''
                if pid and '_' in pid:
                    game_pks.add(pid.split('_')[0])
                if d_c is not None and d_c < len(row) and row[d_c]:
                    dates.add(row[d_c])
            if not game_pks:
                print("  no PitchIDs, skipping")
                continue
            print(f"  {len(rows) - 1} rows, {len(game_pks)} games")

            # --- feed pass ---
            feed_by_pid = {}
            for n, gpk in enumerate(sorted(game_pks), 1):
                try:
                    feed_by_pid.update(feed_rows(gpk, session,
                                                 refresh=refresh_feed))
                except RuntimeError as e:
                    # Fail closed per game, not per tab: a game we could not
                    # read must not be reported as missing pitches.
                    print(f"    FEED FAILED {gpk}: {e}")
                    feed_failures.append((tab, gpk, str(e)))
                if n % 25 == 0:
                    print(f"    feed {n}/{len(game_pks)} games", flush=True)
            print(f"    feed gave {len(feed_by_pid)} pitches")

            # --- savant pass ---
            savant_lookup = {}
            if dates:
                time.sleep(3)
                savant_lookup = download_statcast(ws.title, min(dates),
                                                  max(dates), session) or {}

            tab_handles[tab] = (ws, header)
            changes, missing = diff_tab(tab, rows, header, feed_by_pid,
                                        savant_lookup, ledger, zone_obs)
            if feed_failures and any(f[0] == tab for f in feed_failures):
                bad = {f[1] for f in feed_failures if f[0] == tab}
                missing = [m for m in missing if m[0].split('_')[0] not in bad]
                print(f"    NOTE: {len(bad)} unreadable games excluded from the "
                      f"missing-pitch list for this tab")

            k = collections.Counter(c.kind for c in changes)
            print(f"  to read: {k['new']} fills + {k['drift']} drift | "
                  f"auto: {k['drift_small'] + k['drift_sub']} small + "
                  f"{k['precision']} precision | held: {k['suppressed']} "
                  f"suppressed + {k['fill_implausible']} implausible | "
                  f"missing {len(missing)}")

            all_changes.extend(changes)
            all_missing.extend(missing)

            if apply:
                todo = [c for c in changes if _wanted(c, decisions, kinds)]
                if todo:
                    n = apply_changes(ws, header, todo)
                    print(f"  wrote {n} cells")
                else:
                    print("  nothing to write")

    # ---- strike-zone repair, after every tab is in ----
    # Must run across all tabs at once: a hitter appears in the tab of every team
    # that pitched to him, and the modal zone needs all of those observations.
    unexplained_zones = []
    if zone_obs:
        zc, unexplained_zones = zone_outlier_changes(zone_obs)
        print(f"\nzone repair: {len(zc)} cells carry another hitter's zone and "
              f"are repaired; {len(unexplained_zones)} at-bats do not match any "
              f"hitter in their game and are reported instead")
        all_changes.extend(zc)
        if apply and zc:
            zc = [c for c in zc if _wanted(c, decisions, kinds)]
            for tab, group in collections.groupby(
                    sorted(zc, key=lambda c: c.tab), key=lambda c: c.tab):
                ws, header = tab_handles[tab]
                n = apply_changes(ws, header, list(group))
                print(f"  {tab}: wrote {n} zone cells")

    # ---- summary ----
    k = collections.Counter(c.kind for c in all_changes)
    print(f"\n{'=' * 62}")
    print(f"TO READ")
    print(f"  blank fills to approve   {k['new']}")
    print(f"  drift to approve         {k['drift']}")
    print(f"  zone at-bats unexplained {len(unexplained_zones)}")
    print(f"WRITTEN WITHOUT ASKING")
    print(f"  precision rewrites       {k['precision']}")
    print(f"  drift below precision    {k['drift_sub']}")
    print(f"  drift below 1 noise unit {k['drift_small']}")
    print(f"  strike-zone repairs      {k['zone_fix']} matched to a donor, "
          f"{k['zone_fix_nodonor']} without one")
    print(f"NOT WRITTEN")
    print(f"  suppressed by the ledger {k['suppressed']}")
    print(f"  fills judged implausible {k['fill_implausible']}")
    print(f"  blocked, bad at source   {k['blocked_source']}")
    print(f"  missing pitches found    {len(all_missing)}")
    print(f"cells this run would write "
          f"{sum(k[x] for x in AUTO_KINDS) + k['new'] + k['drift']}")
    if feed_failures:
        print(f"\nWARNING: {len(feed_failures)} games could not be read. Their "
              f"pitches were left out of the missing-pitch list rather than "
              f"reported as absent. Re-run to retry:")
        for tab, gpk, err in feed_failures[:20]:
            print(f"  {tab} {gpk}: {err}")

    if report:
        os.makedirs(REPORT_DIR, exist_ok=True)
        stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        # The team code goes in the filename. Wally's ask 2026-08-17: with one
        # workbook per tab in flight at a time, a bare timestamp gives no way to
        # tell which is which in Downloads. Uses the tabs actually walked rather
        # than the --teams argument, so a whole-season run says ALL and a run
        # whose tabs were filtered names them.
        walked = sorted({c.tab for c in all_changes} | {m[1].get('_PTeam', '')
                                                        for m in all_missing})
        walked = [t for t in walked if t]
        if not walked:
            label = 'none'
        elif len(walked) > 6:
            label = f'ALL{len(walked)}'
        else:
            label = '-'.join(walked)
        path = os.path.join(REPORT_DIR, f'backfill_{label}_{stamp}.xlsx')
        write_report(all_changes, all_missing, ledger, path,
                     unexplained_zones=unexplained_zones)
        print(f"\nReport: {path}")

    if record:
        n_new = record_sweep(all_changes, ledger)
        ledger.save()
        print(f"\nRecorded {n_new} newly surfaced values in "
              f"{os.path.relpath(LEDGER_PATH, _ROOT)} ({len(ledger.entries)} "
              f"cells tracked). A later sweep will not raise these again unless "
              f"the source value changes.")
    else:
        print(f"\n--no-record: nothing written to "
              f"{os.path.relpath(LEDGER_PATH, _ROOT)}; a later sweep will raise "
              f"these same values again.")
    return all_changes, all_missing


if __name__ == '__main__':
    ap = argparse.ArgumentParser(
        description='Reconcile every sheet column against its source.')
    ap.add_argument('--teams', default=None,
                    help='comma-separated tabs, e.g. ARI,WSH')
    ap.add_argument('--apply', action='store_true',
                    help='write to the sheets. Omit for a dry run.')
    ap.add_argument('--refresh-feed', action='store_true',
                    help='re-pull every game instead of using data/_feed_cache')
    ap.add_argument('--kinds',
                    default=('new,drift,drift_sub,drift_small,precision,'
                             'zone_fix,zone_fix_nodonor'),
                    help='which change classes --apply writes')
    ap.add_argument('--decisions-from', default=None, metavar='XLSX',
                    help='a reviewed dry-run workbook. Each row follows its '
                         'Recommend column, inverted where the Reject box is '
                         'ticked. Without this, every recommendation stands.')
    ap.add_argument('--no-record', action='store_true',
                    help='do not write what this sweep surfaced into '
                         'data/backfill_decisions.json. Use for an exploratory '
                         'run that should not stop a later sweep from raising '
                         'the same values again.')
    ap.add_argument('--no-report', action='store_true')
    a = ap.parse_args()
    main(filter_teams=[t.strip().upper() for t in a.teams.split(',')] if a.teams else None,
         apply=a.apply,
         refresh_feed=a.refresh_feed,
         kinds={k.strip() for k in a.kinds.split(',') if k.strip()},
         report=not a.no_report,
         record=not a.no_record,
         decisions_from=a.decisions_from)
